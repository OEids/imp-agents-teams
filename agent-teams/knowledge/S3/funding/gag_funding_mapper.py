#!/usr/bin/env python3
"""
GAG Funding Data Extraction and Mapping Script
===============================================

This script extracts funding data from DfE GAG funding statements and maps
the values into IMP Planner format.

Mapping is performed using SchoolCode and Description fields, as the
FinanceCode field can vary between trusts and configurations.

Supports two output modes:
1. NEW CUSTOMER - Produces Funding_Tab_Values.xlsx (for pasting into Build Workbook)
2. EXISTING CUSTOMER - Produces ScenarioYearValues_Updated.csv (for IMP Planner import)

Usage:
    1. Set CUSTOMER_TYPE to 'NEW' or 'EXISTING'
    2. Populate the pre16_data and post16_data dictionaries with values from PDFs
    3. Run: python gag_funding_mapper.py

Outputs:
    NEW:      Funding_Tab_Values.xlsx + Funding_Validation_Report.xlsx
    EXISTING: [TrustCode]_ScenarioYearValues_Updated.csv + Funding_Validation_Report.xlsx

Author: GAG Funding Project
Version: 2.2
"""

import os
import glob
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import datetime

# ============================================
# CONFIGURATION
# ============================================

# Set to 'NEW' for new customers (Excel Build Workbook)
# Set to 'EXISTING' for existing customers (IMP Planner CSV exports)
CUSTOMER_TYPE = 'NEW'

# File paths (leave as None to auto-detect)
# NEW CUSTOMER:
BUILD_WORKBOOK_PATH = None  # e.g., 'WIP_AA_New_-_Strand_3_Standard_Workbook_API_2_200.xlsx'

# EXISTING CUSTOMER:
SCHOOLS_CSV_PATH = None     # e.g., 'TrustCode_-_Schools.csv'
SCENARIO_CSV_PATH = None    # e.g., 'TrustCode_-_ScenarioYearValues.csv'

# Output settings
VALIDATION_REPORT_NAME = 'Funding_Validation_Report.xlsx'

# ============================================
# PRE-16 GAG FUNDING DATA
# ============================================
# Extract these values from the Pre-16 GAG Funding Statements (PDFs)
# Key = SchoolCode from the Schools tab/CSV
#
# For PRIMARY schools, use fields prefixed with "Primary"
# For SECONDARY schools, use fields prefixed with "Secondary" and "KS3/KS4"
# For ALLTHROUGH schools, use BOTH Primary AND Secondary fields
# For SPECIAL schools (High Needs only), use only "High Needs Pre-16"

pre16_data = {
    # ==========================================
    # EXAMPLE - PRIMARY SCHOOL
    # ==========================================
    # "BBP": {
    #     "Total Allocation": 1088255.20,
    #     "Basic Entitlement - Primary": 798835.80,
    #     "Primary IDACI Band A": 0.00,
    #     "Primary IDACI Band B": 1040.18,
    #     "Primary IDACI Band C": 0.00,
    #     "Primary IDACI Band D": 890.16,
    #     "Primary IDACI Band E": 13682.40,
    #     "Primary IDACI Band F": 4935.84,
    #     "Primary FSM": 17327.80,
    #     "Primary FSM6": 38166.48,
    #     "Primary Low Prior Attainment": 67557.59,
    #     "Primary EAL": 694.28,
    #     "Primary Mobility": 0.00,
    #     "London Fringe": 0.00,
    #     "Primary Lump Sum": 145124.67,
    #     "PFI": 0.00,
    #     "Split Sites": 0.00,
    #     "Sparsity": 0.00,
    #     "Min Per Pupil": 0.00,
    #     "MFG": 0.00,
    #     "Adjustment": 0.00,
    #     "High Needs Pre-16": 0.00,
    # },
    
    # ==========================================
    # EXAMPLE - SECONDARY SCHOOL
    # ==========================================
    # "BBS": {
    #     "Total Allocation": 9392633.13,
    #     "Basic Entitlement - KS3": 4310557.56,
    #     "Basic Entitlement - KS4": 3264118.34,
    #     "Secondary IDACI Band A": 84564.24,
    #     "Secondary IDACI Band B": 73022.74,
    #     "Secondary IDACI Band C": 2085.36,
    #     "Secondary IDACI Band D": 89550.51,
    #     "Secondary IDACI Band E": 15302.72,
    #     "Secondary IDACI Band F": 76173.44,
    #     "Secondary FSM": 185159.92,
    #     "Secondary FSM6": 589443.54,
    #     "Secondary Low Prior Attainment": 525624.69,
    #     "Secondary EAL": 31905.40,
    #     "Secondary Mobility": 0.00,
    #     "London Fringe": 0.00,
    #     "Secondary Lump Sum": 145124.67,
    #     "PFI": 0.00,
    #     "Split Sites": 0.00,
    #     "Sparsity": 0.00,
    #     "Min Per Pupil": 0.00,
    #     "MFG": 0.00,
    #     "Adjustment": 0.00,
    #     "High Needs Pre-16": 0.00,
    # },
    
    # ==========================================
    # EXAMPLE - SPECIAL SCHOOL (High Needs only)
    # ==========================================
    # "CHS": {
    #     "Total Allocation": 1000000.00,
    #     "High Needs Pre-16": 1000000.00,
    # },
}

# ============================================
# POST-16 FUNDING DATA
# ============================================
# Extract these values from the 16-19 Allocation Statements (PDFs)
# Only applicable to schools with sixth forms

post16_data = {
    # ==========================================
    # EXAMPLE - SECONDARY WITH SIXTH FORM
    # ==========================================
    # "BBS": {
    #     "Total Allocation": 760565.00,
    #     "Core Programme Funding": 707849.00,
    #     "Condition of Funding Adjustment": 0.00,
    #     "Advanced Maths Premium": 0.00,
    #     "Core Maths Premium": 5400.00,
    #     "High Value Courses Premium": 16200.00,
    #     "High Needs Post-16": 18000.00,
    #     "Student Financial Support": 13115.00,
    # },
    
    # ==========================================
    # EXAMPLE - SPECIAL SCHOOL POST-16
    # ==========================================
    # "CHS": {
    #     "Total Allocation": 244204.00,
    #     "Core Programme Funding": 0.00,
    #     "Condition of Funding Adjustment": 0.00,
    #     "Advanced Maths Premium": 0.00,
    #     "Core Maths Premium": 0.00,
    #     "High Value Courses Premium": 0.00,
    #     "High Needs Post-16": 240000.00,
    #     "Student Financial Support": 4204.00,
    # },
}

# ============================================
# FIELD MAPPING DEFINITIONS
# ============================================
# These map IMP Planner Description fields to the funding data field names.
# Mapping is done by SchoolCode + Description (FinanceCode is ignored as it
# can vary between trusts).

# Pre-16 Primary school mapping
primary_mapping = {
    "01 - GAG AWPU/Basic Entitlement Primary": "Basic Entitlement - Primary",
    "02 - GAG Deprivation Primary IDACI Band A": "Primary IDACI Band A",
    "03 - GAG Deprivation Primary IDACI Band B": "Primary IDACI Band B",
    "04 - GAG Deprivation Primary IDACI Band C": "Primary IDACI Band C",
    "05 - GAG Deprivation Primary IDACI Band D": "Primary IDACI Band D",
    "06 - GAG Deprivation Primary IDACI Band E": "Primary IDACI Band E",
    "07 - GAG Deprivation Primary IDACI Band F": "Primary IDACI Band F",
    "08 - GAG Deprivation Primary FSM": "Primary FSM",
    "09 - GAG Deprivation Primary FSM6": "Primary FSM6",
    "10 - GAG Prior Attainment Primary": "Primary Low Prior Attainment",
    "11 - GAG EAL Primary": "Primary EAL",
    "12 - GAG Mobility Primary": "Primary Mobility",
    "13 - GAG London Fringe": "London Fringe",
    "14 - GAG Lump Sum Primary": "Primary Lump Sum",
    "15 - GAG PFI": "PFI",
    "16 - GAG Split Sites": "Split Sites",
    "17 - GAG Sparsity": "Sparsity",
    "18 - GAG Minimum per pupil Funding Level": "Min Per Pupil",
    "19 - GAG Minimum Funding Guarantee": "MFG",
    "20 - GAG Funding statement adjustment": "Adjustment",
}

# Pre-16 Secondary school mapping
secondary_mapping = {
    "01 - GAG AWPU/Basic Entitlement KS3": "Basic Entitlement - KS3",
    "02 - GAG AWPU/Basic Entitlement KS4": "Basic Entitlement - KS4",
    "03 - GAG Deprivation Secondary IDACI Band A": "Secondary IDACI Band A",
    "04 - GAG Deprivation Secondary IDACI Band B": "Secondary IDACI Band B",
    "05 - GAG Deprivation Secondary IDACI Band C": "Secondary IDACI Band C",
    "06 - GAG Deprivation Secondary IDACI Band D": "Secondary IDACI Band D",
    "07 - GAG Deprivation Secondary IDACI Band E": "Secondary IDACI Band E",
    "08 - GAG Deprivation Secondary IDACI Band F": "Secondary IDACI Band F",
    "09 - GAG Deprivation Secondary FSM": "Secondary FSM",
    "10 - GAG Deprivation Secondary FSM6": "Secondary FSM6",
    "11 - GAG Prior Attainment Secondary": "Secondary Low Prior Attainment",
    "12 - GAG EAL Secondary": "Secondary EAL",
    "13 - GAG Mobility Secondary": "Secondary Mobility",
    "14 - GAG London Fringe": "London Fringe",
    "15 - GAG Lump Sum Secondary": "Secondary Lump Sum",
    "16 - GAG PFI": "PFI",
    "17 - GAG Split Sites": "Split Sites",
    "18 - GAG Sparsity": "Sparsity",
    "19 - GAG Minimum per pupil Funding Level": "Min Per Pupil",
    "20 - GAG Minimum Funding Guarantee": "MFG",
    "21 - GAG Funding statement adjustment": "Adjustment",
}

# Pre-16 All-Through school mapping
allthrough_mapping = {
    "01 - GAG AWPU/Basic Entitlement Primary": "Basic Entitlement - Primary",
    "02 - GAG AWPU/Basic Entitlement KS3": "Basic Entitlement - KS3",
    "03 - GAG AWPU/Basic Entitlement KS4": "Basic Entitlement - KS4",
    "04 - GAG Deprivation Primary IDACI Band A": "Primary IDACI Band A",
    "05 - GAG Deprivation Primary IDACI Band B": "Primary IDACI Band B",
    "06 - GAG Deprivation Primary IDACI Band C": "Primary IDACI Band C",
    "07 - GAG Deprivation Primary IDACI Band D": "Primary IDACI Band D",
    "08 - GAG Deprivation Primary IDACI Band E": "Primary IDACI Band E",
    "09 - GAG Deprivation Primary IDACI Band F": "Primary IDACI Band F",
    "10 - GAG Deprivation Primary FSM": "Primary FSM",
    "11 - GAG Deprivation Primary FSM6": "Primary FSM6",
    "12 - GAG Deprivation Secondary IDACI Band A": "Secondary IDACI Band A",
    "13 - GAG Deprivation Secondary IDACI Band B": "Secondary IDACI Band B",
    "14 - GAG Deprivation Secondary IDACI Band C": "Secondary IDACI Band C",
    "15 - GAG Deprivation Secondary IDACI Band D": "Secondary IDACI Band D",
    "16 - GAG Deprivation Secondary IDACI Band E": "Secondary IDACI Band E",
    "17 - GAG Deprivation Secondary IDACI Band F": "Secondary IDACI Band F",
    "18 - GAG Deprivation Secondary FSM": "Secondary FSM",
    "19 - GAG Deprivation Secondary FSM6": "Secondary FSM6",
    "20 - GAG Prior Attainment Primary": "Primary Low Prior Attainment",
    "21 - GAG Prior Attainment Secondary": "Secondary Low Prior Attainment",
    "22 - GAG EAL Primary": "Primary EAL",
    "23 - GAG EAL Secondary": "Secondary EAL",
    "24 - GAG Mobility Primary": "Primary Mobility",
    "25 - GAG Mobility Secondary": "Secondary Mobility",
    "26 - GAG London Fringe": "London Fringe",
    "27 - GAG Lump Sum Primary": "Primary Lump Sum",
    "28 - GAG Lump Sum Secondary": "Secondary Lump Sum",
    "29 - GAG PFI": "PFI",
    "30 - GAG Split Sites": "Split Sites",
    "31 - GAG Sparsity": "Sparsity",
    "32 - GAG Minimum per pupil Funding Level": "Min Per Pupil",
    "33 - GAG Minimum Funding Guarantee": "MFG",
    "34 - GAG Funding statement adjustment": "Adjustment",
}

# Post-16 mapping
post16_mapping = {
    "01 - GAG Post 16 Core Programme Funding": "Core Programme Funding",
    "02 - GAG Post 16 Condition of funding Adjustment": "Condition of Funding Adjustment",
    "03 - GAG Post 16 Advanced Maths premium": "Advanced Maths Premium",
    "04 - GAG Post 16 Core maths premium": "Core Maths Premium",
    "05- GAG Post 16 High value courses premium": "High Value Courses Premium",
    "06 - GAG Post 16 Student financial support": "Student Financial Support",
}

# Build combined sets of all known Pre-16 GAG descriptions (for validation)
ALL_PRE16_DESCRIPTIONS = set()
ALL_PRE16_DESCRIPTIONS.update(primary_mapping.keys())
ALL_PRE16_DESCRIPTIONS.update(secondary_mapping.keys())
ALL_PRE16_DESCRIPTIONS.update(allthrough_mapping.keys())

# All known Post-16 descriptions (for validation)
ALL_POST16_DESCRIPTIONS = set(post16_mapping.keys())

# High Needs and SFS description patterns
HN_PRE16_DESCRIPTION = "Pre 16 High Needs"
HN_POST16_DESCRIPTION = "Post 16 High Needs"
SFS_DESCRIPTION_PATTERN = "Student financial support"


# ============================================
# HELPER FUNCTIONS
# ============================================

def find_build_workbook():
    """Auto-detect the build workbook in the current directory."""
    patterns = [
        "*Strand*3*Workbook*.xlsx",
        "*Build*Workbook*.xlsx",
        "*_New_*.xlsx",
    ]
    for pattern in patterns:
        matches = glob.glob(pattern)
        if matches:
            for match in matches:
                if not match.startswith('~'):
                    return match
    return None


def find_csv_files():
    """Auto-detect IMP Planner CSV files in the current directory."""
    schools_file = None
    scenario_file = None
    
    schools_matches = glob.glob("*_-_Schools.csv") + glob.glob("*Schools.csv")
    if schools_matches:
        schools_file = schools_matches[0]
    
    scenario_matches = glob.glob("*_-_ScenarioYearValues.csv") + glob.glob("*ScenarioYearValues.csv")
    if scenario_matches:
        scenario_file = scenario_matches[0]
    
    return schools_file, scenario_file


def get_mapping_for_school_type(school_type):
    """Get the appropriate field mapping based on school type."""
    if school_type == 'ALLTHROUGH':
        return allthrough_mapping
    elif school_type == 'SECONDARY':
        return secondary_mapping
    elif school_type == 'PRIMARY':
        return primary_mapping
    else:
        return primary_mapping


def get_value(school_code, description, school_info):
    """Get the funding value for a school/description combination.
    
    Mapping is based on SchoolCode + Description only. The FinanceCode field
    is not used as it can vary between IMP Planner configurations.
    """
    
    if not description or not school_code:
        return None
    
    # 1. Check Pre-16 GAG mappings (based on school type)
    if school_code in pre16_data:
        school_type = school_info.get(school_code, {}).get('type', 'PRIMARY')
        mapping = get_mapping_for_school_type(school_type)
        
        if description in mapping:
            field = mapping[description]
            return pre16_data[school_code].get(field, None)
    
    # 2. Check Post-16 mappings
    if school_code in post16_data and description in post16_mapping:
        field = post16_mapping[description]
        return post16_data[school_code].get(field, None)
    
    # 3. Check Student Financial Support (description-based match)
    if school_code in post16_data:
        if SFS_DESCRIPTION_PATTERN in description:
            return post16_data[school_code].get("Student Financial Support", None)
    
    # 4. Check Pre-16 High Needs
    if description == HN_PRE16_DESCRIPTION and school_code in pre16_data:
        return pre16_data[school_code].get("High Needs Pre-16", None)
    
    # 5. Check Post-16 High Needs
    if description == HN_POST16_DESCRIPTION and school_code in post16_data:
        return post16_data[school_code].get("High Needs Post-16", None)
    
    return None


def is_pre16_gag_description(description):
    """Check if a description belongs to Pre-16 GAG funding."""
    return description in ALL_PRE16_DESCRIPTIONS


def is_post16_description(description):
    """Check if a description belongs to Post-16 funding (core, SFS, or HN)."""
    if description in ALL_POST16_DESCRIPTIONS:
        return True
    if description == HN_POST16_DESCRIPTION:
        return True
    if SFS_DESCRIPTION_PATTERN in str(description):
        return True
    return False


# ============================================
# NEW CUSTOMER FUNCTIONS
# ============================================

def load_school_info_from_excel(wb):
    """Load school information from the Schools tab in Excel workbook."""
    ws = wb['Schools']
    school_info = {}
    
    for row in range(3, ws.max_row + 1):
        school_code = ws.cell(row=row, column=11).value  # Column K
        if school_code and school_code != 'SchoolCode':
            school_info[school_code] = {
                'name': ws.cell(row=row, column=14).value,   # Column N
                'type': ws.cell(row=row, column=13).value,   # Column M
                'urn': ws.cell(row=row, column=16).value,    # Column P
            }
    
    return school_info


def process_new_customer(workbook_path):
    """Process a new customer's Excel Build Workbook.
    
    Output: Funding_Tab_Values.xlsx
    """
    
    print(f"Loading workbook: {workbook_path}")
    wb = load_workbook(workbook_path)
    
    # Load school info
    school_info = load_school_info_from_excel(wb)
    print(f"Found {len(school_info)} schools")
    
    # Process Funding tab
    ws = wb['Funding']
    
    SCHOOL_CODE_COL = 27   # AA
    DESCRIPTION_COL = 33   # AG
    YEAR_VALUE_COL = 39    # AM
    
    updates = 0
    
    for row in range(2, ws.max_row + 1):
        school_code = ws.cell(row=row, column=SCHOOL_CODE_COL).value
        description = ws.cell(row=row, column=DESCRIPTION_COL).value
        
        if school_code and description:
            value = get_value(school_code, description, school_info)
            if value is not None:
                ws.cell(row=row, column=YEAR_VALUE_COL).value = value
                updates += 1
    
    print(f"Updated {updates} cells")
    
    # Extract Funding tab values to output file
    create_funding_tab_output(ws)
    
    # Validate and return results
    return validate_totals_excel(ws, school_info)


def create_funding_tab_output(ws):
    """Create Funding_Tab_Values.xlsx from the processed worksheet.
    
    This is the primary output for NEW customers - values to paste into Build Workbook.
    """
    
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Funding"
    
    # Headers (columns Y through AM)
    headers = ['ScenarioCode', 'FinanceCode', 'SchoolCode', 'LedgerCode', 'DepartmentCode',
               'FundCode', 'MonthProfileCode', 'CalculatorCode', 'Description', 'Notes',
               'YearNotes', 'MatEditOnly', 'FinancialYearCode', 'Calculated', 'YearValue']
    
    for col, header in enumerate(headers, 1):
        ws_out.cell(row=1, column=col, value=header)
    
    # Data
    source_cols = list(range(25, 40))  # Y=25 through AM=39
    row_out = 2
    
    for row in range(2, ws.max_row + 1):
        school_code = ws.cell(row=row, column=27).value
        description = ws.cell(row=row, column=33).value
        
        if school_code and description:
            for col_out, col_src in enumerate(source_cols, 1):
                value = ws.cell(row=row, column=col_src).value
                ws_out.cell(row=row_out, column=col_out, value=value)
            row_out += 1
    
    output_path = 'Funding_Tab_Values.xlsx'
    wb_out.save(output_path)
    print(f"Output saved to: {output_path}")


def validate_totals_excel(ws, school_info):
    """Validate totals from Excel worksheet using Description-based matching."""
    
    SCHOOL_CODE_COL = 27
    DESCRIPTION_COL = 33
    YEAR_VALUE_COL = 39
    
    validation_results = []
    
    # Pre-16 validation
    for school_code in sorted(pre16_data.keys()):
        if school_code not in school_info:
            continue
            
        school_name = school_info[school_code].get('name', school_code)
        expected_total = pre16_data[school_code].get("Total Allocation", 0)
        
        gag_sum = 0
        hn_sum = 0
        
        for row in range(2, ws.max_row + 1):
            row_school = ws.cell(row=row, column=SCHOOL_CODE_COL).value
            row_desc = ws.cell(row=row, column=DESCRIPTION_COL).value
            row_val = ws.cell(row=row, column=YEAR_VALUE_COL).value
            
            if row_school == school_code and row_val and row_desc:
                try:
                    val = float(row_val)
                    if is_pre16_gag_description(row_desc):
                        gag_sum += val
                    elif row_desc == HN_PRE16_DESCRIPTION:
                        hn_sum += val
                except (ValueError, TypeError):
                    pass
        
        calculated_total = gag_sum + hn_sum
        diff = abs(expected_total - calculated_total)
        status = "PASS" if diff < 1 else "FAIL"
        
        validation_results.append({
            "School": school_name,
            "Code": school_code,
            "Type": "Pre-16 GAG",
            "Expected": expected_total,
            "Calculated": calculated_total,
            "Difference": diff,
            "Status": status
        })
    
    # Post-16 validation
    for school_code in sorted(post16_data.keys()):
        if school_code not in school_info:
            continue
            
        school_name = school_info[school_code].get('name', school_code)
        expected_total = post16_data[school_code].get("Total Allocation", 0)
        
        post16_sum = 0
        
        for row in range(2, ws.max_row + 1):
            row_school = ws.cell(row=row, column=SCHOOL_CODE_COL).value
            row_desc = ws.cell(row=row, column=DESCRIPTION_COL).value
            row_val = ws.cell(row=row, column=YEAR_VALUE_COL).value
            
            if row_school == school_code and row_val and row_desc:
                try:
                    val = float(row_val)
                    if is_post16_description(row_desc):
                        post16_sum += val
                except (ValueError, TypeError):
                    pass
        
        diff = abs(expected_total - post16_sum)
        status = "PASS" if diff < 2 else "FAIL"
        
        validation_results.append({
            "School": school_name,
            "Code": school_code,
            "Type": "Post-16",
            "Expected": expected_total,
            "Calculated": post16_sum,
            "Difference": diff,
            "Status": status
        })
    
    return validation_results


# ============================================
# EXISTING CUSTOMER FUNCTIONS
# ============================================

def load_school_info_from_csv(csv_path):
    """Load school information from IMP Planner Schools.csv."""
    df = pd.read_csv(csv_path)
    
    school_info = {}
    for _, row in df.iterrows():
        school_code = row['SchoolCode']
        school_info[school_code] = {
            "name": row.get('Title', school_code),
            "type": row.get('SchoolType', 'PRIMARY'),
            "urn": row.get('UniqueReferenceNumber', ''),
        }
    
    return school_info


def process_existing_customer(schools_csv, scenario_csv):
    """Process an existing customer's IMP Planner CSV files.
    
    Mapping is based on SchoolCode + Description. FinanceCode is preserved
    in the output but not used for matching as it varies between trusts.
    
    Output: [TrustCode]_ScenarioYearValues_Updated.csv
    """
    
    print(f"Schools file: {schools_csv}")
    print(f"Scenario file: {scenario_csv}")
    
    # Load school mapping
    school_info = load_school_info_from_csv(schools_csv)
    print(f"Found {len(school_info)} schools")
    
    # Load scenario data
    df = pd.read_csv(scenario_csv)
    
    # Ensure YearValue column is float to accept decimal values
    df['YearValue'] = pd.to_numeric(df['YearValue'], errors='coerce').astype(float)
    
    # Update YearValue column using SchoolCode + Description matching
    updates = 0
    for idx, row in df.iterrows():
        school_code = row['SchoolCode']
        description = row['Description']
        
        value = get_value(school_code, description, school_info)
        if value is not None:
            df.at[idx, 'YearValue'] = value
            updates += 1
    
    print(f"Updated {updates} rows")
    
    # Save updated CSV (primary output for existing customers)
    output_csv = scenario_csv.replace('.csv', '_Updated.csv')
    df.to_csv(output_csv, index=False)
    print(f"Output saved to: {output_csv}")
    
    # Validate and return results
    return validate_totals_csv(df, school_info)


def validate_totals_csv(df, school_info):
    """Validate totals from CSV data using Description-based matching."""
    
    validation_results = []
    
    # Pre-16 validation
    for school_code in sorted(pre16_data.keys()):
        if school_code not in school_info:
            continue
            
        school_name = school_info[school_code].get('name', school_code)
        expected_total = pre16_data[school_code].get("Total Allocation", 0)
        
        school_rows = df[df['SchoolCode'] == school_code]
        
        # Sum Pre-16 GAG rows by description
        gag_mask = school_rows['Description'].apply(lambda d: is_pre16_gag_description(d) if pd.notna(d) else False)
        gag_sum = pd.to_numeric(school_rows.loc[gag_mask, 'YearValue'], errors='coerce').fillna(0).sum()
        
        # Sum High Needs Pre-16 rows by description
        hn_mask = school_rows['Description'] == HN_PRE16_DESCRIPTION
        hn_sum = pd.to_numeric(school_rows.loc[hn_mask, 'YearValue'], errors='coerce').fillna(0).sum()
        
        calculated_total = gag_sum + hn_sum
        diff = abs(expected_total - calculated_total)
        status = "PASS" if diff < 1 else "FAIL"
        
        validation_results.append({
            "School": school_name,
            "Code": school_code,
            "Type": "Pre-16 GAG",
            "Expected": expected_total,
            "Calculated": calculated_total,
            "Difference": diff,
            "Status": status
        })
    
    # Post-16 validation
    for school_code in sorted(post16_data.keys()):
        if school_code not in school_info:
            continue
            
        school_name = school_info[school_code].get('name', school_code)
        expected_total = post16_data[school_code].get("Total Allocation", 0)
        
        school_rows = df[df['SchoolCode'] == school_code]
        
        # Sum all Post-16 rows by description
        post16_mask = school_rows['Description'].apply(lambda d: is_post16_description(d) if pd.notna(d) else False)
        post16_sum = pd.to_numeric(school_rows.loc[post16_mask, 'YearValue'], errors='coerce').fillna(0).sum()
        
        diff = abs(expected_total - post16_sum)
        status = "PASS" if diff < 2 else "FAIL"
        
        validation_results.append({
            "School": school_name,
            "Code": school_code,
            "Type": "Post-16",
            "Expected": expected_total,
            "Calculated": post16_sum,
            "Difference": diff,
            "Status": status
        })
    
    return validation_results


# ============================================
# SHARED FUNCTIONS
# ============================================

def create_validation_report(validation_results, output_path):
    """Create an Excel validation report."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Summary"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["School Name", "Code", "Type", "Expected Total", "Calculated Total", "Difference", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Data
    for row_idx, result in enumerate(validation_results, 2):
        ws.cell(row=row_idx, column=1, value=result["School"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=result["Code"]).border = thin_border
        ws.cell(row=row_idx, column=3, value=result["Type"]).border = thin_border
        
        cell4 = ws.cell(row=row_idx, column=4, value=result["Expected"])
        cell4.number_format = '\u00a3#,##0.00'
        cell4.border = thin_border
        
        cell5 = ws.cell(row=row_idx, column=5, value=result["Calculated"])
        cell5.number_format = '\u00a3#,##0.00'
        cell5.border = thin_border
        
        cell6 = ws.cell(row=row_idx, column=6, value=result["Difference"])
        cell6.number_format = '\u00a3#,##0.00'
        cell6.border = thin_border
        
        cell7 = ws.cell(row=row_idx, column=7, value=result["Status"])
        cell7.fill = pass_fill if result["Status"] == "PASS" else fail_fill
        cell7.font = Font(bold=True)
        cell7.border = thin_border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 10
    
    wb.save(output_path)
    print(f"Validation report saved to: {output_path}")


def print_validation_summary(validation_results):
    """Print validation summary to console."""
    
    print("\n" + "=" * 80)
    print("VALIDATION SUMMARY")
    print("=" * 80)
    
    for result in validation_results:
        status_icon = "\u2713" if result["Status"] == "PASS" else "\u2717"
        print(f"\n{result['School']} ({result['Code']}) - {result['Type']}")
        print(f"  Expected:   \u00a3{result['Expected']:,.2f}")
        print(f"  Calculated: \u00a3{result['Calculated']:,.2f}")
        print(f"  Status:     {status_icon} {result['Status']}")
    
    passes = len([r for r in validation_results if r["Status"] == "PASS"])
    fails = len([r for r in validation_results if r["Status"] == "FAIL"])
    
    print("\n" + "=" * 80)
    print(f"TOTAL: {passes} PASS, {fails} FAIL out of {len(validation_results)} checks")
    print("=" * 80)


# ============================================
# MAIN EXECUTION
# ============================================

def main():
    """Main execution function."""
    
    print("=" * 80)
    print("GAG FUNDING DATA MAPPING SCRIPT (IMP Planner Format)")
    print(f"Customer Type: {CUSTOMER_TYPE}")
    print("=" * 80)
    
    # Check if data has been populated
    if not pre16_data and not post16_data:
        print("\nWARNING: No funding data has been entered!")
        print("Please populate the pre16_data and/or post16_data dictionaries")
        return
    
    if CUSTOMER_TYPE == 'NEW':
        # New customer - Excel Build Workbook
        # Output: Funding_Tab_Values.xlsx
        workbook_path = BUILD_WORKBOOK_PATH or find_build_workbook()
        
        if not workbook_path or not os.path.exists(workbook_path):
            print("ERROR: Build workbook not found")
            print("Please ensure a build workbook exists or set BUILD_WORKBOOK_PATH")
            return
        
        validation_results = process_new_customer(workbook_path)
        
    elif CUSTOMER_TYPE == 'EXISTING':
        # Existing customer - IMP Planner CSV exports
        # Output: [TrustCode]_ScenarioYearValues_Updated.csv
        schools_csv = SCHOOLS_CSV_PATH
        scenario_csv = SCENARIO_CSV_PATH
        
        if not schools_csv or not scenario_csv:
            detected_schools, detected_scenario = find_csv_files()
            schools_csv = schools_csv or detected_schools
            scenario_csv = scenario_csv or detected_scenario
        
        if not schools_csv or not os.path.exists(schools_csv):
            print("ERROR: Schools CSV file not found")
            return
        
        if not scenario_csv or not os.path.exists(scenario_csv):
            print("ERROR: ScenarioYearValues CSV file not found")
            return
        
        validation_results = process_existing_customer(schools_csv, scenario_csv)
    
    else:
        print(f"ERROR: Invalid CUSTOMER_TYPE: {CUSTOMER_TYPE}")
        print("Set CUSTOMER_TYPE to 'NEW' or 'EXISTING'")
        return
    
    # Print and save validation results
    print_validation_summary(validation_results)
    create_validation_report(validation_results, VALIDATION_REPORT_NAME)
    
    print("\nProcessing complete!")


if __name__ == "__main__":
    main()
