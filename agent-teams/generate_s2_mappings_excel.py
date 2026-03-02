"""Generate S2 Data Field Mappings Excel Workbook"""
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_header_style():
    return {
        'font': Font(bold=True, color='FFFFFF', size=10),
        'fill': PatternFill(start_color='003366', end_color='003366', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def create_cell_style(alternate=False):
    fill_color = 'F5F5F5' if alternate else 'FFFFFF'
    return {
        'fill': PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid'),
        'alignment': Alignment(vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def apply_style(cell, style_dict):
    for attr, value in style_dict.items():
        setattr(cell, attr, value)

def add_table(ws, headers, data, start_row=1, col_widths=None):
    header_style = create_header_style()

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        apply_style(cell, header_style)

    # Write data
    for row_idx, row_data in enumerate(data):
        cell_style = create_cell_style(alternate=(row_idx % 2 == 1))
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=start_row + row_idx + 1, column=col_idx, value=value)
            apply_style(cell, cell_style)

    # Set column widths
    if col_widths:
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    return start_row + len(data) + 2

def generate_excel():
    wb = Workbook()

    # ========== Sheet 1: Sheet Mappings ==========
    ws = wb.active
    ws.title = "1_Sheet_Mappings"
    add_table(ws,
        ['Internal Name', 'S2 Template Sheet'],
        [
            ['PayScales', '19_PayScales'],
            ['PayScalePoints', '20_PayScalePoints'],
            ['PayScaleGrades', '22_PayScaleGrades'],
            ['PayScaleIncreasePercen', '21_PayScaleIncreasePercen'],
            ['AllowanceTypes', '16_AllowanceTypes'],
            ['AllowanceTypePoint', '17_AllowanceTypePoint'],
            ['AllowanceIncreasePercen', '18_AllowanceIncreasePercen'],
            ['Pensions', '24_Pensions'],
            ['EQWPatterns', '23_EQWPatterns'],
            ['StfRoleGroup', '26_StfRoleGroup'],
            ['StfRole', '27_StfRole'],
            ['StaffMembers', '25_StaffMembers'],
            ['ContractsTeachFTE', '28_ContractsTeachFTE'],
            ['ContractsSupportHours', '29_ContractsSupportHours'],
            ['ContractAllowances', '34_ContractAllowances'],
            ['ContractAdjustments', '33_ContractAdjustments'],
            ['Finance Codes S2', '11_Finance Codes S2'],
            ['Genders', '12_Genders'],
            ['ContractTypes', '13_ContractTypes'],
            ['LeaveTypes', '30_LeaveTypes'],
            ['AdjustmentTypes', '31_AdjustmentTypes'],
            ['StaffMemberLeaves', '32_StaffMemberLeaves'],
        ],
        col_widths=[30, 35]
    )

    # ========== Sheet 2: Column Mappings ==========
    ws = wb.create_sheet("2_Column_Mappings")
    add_table(ws,
        ['Source Variants', 'Maps To'],
        [
            ['first_name, firstname, forename, given_name, fname', 'first_name'],
            ['last_name, lastname, surname, family_name, lname', 'last_name'],
            ['job_title, jobtitle, position, role_title, post', 'job_title'],
            ['spot_salary, spot scale, spot amount', 'spot_salary'],
            ['annual_salary, salary, annual_pay, gross_salary, basic_salary', 'annual_salary'],
            ['fte, full_time_equivalent, weekly_fte', 'weekly_fte'],
            ['weekly_hours, hours_per_week, contracted_hours, hours', 'weekly_hours'],
            ['full_time_hours, ft_hours, standard_hours', 'full_time_hours'],
            ['scale_point, scalepoint, spine_point, current_point, pay_point, scp', 'current_scale_point'],
            ['pay_scale, payscale, pay_range, salary_scale, pay_grade', 'pay_scale'],
            ['pension, pension_code, pension_scheme, superannuation', 'pension_code'],
            ['start_date, startdate, commencement, hire_date', 'start_date'],
            ['end_date, enddate, termination, leaving_date', 'end_date'],
            ['contract_type, contracttype, employment_type', 'contract_type'],
            ['equated_weeks, eqw, term_weeks, working_weeks', 'equated_week_pattern'],
            ['payroll, emp_no, employee_number, staff_id, personnel_no', 'payroll_number'],
            ['role_group, staff_role_group, rolegroup, category', 'staff_role_group'],
            ['gross_salary_fc, salary_finance_code, gross_fc', 'gross_salary_finance_code'],
            ['ni_finance_code, employers_ni_fc, ni_fc', 'employers_ni_finance_code'],
            ['pension_finance_code, pension_fc, super_fc', 'pension_finance_code'],
            ['finance_code, financecode, nominal, nominal_code, account_code', 'finance_code'],
        ],
        col_widths=[55, 30]
    )

    # ========== Sheet 3: Role Group Finance Codes ==========
    ws = wb.create_sheet("3_RoleGroup_FinanceCodes")
    add_table(ws,
        ['Role Group', 'Title', 'Gross Salary', 'Employers NI', 'Pension'],
        [
            ['LST', 'Leadership Teaching', '610100', '610200', '610300'],
            ['LSN', 'Leadership Non-Teaching', '625100', '625200', '625300'],
            ['TEA', 'Teachers', '612100', '612200', '612300'],
            ['TA', 'Teaching Assistants', '615100', '615200', '615300'],
            ['ADM', 'Finance and Admin', '625100', '625200', '625300'],
            ['CAT', 'Catering Staff', '2700', '2705', '2710'],
            ['CLE', 'Cleaning Staff', '630100', '630200', '630300'],
            ['PRE', 'Site Staff', '627100', '627200', '627300'],
            ['MDS', 'Midday Supervisors', '635100', '635200', '635300'],
            ['NUR', 'Nursery Staff', '637100', '637200', '637300'],
            ['TEC', 'Technicians', '622100', '622200', '622300'],
            ['LIB', 'Librarians', '620100', '620200', '620300'],
            ['FSW', 'Family Support Workers', '640100', '640200', '640300'],
            ['COV', 'Cover Supervisors', '2720', '2725', '2730'],
            ['COM', 'Community Facilities Staff', '645100', '645200', '645300'],
            ['OTH', 'Other Staff', '650400', '650410', '650420'],
            ['EDS', 'Educational Support', '619100', '619200', '619300'],
        ],
        col_widths=[15, 30, 15, 15, 15]
    )

    # ========== Sheet 4: Job Title Mappings ==========
    ws = wb.create_sheet("4_JobTitle_Mappings")
    add_table(ws,
        ['Job Title Keywords', 'Role Group', 'Role Code', 'Category'],
        [
            # Leadership
            ['headteacher, head, principal', 'LST', 'HT', 'Teaching Leadership'],
            ['executive headteacher, exec head', 'LST', 'EHT', 'Teaching Leadership'],
            ['deputy headteacher, deputy head', 'LST', 'DHT', 'Teaching Leadership'],
            ['assistant headteacher, asst head', 'LST', 'AHT', 'Teaching Leadership'],
            # Teachers
            ['teacher, class teacher, classroom teacher', 'TEA', 'TEA', 'Teachers'],
            ['unqualified teacher, trainee teacher', 'TEA', 'UQT', 'Teachers'],
            ['upper pay scale, ups teacher', 'TEA', 'TEA_UPS', 'Teachers'],
            ['lead practitioner, advanced skills teacher, ast', 'TEA', 'LP', 'Teachers'],
            # Teaching Assistants
            ['teaching assistant, ta, classroom assistant', 'TA', 'TA', 'Teaching Assistants'],
            ['higher level teaching assistant, hlta', 'TA', 'HLTA', 'Teaching Assistants'],
            ['apprentice ta', 'TA', 'APP_TA', 'Teaching Assistants'],
            # Admin
            ['school business manager, sbm', 'ADM', 'SBM', 'Admin/Finance'],
            ['office manager, admin manager', 'ADM', 'ADM_MGR', 'Admin/Finance'],
            ['admin assistant, administrative assistant', 'ADM', 'ADM_AST', 'Admin/Finance'],
            ['finance manager, finance officer, bursar', 'ADM', 'FIN_MGR', 'Admin/Finance'],
            ['finance assistant, accounts assistant', 'ADM', 'FIN_AST', 'Admin/Finance'],
            ['hr manager, human resources manager', 'ADM', 'HR_MGR', 'Admin/Finance'],
            ['hr assistant, human resources assistant', 'ADM', 'HR_AST', 'Admin/Finance'],
            ['receptionist, reception, front desk', 'ADM', 'REC', 'Admin/Finance'],
            # Other
            ['site manager, premises manager, facilities manager', 'PRE', 'SITE_MGR', 'Premises'],
            ['caretaker, janitor', 'PRE', 'CT', 'Premises'],
            ['catering manager, kitchen manager, chef', 'CAT', 'CAT_MGR', 'Catering'],
            ['catering assistant, kitchen assistant, cook', 'CAT', 'CAT_AST', 'Catering'],
            ['cleaner, cleaning staff, domestic', 'CLE', 'CLE', 'Cleaning'],
            ['midday supervisor, lunchtime supervisor', 'MDS', 'MDS', 'Midday'],
            ['nursery manager', 'NUR', 'NUR_MGR', 'Nursery'],
            ['nursery nurse, early years practitioner', 'NUR', 'NUR_NUR', 'Nursery'],
            ['technician, ict technician, science technician', 'TEC', 'TEC', 'Technicians'],
            ['librarian, library assistant', 'LIB', 'LIB', 'Library'],
            ['family support worker, family liaison', 'FSW', 'FSW', 'Family Support'],
            ['cover supervisor', 'COV', 'COV_SUP', 'Cover'],
        ],
        col_widths=[45, 15, 15, 20]
    )

    # ========== Sheet 5: Pay Scale by Location ==========
    ws = wb.create_sheet("5_PayScale_Location")
    row = add_table(ws,
        ['Role Group', 'Type', 'EW', 'FRI', 'IL', 'OL', 'KEN'],
        [
            ['LST', 'Teaching', 'LS_EW', 'LS_FRI', 'LS_IL', 'LS_OL', '-'],
            ['TEA', 'Teaching', 'MAIN_EW', 'MAIN_FRI', 'MAIN_IL', 'MAIN_OL', '-'],
            ['ADM', 'Support', 'NJC_EW', 'NJC_FRI', 'NJC_IL', 'NJC_OL', 'KENT'],
            ['TA', 'Support', 'NJC_EW', 'NJC_FRI', 'NJC_IL', 'NJC_OL', 'KENT'],
            ['CAT', 'Support', 'NJC_EW', 'NJC_FRI', 'NJC_IL', 'NJC_OL', 'KENT'],
            ['CLE', 'Support', 'NJC_EW', 'NJC_FRI', 'NJC_IL', 'NJC_OL', 'KENT'],
            ['PRE', 'Support', 'NJC_EW', 'NJC_FRI', 'NJC_IL', 'NJC_OL', 'KENT'],
            ['MDS', 'Support', 'NJC_EW', 'NJC_FRI', 'NJC_IL', 'NJC_OL', 'KENT'],
            ['NUR', 'Support', 'NJC_EW', 'NJC_FRI', 'NJC_IL', 'NJC_OL', 'KENT'],
            ['TEC', 'Support', 'NJC_EW', 'NJC_FRI', 'NJC_IL', 'NJC_OL', 'KENT'],
            ['LIB', 'Support', 'NJC_EW', 'NJC_FRI', 'NJC_IL', 'NJC_OL', 'KENT'],
            ['FSW', 'Support', 'NJC_EW', 'NJC_FRI', 'NJC_IL', 'NJC_OL', 'KENT'],
            ['COV', 'Support', 'NJC_EW', 'NJC_FRI', 'NJC_IL', 'NJC_OL', 'KENT'],
            ['OTH', 'Support', 'NJC_EW', 'NJC_FRI', 'NJC_IL', 'NJC_OL', 'KENT'],
        ],
        col_widths=[15, 12, 15, 15, 15, 15, 12]
    )

    # ========== Sheet 6: Department Codes ==========
    ws = wb.create_sheet("6_Department_Codes")
    add_table(ws,
        ['Role Group', 'Title', 'Department Code'],
        [
            ['LST', 'Leadership Teaching', 'SAL_LST'],
            ['LSN', 'Leadership Non-Teaching', 'SAL_LST'],
            ['TEA', 'Teachers', 'SAL_TEA'],
            ['TA', 'Teaching Assistants', 'SAL_TA'],
            ['ADM', 'Admin', 'SAL_ADM'],
            ['CAT', 'Catering', 'SAL_CAT'],
            ['CLE', 'Cleaning', 'SAL_CLE'],
            ['PRE', 'Premises', 'SAL_PRE'],
            ['MDS', 'Midday', 'SAL_MDS'],
            ['NUR', 'Nursery', 'SAL_NUR'],
            ['TEC', 'Technicians', 'SAL_TEC'],
            ['LIB', 'Library', 'SAL_LIB'],
            ['FSW', 'Family Support', 'SAL_FSW'],
            ['COV', 'Cover', 'SAL_COV'],
            ['COM', 'Community', 'SAL_COM'],
            ['OTH', 'Other', 'SAL_OTH'],
            ['ESTF', 'External Staff', 'SAL_EXT'],
        ],
        col_widths=[15, 25, 20]
    )

    # ========== Sheet 7: Location & Hours ==========
    ws = wb.create_sheet("7_Location_Hours")
    add_table(ws,
        ['Code', 'Location Name', 'Teaching Hours', 'Support Hours', 'Pay Scale Suffix'],
        [
            ['EW', 'England & Wales', '32.5', '37.0', '_EW'],
            ['FRI', 'Fringe', '32.43', '37.0', '_FRI'],
            ['IL', 'Inner London', '25.0', '36.0', '_IL'],
            ['OL', 'Outer London', '27.5', '37.5', '_OL'],
            ['KEN', 'Kent', '32.5', '37.0', '_KEN'],
            ['NMW', 'National Minimum Wage', '37.0', '37.0', '_NMW'],
        ],
        col_widths=[10, 25, 18, 18, 18]
    )

    # ========== Sheet 8: Pension Schemes ==========
    ws = wb.create_sheet("8_Pension_Schemes")
    add_table(ws,
        ['Code', 'Title', 'Employer %', 'For Roles'],
        [
            ['TPS', 'Teachers Pensions Scheme', '23.6%', 'Teaching Staff'],
            ['LGPS_IMP', 'LGPS IMP', '~20%', 'Support Staff'],
            ['LGPS_BNT', 'LGPS Barnett', 'Varies', 'Support Staff'],
            ['LGPS_KEN', 'LGPS Kent', 'Varies', 'Support Staff'],
            ['LGPS_LNS', 'LGPS Lincolnshire', 'Varies', 'Support Staff'],
            ['OPTOUT', 'Opted Out', '0%', 'Any'],
            ['0%', 'No Pension', '0%', 'Any'],
        ],
        col_widths=[15, 30, 15, 20]
    )

    # ========== Sheet 9: Contract Types ==========
    ws = wb.create_sheet("9_Contract_Types")
    add_table(ws,
        ['Code', 'Title', 'Description'],
        [
            ['PERM', 'Permanent', 'Ongoing permanent contract'],
            ['FXT', 'Fixed Term', 'Contract with defined end date'],
            ['MAT', 'Maternity Cover', 'Covering maternity leave'],
            ['ZZZ', 'Standard/Not Selected', 'Default permanent contract'],
        ],
        col_widths=[15, 25, 40]
    )

    # ========== Sheet 10: Equated Week Patterns ==========
    ws = wb.create_sheet("10_EQW_Patterns")
    add_table(ws,
        ['Code', 'Title', 'Full-Time Weeks', 'Typical Usage'],
        [
            ['39WEEKSWORKED', '39 weeks worked', '52.143', 'Teachers (term-time)'],
            ['TEA_ALLYEAR', 'Teaching Staff All Year', '52.14', 'Teachers/Leadership'],
            ['SUPALLYEAR', 'Support Staff All Year', '52.14', 'Support/Admin'],
            ['SUPTTO', 'Support TTO', '52.14', 'Support/TAs'],
            ['SUPTTO+1', 'Support TTO Plus 1', '52.14', 'Support'],
            ['SUPTTO+2', 'Support TTO Plus 2', '52.14', 'Support'],
            ['MONTHLYADJUSTMENT', 'Monthly Adjustment', '3.2439', 'Adjustments'],
        ],
        col_widths=[22, 28, 18, 25]
    )

    # ========== Sheet 11: Combined Column Parsing ==========
    ws = wb.create_sheet("11_Combined_Columns")
    add_table(ws,
        ['Source Combined Column', 'Code Output', 'Title Output'],
        [
            ['Staff Member Combined', 'StaffMemberCode', 'StaffMemberName'],
            ['Staff Role Combined', 'StaffRoleCode', 'StaffRoleTitle'],
            ['Contract Type Combined', 'ContractTypeCode', 'ContractTypeTitle'],
            ['Pay Scale Combined', 'PayScaleCode', 'PayScaleTitle'],
            ['Pay Scale Grade Combined', 'PayScaleGradeCode', 'PayScaleGradeTitle'],
            ['Pay Scale Point Combined', 'PayScalePointCode', 'PayScalePointTitle'],
            ['Pension Combined', 'PensionCode', 'PensionTitle'],
            ['Equated Week Pattern Combined', 'EquatedWeekPatternCode', 'EquatedWeekPatternTitle'],
            ['Department Combined', 'DepartmentCode', 'DepartmentTitle'],
            ['Fund Combined', 'FundCode', 'FundTitle'],
            ['School Combined', 'SchoolCode', 'SchoolName'],
            ['Gender Combined', 'GenderCode', 'GenderTitle'],
            ['Activity Combined', 'ActivityCode', 'ActivityTitle'],
            ['Ledger Combined', 'LedgerCode', 'LedgerTitle'],
        ],
        col_widths=[32, 28, 28]
    )

    # ========== Sheet 12: FTE Finance Codes ==========
    ws = wb.create_sheet("12_FTE_Finance_Codes")
    add_table(ws,
        ['Role Group', 'Weekly FTE Code', 'Annual FTE Code', 'Weekly Leave Adj', 'Annual Leave Adj'],
        [
            ['LST', 'WK_FTE_LST', 'A_FTE_LST', 'WK_FTE_LEAVE_ADJ_LST', 'A_FTE_LEAVE_ADJ_LST'],
            ['LSN', 'WK_FTE_LSN', 'A_FTE_LSN', 'WK_FTE_LEAVE_ADJ_LSN', 'A_FTE_LEAVE_ADJ_LSN'],
            ['TEA', 'WK_FTE_TEA', 'A_FTE_TEA', 'WK_FTE_LEAVE_ADJ_TEA', 'A_FTE_LEAVE_ADJ_TEA'],
            ['TA', 'WK_FTE_TA', 'A_FTE_TA', 'WK_FTE_LEAVE_ADJ_TA', 'A_FTE_LEAVE_ADJ_TA'],
            ['ADM', 'WK_FTE_ADM', 'A_FTE_ADM', 'WK_FTE_LEAVE_ADJ_ADM', 'A_FTE_LEAVE_ADJ_ADM'],
            ['CLE', 'WK_FTE_CLE', 'A_FTE_CLE', 'WK_FTE_LEAVE_ADJ_CLE', 'A_FTE_LEAVE_ADJ_CLE'],
            ['PRE', 'WK_FTE_PRE', 'A_FTE_PRE', 'WK_FTE_LEAVE_ADJ_PRE', 'A_FTE_LEAVE_ADJ_PRE'],
            ['MDS', 'WK_FTE_MDS', 'A_FTE_MDS', 'WK_FTE_LEAVE_ADJ_MDS', 'A_FTE_LEAVE_ADJ_MDS'],
            ['NUR', 'WK_FTE_NUR', 'A_FTE_NUR', 'WK_FTE_LEAVE_ADJ_NUR', 'A_FTE_LEAVE_ADJ_NUR'],
            ['TEC', 'WK_FTE_TEC', 'A_FTE_TEC', 'WK_FTE_LEAVE_ADJ_TEC', 'A_FTE_LEAVE_ADJ_TEC'],
        ],
        col_widths=[12, 18, 18, 25, 25]
    )

    # ========== Sheet 13: Staff Role Groups ==========
    ws = wb.create_sheet("13_All_Role_Groups")
    add_table(ws,
        ['Code', 'Title', 'Category'],
        [
            ['TEA', 'Teachers', 'Teaching'],
            ['LST', 'Leadership Teaching', 'Teaching'],
            ['SLT', 'Senior Leadership', 'Teaching'],
            ['ADM', 'Finance and Admin', 'Support'],
            ['APP', 'Apprentice', 'Support'],
            ['ASC', 'After School Club', 'Support'],
            ['BFC', 'Breakfast Club', 'Support'],
            ['BRD_DIR', 'Board Director', 'Governance'],
            ['BRD_IND', 'Board Independent', 'Governance'],
            ['CAT', 'Catering', 'Support'],
            ['CEN', 'Central', 'Support'],
            ['CLE', 'Cleaning', 'Support'],
            ['COM', 'Community', 'Support'],
            ['COV', 'Cover Supervisors', 'Support'],
            ['CT', 'Caretaker', 'Support'],
            ['CUR', 'Curriculum', 'Support'],
            ['DEV', 'Development', 'Support'],
            ['DRI', 'Driver', 'Support'],
            ['EDS', 'Educational Support', 'Support'],
            ['EST', 'Estate', 'Support'],
            ['EXA', 'Exams', 'Support'],
            ['EXT', 'External', 'Support'],
            ['FIN', 'Finance', 'Support'],
            ['FLA', 'First Language Assistant', 'Support'],
            ['FSW', 'Family Support Worker', 'Support'],
            ['GOV', 'Governance', 'Governance'],
            ['HLTA', 'Higher Level TA', 'Support'],
            ['HR', 'Human Resources', 'Support'],
            ['INC', 'Inclusion', 'Support'],
            ['INV', 'Invigilator', 'Support'],
            ['IT', 'IT Support', 'Support'],
            ['LIB', 'Librarian', 'Support'],
            ['LM', 'Line Manager', 'Support'],
            ['LTS', 'Learning Support', 'Support'],
            ['MDS', 'Midday Supervisor', 'Support'],
            ['NUR', 'Nursery', 'Support'],
            ['OTH', 'Other', 'Support'],
            ['OTH_DIR', 'Other Director', 'Governance'],
            ['OUT', 'Outreach', 'Support'],
            ['PAS', 'Pastoral', 'Support'],
            ['PERI', 'Peripatetic', 'Support'],
            ['PRE', 'Premises', 'Support'],
            ['PS', 'Personal Support', 'Support'],
            ['SC', 'School Crossing', 'Support'],
            ['SCITT', 'SCITT', 'Teaching'],
            ['SPO', 'Sports', 'Support'],
            ['SSUP', 'Senior Support', 'Support'],
            ['STA', 'Staff', 'Support'],
            ['STEA_LT', 'Supply Teacher Long Term', 'Teaching'],
            ['STEA_ST', 'Supply Teacher Short Term', 'Teaching'],
            ['TA', 'Teaching Assistant', 'Support'],
            ['TEC', 'Technician', 'Support'],
            ['WEL', 'Welfare', 'Support'],
        ],
        col_widths=[15, 30, 15]
    )

    # ========== Sheet 14: Validation Rules ==========
    ws = wb.create_sheet("14_Validation_Rules")

    row = add_table(ws,
        ['Format Type', 'Format', 'Example'],
        [
            ['Full dates', 'YYYY-MM-DD (ISO)', '2024-09-01'],
            ['Partial dates', 'MM/DD', '01/04'],
            ['Financial years', 'YYYY/YY', '2024/25'],
            ['UK display', 'dd/mm/yy', '01/09/24'],
        ],
        col_widths=[20, 25, 20]
    )

    row = add_table(ws,
        ['Data Type', 'Decimal Places', 'Examples'],
        [
            ['FTE', '1', '0.2, 0.4, 0.5, 1.0'],
            ['Hours per week', '1-2', '37.0, 37.5, 32.43'],
            ['Full-time weeks', '3', '52.143, 52.1429'],
            ['Pay rate', '2', '50025.00, 478.51'],
            ['Percentage', '0-2', '0, 3, 1.75, 2.25'],
            ['Scale point', 'Integer', '1, 2, 3, 18, 43'],
        ],
        start_row=row,
        col_widths=[20, 25, 20]
    )

    row = add_table(ws,
        ['Field', 'Minimum', 'Maximum'],
        [
            ['WeeklyFteOrHpw (FTE)', '0.0', '1.0'],
            ['FullTimeHoursPerWeek', '0.0', '50.0'],
            ['ScalePointNumber', '1', '100'],
            ['ServiceYears', '0', '99'],
            ['Percentage fields', '0', '100'],
        ],
        start_row=row,
        col_widths=[25, 15, 15]
    )

    row = add_table(ws,
        ['Boolean Field', 'Valid Values', 'Invalid Values'],
        [
            ['All boolean fields', 'True, False', 'true, false, TRUE, FALSE, Yes, No, 1, 0'],
        ],
        start_row=row,
        col_widths=[25, 20, 40]
    )

    row = add_table(ws,
        ['Empty Value Handling', 'Rule'],
        [
            ['Missing values', "Use empty string ''"],
            ['NOT allowed', 'nan, NaN, None, null, N/A, #N/A'],
            ['DateTo', 'Can be empty for ongoing contracts'],
            ['ScenarioCode', 'Typically empty (global scenario)'],
        ],
        start_row=row,
        col_widths=[25, 45]
    )

    # ========== Sheet 15: Fund Codes ==========
    ws = wb.create_sheet("15_Fund_Codes")
    add_table(ws,
        ['Code', 'Title', 'Description'],
        [
            ['GAG', 'General Annual Grant', 'Main school funding (default)'],
            ['PP', 'Pupil Premium', 'Additional funding for disadvantaged pupils'],
            ['PE', 'PE & Sports Premium', 'Sports and physical education funding'],
            ['UIFSM', 'Universal Infant Free School Meals', 'Free meals funding'],
            ['SCA', 'School Condition Allocation', 'Building maintenance funding'],
            ['DFC', 'Devolved Formula Capital', 'Capital funding'],
        ],
        col_widths=[15, 35, 45]
    )

    # ========== Sheet 16: Teacher Pay Points ==========
    ws = wb.create_sheet("16_Teacher_Pay_Points")
    add_table(ws,
        ['Scale', 'Point', 'Salary (2024-25)'],
        [
            ['Main Pay Scale', 'M1', '31,650'],
            ['Main Pay Scale', 'M2', '33,483'],
            ['Main Pay Scale', 'M3', '35,674'],
            ['Main Pay Scale', 'M4', '37,895'],
            ['Main Pay Scale', 'M5', '40,377'],
            ['Main Pay Scale', 'M6', '43,607'],
            ['Upper Pay Scale', 'U1', '45,646'],
            ['Upper Pay Scale', 'U2', '47,340'],
            ['Upper Pay Scale', 'U3', '49,084'],
            ['Unqualified', 'UQ1', '22,637'],
            ['Unqualified', 'UQ2', '24,781'],
            ['Unqualified', 'UQ3', '26,878'],
            ['Unqualified', 'UQ4', '28,969'],
            ['Unqualified', 'UQ5', '31,077'],
            ['Unqualified', 'UQ6', '33,455'],
            ['Leadership', 'L01', '47,185'],
            ['Leadership', 'L06', '52,659'],
            ['Leadership', 'L11', '60,488'],
            ['Leadership', 'L18', '72,162'],
            ['Leadership', 'L27', '91,633'],
            ['Leadership', 'L35', '109,366'],
            ['Leadership', 'L43', '128,443'],
        ],
        col_widths=[20, 15, 20]
    )

    # Save workbook
    output_path = 'reports/S2_Data_Field_Mappings.xlsx'
    wb.save(output_path)
    return output_path

if __name__ == '__main__':
    output = generate_excel()
    print(f"Excel file generated: {output}")
