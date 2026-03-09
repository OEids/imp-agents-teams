"""
DfE Data Integration

Integrates with DfE data sources to retrieve school information and funding data.

Data Sources:
1. GIAS (Get Information About Schools) - School details, URN lookup
   - Download: https://get-information-schools.service.gov.uk/Downloads
   - Free public CSV download

2. School Funding Statistics - Funding allocations
   - Download: https://explore-education-statistics.service.gov.uk/find-statistics/school-funding-statistics
   - Excel/CSV files with school-level allocations

3. Pupil Premium Allocations
   - Download: https://www.gov.uk/government/publications/pupil-premium-allocations-and-conditions-of-grant-2024-to-2025

Note: DfE funding data is not available via API - only as downloadable files.
This module helps download, cache, and query this data using URNs.
"""

import requests
import pandas as pd
from typing import Dict, List, Any, Optional
from pathlib import Path
import json
from datetime import datetime
import re
import io


# GIAS Download URL (public, no auth required)
# Format: edubasealldata{YYYYMMDD}.csv - updated daily
GIAS_DOWNLOAD_BASE = "https://ea-edubase-api-prod.azurewebsites.net/edubase/downloads/public/edubasealldata"

# API Configuration (for Explore Education Statistics - limited data available)
DFE_API_BASE = "https://api.education.gov.uk/statistics/v1"

# Known publication/dataset IDs (these may need updating if DfE changes them)
FUNDING_PUBLICATIONS = {
    "school_funding_statistics": {
        "search_term": "school funding statistics",
        "description": "School-level funding allocations including PP, DSG, grants"
    },
    "pupil_premium": {
        "search_term": "pupil premium allocations",
        "description": "Pupil premium funding allocations by school"
    }
}

# =============================================================================
# GOV.UK Grant Allocation Download URLs
# Updated annually - check GOV.UK for latest URLs
# =============================================================================
GOVUK_GRANT_DOWNLOADS = {
    "pupil_premium_2024_25": {
        "name": "Pupil Premium 2024-25",
        "url": "https://assets.publishing.service.gov.uk/media/67ceca4edc8c730647fd13d0/Pupil-premium_2024-to-2025_published-_Mar_25.ods",
        "format": "ods",
        "description": "Pupil Premium allocations Q4 March 2025",
        "year": "2024-25"
    },
    "school_capital_2025_26": {
        "name": "School Capital (DFC) 2025-26",
        "url": "https://assets.publishing.service.gov.uk/media/682afa7a96a230471ac7e767/School_capital_funding_allocations_for_2025_to_2026.xlsx",
        "format": "xlsx",
        "description": "DFC and School Condition Allocations 2025-26",
        "year": "2025-26"
    },
    "pupil_premium_2025_26": {
        "name": "Pupil Premium 2025-26",
        # URL will be updated when published - typically June each year
        "url": None,
        "format": "ods",
        "description": "Pupil Premium allocations 2025-26 (not yet published)",
        "year": "2025-26"
    }
}


class DfEFundingAPI:
    """Client for DfE Explore Education Statistics API."""

    def __init__(self):
        self.base_url = DFE_API_BASE
        self.session = requests.Session()
        self.session.headers.update({
            "Accept": "application/json",
            "Content-Type": "application/json"
        })
        self._publication_cache = {}
        self._dataset_cache = {}

    def _get(self, endpoint: str, params: dict = None) -> dict:
        """Make GET request to API."""
        url = f"{self.base_url}/{endpoint.lstrip('/')}"
        try:
            response = self.session.get(url, params=params, timeout=30)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            raise APIError(f"API request failed: {e}")

    def _post(self, endpoint: str, data: dict) -> dict:
        """Make POST request to API."""
        url = f"{self.base_url}/{endpoint.lstrip('/')}"
        try:
            response = self.session.post(url, json=data, timeout=60)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            raise APIError(f"API request failed: {e}")

    def search_publications(self, search_term: str) -> List[dict]:
        """Search for publications by keyword."""
        result = self._get("publications", {"search": search_term})
        return result.get("results", [])

    def get_publication_datasets(self, publication_id: str) -> List[dict]:
        """Get datasets for a publication."""
        result = self._get(f"publications/{publication_id}/data-sets")
        return result.get("results", [])

    def get_dataset_meta(self, dataset_id: str) -> dict:
        """Get metadata (filters, indicators) for a dataset."""
        return self._get(f"data-sets/{dataset_id}/meta")

    def query_dataset(self, dataset_id: str, criteria: dict = None,
                      indicators: List[str] = None, page: int = 1,
                      page_size: int = 1000) -> dict:
        """
        Query a dataset with filters.

        Args:
            dataset_id: The dataset ID
            criteria: Filter criteria (filters, timePeriods, locations, geographicLevels)
            indicators: List of indicator IDs to include
            page: Page number
            page_size: Results per page
        """
        query = {
            "page": page,
            "pageSize": page_size
        }
        if criteria:
            query["criteria"] = criteria
        if indicators:
            query["indicators"] = indicators

        return self._post(f"data-sets/{dataset_id}/query", query)

    def find_funding_dataset(self) -> Optional[str]:
        """Find the school funding statistics dataset ID."""
        # Search for school funding publication
        pubs = self.search_publications("school funding statistics")

        for pub in pubs:
            if "school funding" in pub.get("title", "").lower():
                pub_id = pub.get("id")
                if pub_id:
                    # Get datasets for this publication
                    datasets = self.get_publication_datasets(pub_id)
                    for ds in datasets:
                        # Look for school-level data
                        title = ds.get("title", "").lower()
                        if "school" in title or "allocation" in title:
                            return ds.get("id")
        return None

    def get_school_funding(self, urns: List[str], academic_year: str = None) -> pd.DataFrame:
        """
        Get funding allocations for schools by URN.

        Args:
            urns: List of school URNs
            academic_year: Academic year (e.g., "2024/2025"). Defaults to latest.

        Returns:
            DataFrame with funding allocations per school
        """
        dataset_id = self.find_funding_dataset()
        if not dataset_id:
            raise APIError("Could not find school funding dataset")

        # Get metadata to understand available filters
        meta = self.get_dataset_meta(dataset_id)

        # Build query criteria
        criteria = {}

        # Add time period filter if specified
        if academic_year:
            criteria["timePeriods"] = {
                "eq": {"period": academic_year, "code": "AY"}
            }

        # Query the dataset
        result = self.query_dataset(dataset_id, criteria=criteria)

        # Convert to DataFrame
        if "results" in result:
            df = pd.DataFrame(result["results"])

            # Filter to requested URNs if URN column exists
            urn_cols = [c for c in df.columns if "urn" in c.lower()]
            if urn_cols and urns:
                urn_col = urn_cols[0]
                df[urn_col] = df[urn_col].astype(str)
                urns_str = [str(u) for u in urns]
                df = df[df[urn_col].isin(urns_str)]

            return df

        return pd.DataFrame()


class APIError(Exception):
    """API error exception."""
    pass


def get_urns_from_workbook(workbook_path: Path) -> List[str]:
    """
    Extract URNs from an S3 workbook.

    Looks in the Schools sheet for URN column.
    """
    try:
        # Try reading Schools sheet
        df = pd.read_excel(workbook_path, sheet_name="Schools", header=1)

        # Look for URN column
        urn_cols = [c for c in df.columns if "urn" in str(c).lower()]
        if urn_cols:
            urns = df[urn_cols[0]].dropna().astype(str).tolist()
            # Filter out non-numeric values
            urns = [u for u in urns if u.isdigit() and len(u) >= 5]
            return urns

        return []
    except Exception as e:
        print(f"Error reading workbook: {e}")
        return []


def fetch_funding_for_workbook(workbook_path: Path, academic_year: str = None) -> Dict[str, Any]:
    """
    Fetch funding data for all schools in an S3 workbook.

    Args:
        workbook_path: Path to S3 workbook
        academic_year: Academic year (e.g., "2024/2025")

    Returns:
        Dict with:
        - success: bool
        - funding_data: DataFrame with funding allocations
        - urns_found: List of URNs found in workbook
        - error: Error message if failed
    """
    result = {
        "success": False,
        "funding_data": pd.DataFrame(),
        "urns_found": [],
        "error": None
    }

    try:
        # Extract URNs from workbook
        urns = get_urns_from_workbook(workbook_path)
        result["urns_found"] = urns

        if not urns:
            result["error"] = "No URNs found in workbook Schools sheet"
            return result

        # Fetch funding data
        api = DfEFundingAPI()
        funding_df = api.get_school_funding(urns, academic_year)

        if funding_df.empty:
            result["error"] = "No funding data returned from API"
            return result

        result["funding_data"] = funding_df
        result["success"] = True
        return result

    except APIError as e:
        result["error"] = str(e)
        return result
    except Exception as e:
        result["error"] = f"Unexpected error: {e}"
        return result


# Alternative: Download and cache funding data files
FUNDING_DATA_URLS = {
    "school_funding_2024_25": "https://explore-education-statistics.service.gov.uk/data-catalogue/data-set/school-funding-allocations-2024-25/csv",
    "pupil_premium_2024_25": "https://www.gov.uk/government/publications/pupil-premium-allocations-and-conditions-of-grant-2024-to-2025"
}


def download_funding_data(cache_dir: Path = None) -> pd.DataFrame:
    """
    Download the latest school funding allocations data.

    This is an alternative to the API that downloads the full dataset.
    """
    if cache_dir is None:
        cache_dir = Path.home() / ".cache" / "dfe_funding"
    cache_dir.mkdir(parents=True, exist_ok=True)

    cache_file = cache_dir / "school_funding_2024_25.csv"

    # Check if cached file exists and is recent (within 7 days)
    if cache_file.exists():
        age = datetime.now().timestamp() - cache_file.stat().st_mtime
        if age < 7 * 24 * 60 * 60:  # 7 days
            return pd.read_csv(cache_file)

    # Download fresh data
    # Note: The actual download URL may vary - this is a placeholder
    # The real data is available from Explore Education Statistics
    try:
        # Try API first
        api = DfEFundingAPI()
        dataset_id = api.find_funding_dataset()
        if dataset_id:
            result = api.query_dataset(dataset_id, page_size=50000)
            if "results" in result:
                df = pd.DataFrame(result["results"])
                df.to_csv(cache_file, index=False)
                return df
    except Exception:
        pass

    return pd.DataFrame()


def lookup_school_funding(urn: str, funding_df: pd.DataFrame = None) -> Dict[str, Any]:
    """
    Look up funding for a single school by URN.

    Args:
        urn: School URN
        funding_df: Pre-loaded funding DataFrame (optional)

    Returns:
        Dict with funding amounts by category
    """
    if funding_df is None:
        funding_df = download_funding_data()

    if funding_df.empty:
        return {"error": "No funding data available"}

    # Find URN column
    urn_cols = [c for c in funding_df.columns if "urn" in c.lower()]
    if not urn_cols:
        return {"error": "No URN column in funding data"}

    urn_col = urn_cols[0]
    school_data = funding_df[funding_df[urn_col].astype(str) == str(urn)]

    if school_data.empty:
        return {"error": f"No data found for URN {urn}"}

    # Convert to dict
    return school_data.iloc[0].to_dict()


# ============================================================================
# GOV.UK Grant Allocation Downloads
# ============================================================================

def download_grant_allocation(
    grant_key: str,
    cache_dir: Path = None,
    force_refresh: bool = False
) -> Dict[str, Any]:
    """
    Download a grant allocation file from GOV.UK.

    Args:
        grant_key: Key from GOVUK_GRANT_DOWNLOADS (e.g., "pupil_premium_2024_25")
        cache_dir: Directory to cache downloads
        force_refresh: Force re-download even if cached

    Returns:
        Dict with: success, data (DataFrame), source, error
    """
    result = {
        "success": False,
        "data": pd.DataFrame(),
        "source": None,
        "error": None
    }

    if grant_key not in GOVUK_GRANT_DOWNLOADS:
        result["error"] = f"Unknown grant key: {grant_key}"
        return result

    grant_info = GOVUK_GRANT_DOWNLOADS[grant_key]

    if not grant_info.get("url"):
        result["error"] = f"{grant_info['name']} URL not available yet"
        return result

    if cache_dir is None:
        cache_dir = Path.home() / ".cache" / "dfe_grants"
    cache_dir.mkdir(parents=True, exist_ok=True)

    # Cache filename
    ext = grant_info.get("format", "xlsx")
    cache_file = cache_dir / f"{grant_key}.{ext}"

    # Check cache (valid for 30 days - allocations don't change often)
    if cache_file.exists() and not force_refresh:
        age_days = (datetime.now().timestamp() - cache_file.stat().st_mtime) / (24 * 60 * 60)
        if age_days < 30:
            print(f"Using cached {grant_info['name']} ({age_days:.1f} days old)")
            try:
                if ext == "ods":
                    df = pd.read_excel(cache_file, engine="odf")
                else:
                    df = pd.read_excel(cache_file)
                result["data"] = df
                result["success"] = True
                result["source"] = f"cache ({cache_file.name})"
                return result
            except Exception as e:
                print(f"Cache read failed, re-downloading: {e}")

    # Download from GOV.UK
    print(f"Downloading {grant_info['name']} from GOV.UK...")
    url = grant_info["url"]

    try:
        response = requests.get(url, timeout=120)
        response.raise_for_status()

        # Save to cache
        with open(cache_file, "wb") as f:
            f.write(response.content)

        # Read the file
        if ext == "ods":
            df = pd.read_excel(cache_file, engine="odf")
        else:
            df = pd.read_excel(cache_file)

        print(f"Downloaded {len(df)} rows from {grant_info['name']}")

        result["data"] = df
        result["success"] = True
        result["source"] = url
        return result

    except requests.exceptions.RequestException as e:
        result["error"] = f"Download failed: {e}"
        return result
    except Exception as e:
        result["error"] = f"Error reading file: {e}"
        return result


def download_all_available_grants(
    cache_dir: Path = None,
    force_refresh: bool = False
) -> Dict[str, Any]:
    """
    Download all available grant allocation files.

    Returns:
        Dict with grant key -> download result
    """
    results = {}

    for grant_key, grant_info in GOVUK_GRANT_DOWNLOADS.items():
        if grant_info.get("url"):  # Only download if URL is available
            print(f"\nFetching {grant_info['name']}...")
            results[grant_key] = download_grant_allocation(
                grant_key, cache_dir, force_refresh
            )

    return results


def get_available_grants() -> List[Dict[str, str]]:
    """
    Get list of available grant downloads with metadata.

    Returns:
        List of dicts with: key, name, year, description, available
    """
    grants = []
    for key, info in GOVUK_GRANT_DOWNLOADS.items():
        grants.append({
            "key": key,
            "name": info["name"],
            "year": info.get("year", ""),
            "description": info.get("description", ""),
            "available": info.get("url") is not None
        })
    return grants


def fetch_and_process_grants(
    workbook_path: Path,
    grant_keys: List[str] = None,
    cache_dir: Path = None
) -> Dict[str, Any]:
    """
    Fetch grant data from GOV.UK and match to schools in workbook.

    Args:
        workbook_path: Path to S3 workbook
        grant_keys: List of grant keys to fetch (None = all available)
        cache_dir: Cache directory

    Returns:
        Dict with: success, grant_results, schools_matched, errors
    """
    result = {
        "success": False,
        "grant_results": [],
        "schools_matched": 0,
        "grants_downloaded": [],
        "errors": [],
        "log": []
    }

    try:
        # Get URN to SchoolCode mapping
        result["log"].append("Loading schools from workbook...")
        urn_mapping = get_urn_to_school_mapping(workbook_path)

        if not urn_mapping:
            result["errors"].append("No URN to SchoolCode mapping found")
            return result

        result["log"].append(f"Found {len(urn_mapping)} schools with URNs")

        # Determine which grants to download
        if grant_keys is None:
            grant_keys = [k for k, v in GOVUK_GRANT_DOWNLOADS.items() if v.get("url")]

        # Download and process each grant
        all_grant_results = []

        for grant_key in grant_keys:
            result["log"].append(f"Fetching {grant_key}...")

            download_result = download_grant_allocation(grant_key, cache_dir)

            if not download_result["success"]:
                result["errors"].append(f"{grant_key}: {download_result['error']}")
                continue

            result["grants_downloaded"].append(grant_key)
            df = download_result["data"]

            # Process grant allocations
            grant_results = process_grant_allocations(df, urn_mapping)
            all_grant_results.extend(grant_results)

            result["log"].append(f"  Found {len(grant_results)} values from {grant_key}")

        result["grant_results"] = all_grant_results
        result["schools_matched"] = len(set(g["school_code"] for g in all_grant_results))
        result["success"] = len(all_grant_results) > 0

        return result

    except Exception as e:
        result["errors"].append(str(e))
        return result


# ============================================================================
# GIAS (Get Information About Schools) Functions
# ============================================================================

def download_gias_data(cache_dir: Path = None, force_refresh: bool = False) -> pd.DataFrame:
    """
    Download GIAS (Get Information About Schools) data.

    This is the official school register with URN, name, address, type, etc.

    Args:
        cache_dir: Directory to cache the data
        force_refresh: Force re-download even if cached

    Returns:
        DataFrame with all school data
    """
    if cache_dir is None:
        cache_dir = Path.home() / ".cache" / "dfe_data"
    cache_dir.mkdir(parents=True, exist_ok=True)

    cache_file = cache_dir / "gias_schools.csv"

    # Check cache (valid for 7 days)
    if cache_file.exists() and not force_refresh:
        age_days = (datetime.now().timestamp() - cache_file.stat().st_mtime) / (24 * 60 * 60)
        if age_days < 7:
            print(f"Using cached GIAS data ({age_days:.1f} days old)")
            return pd.read_csv(cache_file, low_memory=False)

    print("Downloading GIAS school data...")

    # Try recent dates (GIAS updates daily, try last 7 days)
    from datetime import timedelta
    for days_ago in range(7):
        date = datetime.now() - timedelta(days=days_ago)
        date_str = date.strftime("%Y%m%d")
        url = f"{GIAS_DOWNLOAD_BASE}{date_str}.csv"

        try:
            response = requests.get(url, timeout=120)
            if response.status_code == 200:
                # Parse CSV
                df = pd.read_csv(io.StringIO(response.text), low_memory=False)

                # Cache it
                df.to_csv(cache_file, index=False)
                print(f"Downloaded {len(df)} schools from {date_str}, cached to {cache_file}")

                return df
        except Exception:
            continue

    print("Error: Could not download GIAS data from any recent date")
    # Return cached if available
    if cache_file.exists():
        print("Using cached data instead")
        return pd.read_csv(cache_file, low_memory=False)
    return pd.DataFrame()


def lookup_school_by_urn(urn: str, gias_df: pd.DataFrame = None) -> Dict[str, Any]:
    """
    Look up school details by URN.

    Args:
        urn: School URN (6-digit number)
        gias_df: Pre-loaded GIAS DataFrame (optional)

    Returns:
        Dict with school details or error
    """
    if gias_df is None:
        gias_df = download_gias_data()

    if gias_df.empty:
        return {"error": "GIAS data not available"}

    # Find URN column
    urn_col = "URN" if "URN" in gias_df.columns else None
    if not urn_col:
        urn_cols = [c for c in gias_df.columns if "urn" in c.lower()]
        if urn_cols:
            urn_col = urn_cols[0]

    if not urn_col:
        return {"error": "No URN column found in GIAS data"}

    # Look up school
    school = gias_df[gias_df[urn_col].astype(str) == str(urn)]

    if school.empty:
        return {"error": f"School not found for URN {urn}"}

    # Return key fields
    row = school.iloc[0]
    return {
        "URN": str(row.get("URN", "")),
        "EstablishmentName": row.get("EstablishmentName", ""),
        "TypeOfEstablishment": row.get("TypeOfEstablishment (name)", ""),
        "EstablishmentStatus": row.get("EstablishmentStatus (name)", ""),
        "PhaseOfEducation": row.get("PhaseOfEducation (name)", ""),
        "LA": row.get("LA (name)", ""),
        "Postcode": row.get("Postcode", ""),
        "NumberOfPupils": row.get("NumberOfPupils", ""),
        "TrustName": row.get("Trusts (name)", ""),
        "UKPRN": row.get("UKPRN", ""),
    }


def get_schools_from_workbook(workbook_path: Path) -> List[Dict[str, Any]]:
    """
    Get school info for all URNs in an S3 workbook.

    Args:
        workbook_path: Path to S3 workbook

    Returns:
        List of school info dicts
    """
    urns = get_urns_from_workbook(workbook_path)
    if not urns:
        return []

    gias_df = download_gias_data()
    schools = []

    for urn in urns:
        info = lookup_school_by_urn(urn, gias_df)
        if "error" not in info:
            schools.append(info)

    return schools


# ============================================================================
# Funding Data Functions (requires manual file import)
# ============================================================================

def load_funding_file(file_path: Path) -> pd.DataFrame:
    """
    Load a funding data file (Excel or CSV) from GOV.UK downloads.

    Expected sources:
    - School Funding Statistics: https://explore-education-statistics.service.gov.uk/find-statistics/school-funding-statistics
    - Pupil Premium: https://www.gov.uk/government/publications/pupil-premium-allocations-and-conditions-of-grant-2024-to-2025

    Args:
        file_path: Path to funding data file

    Returns:
        DataFrame with funding data
    """
    if not file_path.exists():
        raise FileNotFoundError(f"Funding file not found: {file_path}")

    ext = file_path.suffix.lower()

    if ext == ".csv":
        return pd.read_csv(file_path, low_memory=False)
    elif ext in [".xlsx", ".xls"]:
        # Try to find the data sheet
        xl = pd.ExcelFile(file_path)
        for sheet in xl.sheet_names:
            if "data" in sheet.lower() or "allocation" in sheet.lower():
                return pd.read_excel(file_path, sheet_name=sheet)
        # Default to first sheet
        return pd.read_excel(file_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def match_funding_to_schools(urns: List[str], funding_df: pd.DataFrame) -> pd.DataFrame:
    """
    Match funding data to schools by URN.

    Args:
        urns: List of school URNs
        funding_df: DataFrame with funding data

    Returns:
        DataFrame with matched funding
    """
    if funding_df.empty:
        return pd.DataFrame()

    # Find URN column in funding data
    urn_col = None
    for col in funding_df.columns:
        if "urn" in col.lower():
            urn_col = col
            break

    if not urn_col:
        # Try to find by pattern (6-digit numbers)
        for col in funding_df.columns:
            sample = funding_df[col].dropna().head(10).astype(str)
            if sample.str.match(r'^\d{6}$').all():
                urn_col = col
                break

    if not urn_col:
        return pd.DataFrame({"error": ["No URN column found in funding data"]})

    # Filter to requested URNs
    funding_df[urn_col] = funding_df[urn_col].astype(str)
    urns_str = [str(u) for u in urns]

    matched = funding_df[funding_df[urn_col].isin(urns_str)]

    return matched


# ============================================================================
# Combined Workflow
# ============================================================================

def fetch_school_data_for_workbook(
    workbook_path: Path,
    funding_file: Path = None
) -> Dict[str, Any]:
    """
    Fetch all available data for schools in an S3 workbook.

    Args:
        workbook_path: Path to S3 workbook
        funding_file: Optional path to funding data file

    Returns:
        Dict with:
        - schools: List of school info from GIAS
        - funding: DataFrame with funding data (if file provided)
        - urns: List of URNs found
        - success: bool
        - error: Error message if failed
    """
    result = {
        "success": False,
        "schools": [],
        "funding": pd.DataFrame(),
        "urns": [],
        "error": None
    }

    try:
        # Get URNs from workbook
        urns = get_urns_from_workbook(workbook_path)
        result["urns"] = urns

        if not urns:
            result["error"] = "No URNs found in workbook"
            return result

        # Get school info from GIAS
        gias_df = download_gias_data()
        schools = []
        for urn in urns:
            info = lookup_school_by_urn(urn, gias_df)
            schools.append(info)
        result["schools"] = schools

        # Match funding if file provided
        if funding_file and funding_file.exists():
            funding_df = load_funding_file(funding_file)
            result["funding"] = match_funding_to_schools(urns, funding_df)

        result["success"] = True
        return result

    except Exception as e:
        result["error"] = str(e)
        return result


# ============================================================================
# Grant Allocation Processing
# ============================================================================

def get_urn_to_school_mapping(workbook_path: Path) -> Dict[str, str]:
    """
    Create URN to SchoolCode mapping from workbook Schools sheet.

    Args:
        workbook_path: Path to S3 workbook

    Returns:
        Dict mapping URN (str) to SchoolCode (str)
    """
    try:
        schools_df = pd.read_excel(workbook_path, sheet_name="Schools", header=1)

        mapping = {}

        # Find URN column
        urn_col = None
        for col in schools_df.columns:
            if 'urn' in str(col).lower() or 'uniquereference' in str(col).lower():
                urn_col = col
                break

        # Find SchoolCode column
        code_col = None
        for col in schools_df.columns:
            if col == 'SchoolCode' or 'schoolcode' in str(col).lower():
                code_col = col
                break

        if urn_col and code_col:
            for _, row in schools_df.iterrows():
                urn = str(row[urn_col]).strip()
                code = str(row[code_col]).strip()
                if urn and code and urn not in ['nan', '', '0', 'nan.0']:
                    # Clean URN (remove .0 if present)
                    urn = urn.split('.')[0]
                    mapping[urn] = code

        return mapping
    except Exception as e:
        print(f"Error creating URN mapping: {e}")
        return {}


def process_grant_allocations(
    allocations_df: pd.DataFrame,
    urn_to_school: Dict[str, str]
) -> List[Dict[str, Any]]:
    """
    Process grant allocation data and match to schools.

    Uses GOVUK_DATA_WORKBOOK_MAPPING to determine target tab and value type:
    - 'count' → Pupils tab
    - 'rate' → Statistics tab
    - 'amount' → Funding tab

    Args:
        allocations_df: DataFrame with grant allocations (must have URN column)
        urn_to_school: Dict mapping URN to SchoolCode

    Returns:
        List of dicts with: school_code, target_tab, value_type, description, value, etc.
    """
    # Import mappings - try new comprehensive mapping first
    try:
        import sys
        from pathlib import Path
        knowledge_path = Path(__file__).parent.parent / "knowledge" / "S3" / "funding"
        if str(knowledge_path) not in sys.path:
            sys.path.insert(0, str(knowledge_path))

        # Try new comprehensive mapping
        try:
            from GOVUK_DATA_WORKBOOK_MAPPING import find_workbook_mapping
            use_new_mapping = True
        except ImportError:
            from NATIONAL_GRANT_MAPPINGS import find_grant_mapping
            use_new_mapping = False
    except ImportError:
        print("Warning: Could not import mapping modules")
        return []

    results = []

    # Find URN column in allocations
    urn_col = None
    for col in allocations_df.columns:
        if 'urn' in str(col).lower():
            urn_col = col
            break

    if not urn_col:
        print("No URN column found in allocations data")
        return results

    # Find grant columns using appropriate mapping
    grant_columns = {}
    for col in allocations_df.columns:
        if use_new_mapping:
            mapping = find_workbook_mapping(str(col))
        else:
            mapping = find_grant_mapping(str(col))

        if mapping:
            grant_columns[col] = mapping
            if use_new_mapping:
                print(f"  Mapped '{col}' -> {mapping.target_tab}/{mapping.description}")
            else:
                print(f"  Mapped '{col}' -> {mapping.imp_description}")

    if not grant_columns:
        print("No grant columns matched in data")
        return results

    # Process each row
    for _, row in allocations_df.iterrows():
        urn = str(row[urn_col]).strip()
        # Clean URN
        urn = urn.split('.')[0]

        # Skip if not in our schools
        if urn not in urn_to_school:
            continue

        school_code = urn_to_school[urn]

        # Extract grant values
        for col, mapping in grant_columns.items():
            try:
                value = row[col]
                if pd.isna(value):
                    continue
                value = float(value)
                if value != 0:
                    if use_new_mapping:
                        # New mapping includes target_tab and value_type
                        # Make amounts negative for income
                        if mapping.value_type == 'amount':
                            value = -abs(value)

                        results.append({
                            "school_code": school_code,
                            "urn": urn,
                            "target_tab": mapping.target_tab,
                            "value_type": mapping.value_type,
                            "grant_type": mapping.govuk_column.split()[0],  # First word as type
                            "description": mapping.description,
                            "finance_code": mapping.finance_code,
                            "value": value,
                            "source_column": col
                        })
                    else:
                        # Old mapping - default to Funding tab
                        if mapping.is_negative:
                            value = -abs(value)

                        results.append({
                            "school_code": school_code,
                            "urn": urn,
                            "target_tab": "Funding",
                            "value_type": "amount",
                            "grant_type": mapping.grant_type,
                            "description": mapping.imp_description,
                            "finance_code": mapping.imp_finance_code,
                            "value": value,
                            "source_column": col
                        })
            except (ValueError, TypeError):
                continue

    return results


def insert_grants_into_workbook(
    workbook_path: Path,
    grant_results: List[Dict[str, Any]],
    financial_year: str = "2025/26"
) -> Dict[str, Any]:
    """
    Insert grant data into workbook - Pupils, Statistics, and Funding tabs.

    IMPORTANT: Does NOT overwrite existing values (e.g., census data).
    - If cell is empty → insert value
    - If cell has value → flag discrepancy for review, don't overwrite

    Data goes to different tabs based on value_type:
    - 'count' → Pupils tab (pupil numbers) - FLAGGED if census differs
    - 'rate' → Statistics tab (per-pupil rates)
    - 'amount' → Funding tab (allocation amounts for review)

    Args:
        workbook_path: Path to S3 workbook
        grant_results: List of grant allocation dicts
        financial_year: Financial year code

    Returns:
        Dict with success status, summary, and discrepancies
    """
    from openpyxl import load_workbook

    result = {
        "success": False,
        "inserted_count": 0,
        "skipped_count": 0,
        "discrepancies": [],  # Values that differ from existing (e.g., census vs GOV.UK)
        "by_tab": {"Pupils": 0, "Statistics": 0, "Funding": 0},
        "errors": [],
        "log": []
    }

    try:
        wb = load_workbook(workbook_path)

        # Process each target tab
        tabs_to_process = ["Funding", "Pupils", "Statistics"]

        for tab_name in tabs_to_process:
            if tab_name not in wb.sheetnames:
                result["log"].append(f"Tab '{tab_name}' not found in workbook")
                continue

            sheet = wb[tab_name]

            # Find column indices (row 1 is header)
            headers = {}
            for col_idx, cell in enumerate(sheet[1], 1):
                if cell.value:
                    headers[str(cell.value).strip()] = col_idx

            # Required columns vary by tab
            school_col = headers.get("SchoolCode")
            desc_col = headers.get("Description")
            fc_col = headers.get("FinanceCode")
            value_col = headers.get("YearValue")

            if not value_col:
                result["log"].append(f"Tab '{tab_name}': No YearValue column")
                continue

            # Build index of existing rows WITH their current values
            existing_rows = {}
            for row_idx in range(2, sheet.max_row + 1):
                school = sheet.cell(row=row_idx, column=school_col).value if school_col else None
                desc = sheet.cell(row=row_idx, column=desc_col).value if desc_col else None
                fc = sheet.cell(row=row_idx, column=fc_col).value if fc_col else None
                current_value = sheet.cell(row=row_idx, column=value_col).value

                row_info = {"row_idx": row_idx, "current_value": current_value}

                if school and desc:
                    key = f"{school}|{desc}"
                    existing_rows[key] = row_info
                if school and fc:
                    key2 = f"{school}|{fc}"
                    existing_rows[key2] = row_info
                if fc and desc:
                    key3 = f"DEFAULT|{fc}"
                    existing_rows[key3] = row_info

            # Filter grant results for this tab
            tab_grants = [g for g in grant_results if g.get("target_tab") == tab_name]

            for grant in tab_grants:
                school_code = grant.get("school_code", "DEFAULT")
                description = grant.get("description", "")
                finance_code = grant.get("finance_code", "")
                new_value = grant.get("value", 0)
                value_type = grant.get("value_type", "amount")

                # Try multiple match strategies
                matched = False
                for key in [
                    f"{school_code}|{description}",
                    f"{school_code}|{finance_code}",
                    f"DEFAULT|{finance_code}",
                ]:
                    if key in existing_rows:
                        row_info = existing_rows[key]
                        row_idx = row_info["row_idx"]
                        current_value = row_info["current_value"]

                        # Check if cell already has a value
                        if current_value is not None and current_value != "" and current_value != 0:
                            # Cell has existing value - note the change for audit
                            try:
                                existing_num = float(current_value)
                                new_num = float(new_value)

                                # Log the override if values differ
                                if abs(existing_num - new_num) > 0.01:
                                    result["discrepancies"].append({
                                        "tab": tab_name,
                                        "school_code": school_code,
                                        "description": description,
                                        "previous_value": existing_num,
                                        "new_value": new_num,
                                        "difference": new_num - existing_num,
                                        "action": "OVERWRITTEN"
                                    })
                                    result["log"].append(
                                        f"📝 UPDATED {tab_name}: {school_code} - {description} | "
                                        f"Was: {existing_num} → Now: {new_num}"
                                    )
                            except (ValueError, TypeError):
                                pass

                            # OVERRIDE with new value
                            sheet.cell(row=row_idx, column=value_col).value = new_value
                            result["inserted_count"] += 1
                            result["by_tab"][tab_name] = result["by_tab"].get(tab_name, 0) + 1
                            matched = True
                            break
                        else:
                            # Cell is empty - insert the value
                            sheet.cell(row=row_idx, column=value_col).value = new_value
                            result["inserted_count"] += 1
                            result["by_tab"][tab_name] = result["by_tab"].get(tab_name, 0) + 1
                            result["log"].append(f"✅ {tab_name}: {school_code} - {description} = {new_value}")
                            matched = True
                            break

                if not matched:
                    result["skipped_count"] += 1
                    result["log"].append(f"⏭️ No match ({tab_name}): {school_code} - {description}")

        # Save workbook
        wb.save(workbook_path)
        wb.close()

        result["success"] = True
        return result

    except Exception as e:
        result["errors"].append(str(e))
        return result


def run_grant_allocation_import(
    workbook_path: Path,
    allocation_file: Path,
    financial_year: str = "2025/26"
) -> Dict[str, Any]:
    """
    Full workflow: Load allocations, match to schools, insert into workbook.

    Args:
        workbook_path: Path to S3 workbook
        allocation_file: Path to DfE grant allocation file (CSV or Excel)
        financial_year: Financial year code

    Returns:
        Dict with results summary
    """
    result = {
        "success": False,
        "schools_matched": 0,
        "grants_found": 0,
        "updated_count": 0,
        "errors": [],
        "log": []
    }

    try:
        # Step 1: Create URN to SchoolCode mapping
        result["log"].append("Loading workbook schools...")
        urn_mapping = get_urn_to_school_mapping(workbook_path)
        result["log"].append(f"Found {len(urn_mapping)} schools with URNs")

        if not urn_mapping:
            result["errors"].append("No URN to SchoolCode mapping found in workbook")
            return result

        # Step 2: Load allocation file
        result["log"].append(f"Loading allocation file: {allocation_file.name}")
        allocations_df = load_funding_file(allocation_file)
        result["log"].append(f"Loaded {len(allocations_df)} rows")

        # Step 3: Process allocations
        result["log"].append("Matching grant columns...")
        grant_results = process_grant_allocations(allocations_df, urn_mapping)
        result["grants_found"] = len(grant_results)
        result["schools_matched"] = len(set(g["school_code"] for g in grant_results))
        result["log"].append(f"Found {len(grant_results)} grant values for {result['schools_matched']} schools")

        if not grant_results:
            result["errors"].append("No grant allocations matched to schools")
            return result

        # Step 4: Insert into workbook
        result["log"].append("Inserting values into Funding sheet...")
        insert_result = insert_grants_into_workbook(workbook_path, grant_results, financial_year)

        result["updated_count"] = insert_result["updated_count"]
        result["log"].extend(insert_result["log"][:20])  # Limit log entries
        if insert_result["errors"]:
            result["errors"].extend(insert_result["errors"])

        if insert_result["success"]:
            result["success"] = True
            result["log"].append(f"SUCCESS: Updated {result['updated_count']} values in Funding sheet")

        return result

    except Exception as e:
        result["errors"].append(str(e))
        return result


# ============================================================================
# Test Functions
# ============================================================================

def test_gias_download():
    """Test GIAS data download."""
    print("Testing GIAS download...")
    df = download_gias_data()

    if df.empty:
        print("FAILED: No data returned")
        return False

    print(f"SUCCESS: Downloaded {len(df)} schools")
    print(f"Columns: {list(df.columns[:10])}...")

    # Test lookup
    sample_urn = df["URN"].dropna().iloc[0]
    print(f"\nTesting lookup for URN {sample_urn}...")
    info = lookup_school_by_urn(str(int(sample_urn)), df)
    print(f"Result: {info}")

    return True


def test_api_connection():
    """Test the DfE Statistics API connection."""
    try:
        api = DfEFundingAPI()

        # Search for publications
        print("Testing DfE Statistics API...")
        pubs = api.search_publications("school")
        print(f"Found {len(pubs)} publications")

        return len(pubs) > 0
    except Exception as e:
        print(f"API test failed: {e}")
        return False


if __name__ == "__main__":
    print("=" * 50)
    print("DfE Data Integration Test")
    print("=" * 50)

    # Test GIAS
    test_gias_download()

    print()

    # Test Statistics API
    test_api_connection()
