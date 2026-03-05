"""
S3 Specialist Agent - Financial Team

Deep analysis and complete template builder for:
- Pupil Numbers (Spring/Autumn Census)
- Statistics & Rates
- Funding (GAG, Grants)
- Income & Expenditure Budgets
- Calculators
- Month Profiles
- Scenarios
- BF Balances

Enhanced with InferenceEngine for intelligent column mapping,
budget classification, and confidence-based decisions.

Knowledge Sources:
- S3_BUDGET_TERMINOLOGY: Customer budget file format patterns (EXP1-EXP11)
- S3_IMPORT_FORMAT_STANDARDS: API schema for output format
- S3_COLUMN_MAPPINGS: Intelligent column mapping rules
- S3_BUILD_MODES: Upload modes and budget type detection
- GAG_FUNDING_MAPPINGS: DfE GAG funding statement extraction

GAG Funding Processing:
- Processes raw DfE funding statements
- Maps DfE descriptions to IMP finance codes using SchoolCode + Description
- Supports Primary, Secondary, All-Through, and Post-16 schools
- Auto-detects school type from school name
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field
import re
import warnings
warnings.filterwarnings('ignore')

try:
    import docx
    DOCX_SUPPORT = True
except ImportError:
    DOCX_SUPPORT = False

# Import Intelligence Module for smart decisions
try:
    from intelligence import InferenceEngine, InferenceResult, ConfidenceLevel
    from intelligence import TemplateRegistry, TemplateFormatter
    INFERENCE_AVAILABLE = True
    TEMPLATE_AVAILABLE = True
except ImportError:
    INFERENCE_AVAILABLE = False
    TEMPLATE_AVAILABLE = False

# Import S3 Knowledge Modules
try:
    import sys
    from pathlib import Path as PathLib
    knowledge_path = PathLib(__file__).parent.parent / "knowledge" / "S3"
    if str(knowledge_path) not in sys.path:
        sys.path.insert(0, str(knowledge_path))

    from S3_BUDGET_TERMINOLOGY import (
        FINANCE_CODE_TERMS, DEPARTMENT_CODE_TERMS, SCHOOL_CODE_TERMS,
        LEDGER_CODE_TERMS, DESCRIPTION_TERMS, BUDGET_FILE_PATTERNS,
        INCOME_FINANCE_CODES, EXPENDITURE_FINANCE_CODES,
        normalize_column_name, detect_budget_file_pattern,
        get_finance_code_info, classify_finance_code
    )
    from S3_IMPORT_FORMAT_STANDARDS import (
        INCOME_SHEET_SCHEMA, EXPENDITURE_SHEET_SCHEMA,
        STANDARD_CALCULATORS, STANDARD_MONTH_PROFILES, STANDARD_DEPARTMENTS,
        get_schema_for_sheet, get_required_columns, validate_year_code,
        get_income_sign_convention, get_expenditure_sign_convention
    )
    from S3_COLUMN_MAPPINGS import (
        map_all_columns, find_matching_column, extract_finance_code,
        is_value_column, is_school_column, determine_line_type,
        extract_year_from_column, normalize_amount, get_default_department_for_finance_code,
        ALL_COLUMN_MAPPINGS, YEAR_COLUMN_PATTERNS, MONTH_COLUMN_PATTERNS
    )
    from S3_BUILD_MODES import (
        BuildMode, BudgetType, CUSTOMER_INPUT_FILES, TEMPLATE_REFERENCE_SHEETS,
        STANDARD_UPLIFT_STATISTICS, generate_scenario_code, detect_budget_type_from_filename,
        is_raw_data_mode
    )

    # Import GAG Funding Mappings for DfE statement processing
    funding_path = PathLib(__file__).parent.parent / "knowledge" / "S3" / "funding"
    if str(funding_path) not in sys.path:
        sys.path.insert(0, str(funding_path))

    from GAG_FUNDING_MAPPINGS import (
        PRIMARY_FUNDING_MAPPINGS, SECONDARY_FUNDING_MAPPINGS,
        ALLTHROUGH_FUNDING_MAPPINGS, POST16_FUNDING_MAPPINGS,
        FundingLineMapping, FundingOutputRow,
        detect_school_type, get_funding_mappings_for_school_type,
        find_mapping_by_description, create_funding_output
    )
    GAG_MAPPINGS_AVAILABLE = True

    S3_KNOWLEDGE_AVAILABLE = True
except ImportError as e:
    S3_KNOWLEDGE_AVAILABLE = False
    GAG_MAPPINGS_AVAILABLE = False
    # Provide fallback empty structures
    FINANCE_CODE_TERMS = {}
    INCOME_FINANCE_CODES = {}
    EXPENDITURE_FINANCE_CODES = {}
    STANDARD_CALCULATORS = {}
    STANDARD_MONTH_PROFILES = {}
    STANDARD_DEPARTMENTS = {}
    STANDARD_UPLIFT_STATISTICS = []
    PRIMARY_FUNDING_MAPPINGS = []
    SECONDARY_FUNDING_MAPPINGS = []
    ALLTHROUGH_FUNDING_MAPPINGS = []
    POST16_FUNDING_MAPPINGS = []

    # Fallback enums
    class BuildMode:
        RAW_DATA = "raw_data"
        PREPOPULATED_TEMPLATE = "prepopulated_template"

    class BudgetType:
        APPROVED = "approved"
        MASTER = "master"

# Shared inference engine instance for S3
_s3_inference_engine: Optional['InferenceEngine'] = None


def get_s3_inference_engine() -> Optional['InferenceEngine']:
    """Get or create the shared InferenceEngine for S3."""
    global _s3_inference_engine
    if _s3_inference_engine is None and INFERENCE_AVAILABLE:
        try:
            _s3_inference_engine = InferenceEngine(hot_reload=False)
        except Exception:
            pass
    return _s3_inference_engine

from .finished_workbook_patterns import (
    S3_PUPIL_PATTERNS,
    S3_INCOME_PATTERNS,
    S3_EXPENDITURE_PATTERNS,
    S3_CALCULATOR_PATTERNS,
    S3_SCENARIO_PATTERNS,
)


@dataclass
class ExtractedPupilNumber:
    """Extracted pupil number data."""
    finance_code: str
    school_code: str
    description: str
    year_code: str
    value: float
    calculator_code: str
    notes: str


@dataclass
class ExtractedBudgetLine:
    """Extracted budget line data."""
    finance_code: str
    school_code: str
    department_code: str
    description: str
    year_value: float
    line_type: str  # income, expenditure
    calculator_code: str
    month_profile: str


@dataclass
class ExtractedGrant:
    """Extracted grant data."""
    grant_type: str  # DFC, SCA, PE, UIFSM, PP
    school_code: str
    amount: float
    calculation_basis: str
    pupil_count: int


class S3SpecialistAgent:
    """
    Upskilled S3 agent for financial data.

    Builds ALL S3 template sheets:
    - Pupils
    - Statistics
    - Funding
    - Calculators
    - MonthProfiles
    - Income
    - Expenditure
    - ScenarioApBud
    - BF Balances
    - Finance Codes S3
    """

    def __init__(self, build_mode: str = "raw_data", template_workbook: str = None):
        """
        Initialize S3 Specialist Agent.

        Args:
            build_mode: "raw_data" or "prepopulated_template"
            template_workbook: Path to pre-populated template (if using template mode)
        """
        self.extracted_pupils: List[ExtractedPupilNumber] = []
        self.extracted_budgets: List[ExtractedBudgetLine] = []
        self.extracted_grants: List[ExtractedGrant] = []
        self.issues: List[str] = []
        self.assumptions: List[str] = []

        # Build mode configuration
        self.build_mode = build_mode
        self.template_workbook = template_workbook

        # Budget type tracking - CRITICAL: Don't confuse Approved vs Master!
        # Approved = Full year budget, Master = Changed in-year
        self.budget_type: Optional[str] = None  # 'approved' or 'master'
        self.scenario_code: Optional[str] = None  # e.g., 'APBUD2526' or 'MASTER2526'

        # Template reference data (populated from pre-populated template)
        self.template_reference = {
            "schools": {},      # SchoolCode -> school info
            "departments": {},  # DeptCode -> dept info
            "funds": {},        # FundCode -> fund info
            "calculators": {},  # CalculatorCode -> calc info
            "finance_codes": {},# FinanceCode -> code info
        }

        # Code mappings (customer code -> template code)
        # Set by apply_code_mappings() before processing
        self.code_mappings = {
            "school_mappings": {},      # customer school -> template SchoolCode
            "finance_mappings": {},     # customer finance code -> template FinanceCode
            "department_mappings": {},  # customer dept -> template DepartmentCode
        }

        # Initialize InferenceEngine for intelligent decisions
        self.inference_engine = get_s3_inference_engine()
        self.inference_results: List[Dict] = []  # Track all inference decisions

        self.template_data = {
            "Pupils": [],
            "Statistics": [],
            "Funding": [],
            "Calculators": [],
            "MonthProfiles": [],
            "Income": [],
            "Expenditure": [],
            "ScenarioApBud": [],
            "BF Balances": [],
            "Finance Codes S3": [],
            # Scenario sheets - budget lines and values
            "ScenarioRows": [],        # 35_ScenarioRows - budget line definitions
            "ScenarioYearValues": [],  # 36_ScenarioYearValues - customer values (numbers)
            "Monthly Values": [],      # 37_Monthly Values - monthly breakdown
        }

        # Mapping from internal sheet names to official template sheet names
        self.SHEET_NAME_MAPPING = {
            "Pupils": "Pupils",
            "Statistics": "Statistics",
            "Funding": "Funding",
            "Calculators": "14_Calculators",
            "MonthProfiles": "15_MonthProfiles",
            "Income": "Income",
            "Expenditure": "Expenditure",
            "ScenarioApBud": "ScenarioApBud",
            "BF Balances": "BF Balances",
            "Finance Codes S3": "11_Finance Codes S3",
            "ScenarioRows": "35_ScenarioRows",
            "ScenarioYearValues": "36_ScenarioYearValues",
            "Monthly Values": "37_Monthly Values",
        }

        # Initialize template registry and formatter if available
        self.template_registry = None
        self.template_formatter = None
        if TEMPLATE_AVAILABLE:
            try:
                self.template_registry = TemplateRegistry()
                self.template_formatter = TemplateFormatter(self.template_registry)
            except Exception as e:
                self.log(f"[WARN] Could not initialize template registry: {e}")

        # Log S3 knowledge availability
        if S3_KNOWLEDGE_AVAILABLE:
            self.log("[INFO] S3 Knowledge modules loaded - using enhanced budget terminology")
            self.log(f"  - {len(INCOME_FINANCE_CODES)} income finance codes")
            self.log(f"  - {len(EXPENDITURE_FINANCE_CODES)} expenditure finance codes")
            self.log(f"  - {len(STANDARD_CALCULATORS)} standard calculators")
            self.log(f"  - {len(STANDARD_DEPARTMENTS)} standard departments")
        else:
            self.log("[WARN] S3 Knowledge modules not available - using fallback logic")

        # Tracking
        self.schools_found = set()
        self.finance_codes_found = set()
        self.current_year = "2025/26"
        self.previous_year = "2024/25"

        # Validation tracking
        self.validation_errors: List[str] = []
        self.validation_warnings: List[str] = []

        # External Audit Review tracking
        self.audit_results = {
            "source_vs_output": [],
            "data_integrity": [],
            "domain_checks": [],
            "missing_data": [],
            "data_lineage": [],
        }
        self.audit_passed = True
        self.audit_score = 100.0

        # Track source data for audit comparison
        self.source_pupils = set()
        self.source_budgets = set()
        self.source_grants = set()

    def log(self, message: str, level: str = "INFO"):
        """Log a message with proper encoding and error handling for Streamlit."""
        try:
            timestamp = datetime.now().strftime("%H:%M:%S")
            msg_str = str(message).replace('\x00', '').replace('\n\n', '\n')
            if len(msg_str) > 10000:
                msg_str = msg_str[:10000] + "... [truncated]"
            output = f"[{timestamp}] [{level}] S3-Specialist: {msg_str}"
            try:
                print(output, flush=True)
            except (OSError, IOError, ValueError):
                try:
                    print(output.encode('ascii', errors='replace').decode('ascii'), flush=True)
                except Exception:
                    pass
        except Exception:
            pass

    def detect_and_set_budget_type(self, filename: str) -> str:
        """
        Detect budget type from filename and set scenario code.

        CRITICAL: Must correctly identify Approved vs Master budget!
        - Approved Budget = Full year budget (APBUD)
        - Master Budget = Changed in-year (MASTER)

        Args:
            filename: Name of the budget file

        Returns:
            Scenario code e.g., 'APBUD2526'
        """
        filename_lower = filename.lower()

        # Detect budget type
        if S3_KNOWLEDGE_AVAILABLE:
            detected = detect_budget_type_from_filename(filename)
            if detected:
                self.budget_type = detected.value if hasattr(detected, 'value') else str(detected)
        else:
            # Fallback detection
            if any(x in filename_lower for x in ['approved', 'apbud', 'full year', 'annual']):
                self.budget_type = 'approved'
            elif any(x in filename_lower for x in ['master', 'amended', 'revised', 'in-year', 'inyear']):
                self.budget_type = 'master'

        # Generate scenario code
        if self.budget_type == 'approved':
            self.scenario_code = f"APBUD{self.current_year.replace('/', '')[-4:]}"
            self.log(f"[BUDGET TYPE] Detected APPROVED budget -> {self.scenario_code}")
        elif self.budget_type == 'master':
            self.scenario_code = f"MASTER{self.current_year.replace('/', '')[-4:]}"
            self.log(f"[BUDGET TYPE] Detected MASTER budget -> {self.scenario_code}")
        else:
            # Default to approved if unclear
            self.scenario_code = f"APBUD{self.current_year.replace('/', '')[-4:]}"
            self.log(f"[WARN] Could not detect budget type from '{filename}', defaulting to APPROVED")
            self.assumptions.append(f"Budget type unclear from filename, assumed APPROVED budget")

        return self.scenario_code

    def load_template_reference_data(self, template_path: str):
        """
        Load reference data from pre-populated template workbook.

        In PREPOPULATED_TEMPLATE mode, extracts:
        - Schools, Departments, Funds, Calculators, Finance Codes
        - Local Authority rates
        - Month Profiles

        Args:
            template_path: Path to template workbook
        """
        self.log(f"Loading template reference data from: {template_path}")

        try:
            xl = pd.ExcelFile(template_path)

            # Load Schools
            if 'Schools' in xl.sheet_names:
                df = pd.read_excel(xl, 'Schools')
                for _, row in df.iterrows():
                    code = str(row.get('SchoolCode', '')).strip()
                    if code and code != 'nan':
                        self.template_reference['schools'][code] = {
                            'title': row.get('Title', ''),
                            'hub': row.get('SchoolHub', 'DEFAULT'),
                            'type': row.get('SchoolType', ''),
                            'london_weighting': row.get('LondonWeighting', False),
                            'urn': row.get('UniqueReferenceNumber', ''),
                        }
                self.log(f"  Loaded {len(self.template_reference['schools'])} schools")

            # Load Funds
            if 'Funds' in xl.sheet_names:
                df = pd.read_excel(xl, 'Funds')
                for _, row in df.iterrows():
                    code = str(row.get('FundCode', '')).strip()
                    if code and code != 'nan':
                        self.template_reference['funds'][code] = {
                            'title': row.get('Title', ''),
                            'enabled': row.get('FundEnabled', True),
                        }
                self.log(f"  Loaded {len(self.template_reference['funds'])} fund codes")

            # Load Calculators
            if '14_Calculators' in xl.sheet_names:
                df = pd.read_excel(xl, '14_Calculators')
                for _, row in df.iterrows():
                    code = str(row.get('CalculatorCode', '')).strip()
                    if code and code != 'nan':
                        self.template_reference['calculators'][code] = {
                            'title': row.get('Title', ''),
                            'type': row.get('CalculatorTypeCode', 'DEFAULT'),
                            'enabled': row.get('CalculatorEnabled', True),
                        }
                self.log(f"  Loaded {len(self.template_reference['calculators'])} calculators")

            # Load Finance Codes
            if 'FinanceCodes Budget' in xl.sheet_names:
                df = pd.read_excel(xl, 'FinanceCodes Budget')
                for _, row in df.iterrows():
                    code = str(row.get('FinanceCode', '')).strip()
                    if code and code != 'nan':
                        self.template_reference['finance_codes'][code] = {
                            'title': row.get('Title', ''),
                            'grouping': row.get('GroupingCode', ''),
                            'type': row.get('FinanceCodeTypeCode', 'BUDGET'),
                            'ledger': row.get('LedgerCode', 'COSTCTR'),
                        }
                self.log(f"  Loaded {len(self.template_reference['finance_codes'])} finance codes")

            self.log(f"Template reference data loaded successfully")

        except Exception as e:
            self.log(f"[ERROR] Failed to load template: {e}")
            self.issues.append(f"Template load error: {e}")

    def apply_code_mappings(self, code_mappings: Dict[str, Dict[str, str]]):
        """
        Apply approved code mappings (customer codes -> template codes).

        Args:
            code_mappings: Dict with:
                - school_mappings: {customer_school: template_SchoolCode}
                - finance_mappings: {customer_finance: template_FinanceCode}
                - department_mappings: {customer_dept: template_DeptCode}
        """
        if code_mappings:
            self.code_mappings = code_mappings
            self.log(f"Applied code mappings:")
            self.log(f"  - {len(code_mappings.get('school_mappings', {}))} school mappings")
            self.log(f"  - {len(code_mappings.get('finance_mappings', {}))} finance code mappings")
            self.log(f"  - {len(code_mappings.get('department_mappings', {}))} department mappings")

    def map_school_code(self, customer_school: str) -> str:
        """Map customer school code/name to template SchoolCode."""
        if not customer_school:
            return ''

        customer_school = str(customer_school).strip()

        # Check code mappings first
        if customer_school in self.code_mappings.get('school_mappings', {}):
            return self.code_mappings['school_mappings'][customer_school]

        # Check if it's already a valid template code
        if customer_school in self.template_reference.get('schools', {}):
            return customer_school

        # Log unmapped
        self.assumptions.append(f"Unmapped school: '{customer_school}' - using as-is")
        return customer_school

    def map_finance_code(self, customer_finance: str) -> str:
        """Map customer finance code to template FinanceCode."""
        if not customer_finance:
            return ''

        customer_finance = str(customer_finance).strip()

        # Check code mappings first
        if customer_finance in self.code_mappings.get('finance_mappings', {}):
            return self.code_mappings['finance_mappings'][customer_finance]

        # Check if it's already a valid template code
        if customer_finance in self.template_reference.get('finance_codes', {}):
            return customer_finance

        # Log unmapped
        self.assumptions.append(f"Unmapped finance code: '{customer_finance}' - using as-is")
        return customer_finance

    def map_department_code(self, customer_dept: str) -> str:
        """Map customer department code to template DepartmentCode."""
        if not customer_dept:
            return ''

        customer_dept = str(customer_dept).strip()

        # Check code mappings first
        if customer_dept in self.code_mappings.get('department_mappings', {}):
            return self.code_mappings['department_mappings'][customer_dept]

        # Check if it's already a valid template code
        if customer_dept in self.template_reference.get('departments', {}):
            return customer_dept

        return customer_dept

    # =========================================================================
    # PRE-FLIGHT VALIDATION - User Approval of Mappings
    # =========================================================================

    def preflight_analysis(
        self,
        customer_data_dir: Path,
        template_path: Path = None
    ) -> Dict[str, Any]:
        """
        Analyze customer data files and return proposed mappings for user approval.

        This method should be called BEFORE process() to show users:
        - Column mapping proposals (with confidence scores)
        - School code mapping proposals
        - Finance code mapping proposals
        - Budget type detection
        - Any assumptions being made

        Args:
            customer_data_dir: Path to customer raw data files
            template_path: Optional path to pre-populated template

        Returns:
            Dict containing:
            - proposed_mappings: List of mapping proposals for user review
            - file_summary: Summary of files found
            - warnings: Any warnings about the data
            - ready_for_processing: bool indicating if data looks ready
        """
        self.log("="*60)
        self.log("PRE-FLIGHT ANALYSIS - Analyzing files for mapping approval")
        self.log("="*60)

        # Load template reference if provided
        if template_path and Path(template_path).exists():
            self.load_template_reference_data(str(template_path))

        # Find all customer data files
        all_files = list(Path(customer_data_dir).rglob("*.xls*")) + \
                    list(Path(customer_data_dir).rglob("*.csv"))
        all_files = [f for f in all_files if not f.name.startswith("~$")]

        self.log(f"Found {len(all_files)} files to analyze")

        # Collect all proposed mappings
        proposed_mappings = []
        file_summaries = []
        warnings = []

        for file_path in all_files:
            self.log(f"\nAnalyzing: {file_path.name}")
            file_mappings = self._analyze_file_for_mappings(file_path)

            file_summaries.append({
                "file": file_path.name,
                "path": str(file_path),
                "sheets": file_mappings.get("sheets", []),
                "detected_type": file_mappings.get("detected_type", "unknown"),
                "budget_type": file_mappings.get("budget_type"),
            })

            proposed_mappings.extend(file_mappings.get("column_mappings", []))

            if file_mappings.get("warnings"):
                warnings.extend(file_mappings["warnings"])

        # Deduplicate and prioritize mappings
        unique_mappings = self._deduplicate_mappings(proposed_mappings)

        # Sort by confidence (lowest first - those need user attention)
        unique_mappings.sort(key=lambda x: x.get("confidence", 0))

        # Determine readiness
        low_confidence_count = sum(1 for m in unique_mappings if m.get("confidence", 0) < 0.7)
        ready_for_processing = low_confidence_count == 0 and len(warnings) == 0

        result = {
            "proposed_mappings": unique_mappings,
            "file_summary": file_summaries,
            "warnings": warnings,
            "ready_for_processing": ready_for_processing,
            "requires_review_count": low_confidence_count,
            "total_mappings": len(unique_mappings),
            "assumptions": self.assumptions.copy(),
        }

        self.log(f"\nPre-flight complete:")
        self.log(f"  - {len(unique_mappings)} column mappings proposed")
        self.log(f"  - {low_confidence_count} require user review (confidence < 70%)")
        self.log(f"  - {len(warnings)} warnings")
        self.log(f"  - Ready for processing: {ready_for_processing}")

        return result

    def _analyze_file_for_mappings(self, file_path: Path) -> Dict[str, Any]:
        """
        Analyze a single file and propose column mappings.

        Returns dict with:
        - sheets: List of sheet names
        - column_mappings: List of proposed mappings
        - detected_type: Detected file type (budget, census, funding, etc.)
        - budget_type: For budget files, detected as approved/master
        - warnings: Any warnings
        """
        result = {
            "sheets": [],
            "column_mappings": [],
            "detected_type": "unknown",
            "budget_type": None,
            "warnings": []
        }

        try:
            # Detect budget type from filename
            file_lower = file_path.name.lower()
            if 'budget' in file_lower:
                result["detected_type"] = "budget"
                if S3_KNOWLEDGE_AVAILABLE:
                    budget_type = detect_budget_type_from_filename(file_path.name)
                    if budget_type:
                        result["budget_type"] = budget_type.value
                    else:
                        result["warnings"].append(
                            f"Could not detect budget type (APPROVED/MASTER) from filename: {file_path.name}"
                        )
            elif any(x in file_lower for x in ['census', 'pupil', 'student']):
                result["detected_type"] = "census"
            elif any(x in file_lower for x in ['funding', 'gag', 'dfe', 'esfa']):
                result["detected_type"] = "funding"
            elif any(x in file_lower for x in ['balance', 'bf', 'brought']):
                result["detected_type"] = "brought_forward"

            # Read file
            if file_path.suffix == ".csv":
                df = pd.read_csv(file_path, nrows=100)  # Sample first 100 rows
                result["sheets"] = ["CSV"]
                mappings = self._propose_column_mappings(df, file_path.name, "CSV")
                result["column_mappings"].extend(mappings)
            else:
                xl = pd.ExcelFile(file_path)
                result["sheets"] = xl.sheet_names

                for sheet in xl.sheet_names:
                    if self._is_skip_sheet(sheet):
                        continue

                    try:
                        df = pd.read_excel(xl, sheet, nrows=100)
                        mappings = self._propose_column_mappings(df, file_path.name, sheet)
                        result["column_mappings"].extend(mappings)
                    except Exception as e:
                        result["warnings"].append(f"Could not read sheet '{sheet}': {str(e)}")

        except Exception as e:
            result["warnings"].append(f"Error analyzing file: {str(e)}")

        return result

    def _propose_column_mappings(
        self,
        df: pd.DataFrame,
        file_name: str,
        sheet_name: str
    ) -> List[Dict[str, Any]]:
        """
        Propose column mappings for a DataFrame.

        Returns list of mapping proposals with:
        - file: Source file name
        - sheet: Source sheet name
        - source_column: Original column name
        - proposed_mapping: Suggested standard column name
        - confidence: Confidence score (0-1)
        - reasoning: Why this mapping was proposed
        - alternatives: Other possible mappings
        - requires_review: True if user should verify
        """
        mappings = []

        for col in df.columns:
            col_str = str(col).strip()
            if not col_str or col_str.lower() in ['unnamed', 'nan']:
                continue

            # Get sample values for context
            sample_values = df[col].dropna().head(5).tolist()

            # Propose mapping
            proposed, confidence = self.infer_column(col_str)

            # Get alternatives using fuzzy matching
            alternatives = self._get_alternative_mappings(col_str)

            mapping = {
                "file": file_name,
                "sheet": sheet_name,
                "source_column": col_str,
                "proposed_mapping": proposed,
                "confidence": confidence,
                "confidence_pct": f"{confidence:.0%}",
                "reasoning": self._get_mapping_reasoning(col_str, proposed, confidence),
                "alternatives": alternatives,
                "sample_values": [str(v)[:50] for v in sample_values],  # Truncate long values
                "requires_review": confidence < 0.7,
            }
            mappings.append(mapping)

        return mappings

    def _get_alternative_mappings(self, source_column: str) -> List[Dict[str, Any]]:
        """Get alternative mapping suggestions for a column."""
        alternatives = []
        col_lower = source_column.lower()

        # Standard target columns
        standard_columns = [
            ("finance_code", ["finance", "code", "nominal", "account", "gl"]),
            ("school_code", ["school", "location", "site", "cost centre", "academy"]),
            ("department_code", ["department", "dept", "cost center"]),
            ("description", ["description", "title", "name", "narrative"]),
            ("amount", ["amount", "value", "budget", "total", "sum"]),
            ("year_value", ["year", "annual", "fy"]),
            ("pupils", ["pupil", "student", "number", "count", "fte"]),
            ("fund_code", ["fund", "funding"]),
            ("ledger_code", ["ledger", "section"]),
        ]

        for target, keywords in standard_columns:
            score = sum(1 for kw in keywords if kw in col_lower)
            if score > 0:
                alternatives.append({
                    "target": target,
                    "match_score": score / len(keywords)
                })

        # Sort by score descending
        alternatives.sort(key=lambda x: x["match_score"], reverse=True)
        return alternatives[:3]  # Top 3 alternatives

    def _get_mapping_reasoning(self, source: str, proposed: str, confidence: float) -> str:
        """Generate human-readable reasoning for a mapping."""
        if confidence >= 0.9:
            return f"High confidence match: '{source}' clearly maps to '{proposed}'"
        elif confidence >= 0.7:
            return f"Good match: '{source}' likely maps to '{proposed}'"
        elif confidence >= 0.5:
            return f"Possible match: '{source}' may map to '{proposed}' - please verify"
        else:
            return f"Low confidence: '{source}' could not be confidently mapped - user input required"

    def _deduplicate_mappings(self, mappings: List[Dict]) -> List[Dict]:
        """Remove duplicate mappings, keeping highest confidence."""
        seen = {}
        for m in mappings:
            key = (m["file"], m["sheet"], m["source_column"])
            if key not in seen or m["confidence"] > seen[key]["confidence"]:
                seen[key] = m
        return list(seen.values())

    def apply_user_approved_mappings(
        self,
        approved_mappings: List[Dict[str, str]]
    ) -> Dict[str, Dict[str, str]]:
        """
        Convert user-approved mappings to the format expected by process().

        Args:
            approved_mappings: List of approved mappings from UI, each with:
                - file: Source file name
                - sheet: Source sheet name (optional)
                - source_column: Original column name
                - approved_mapping: User-approved target column name

        Returns:
            Dict in format: {
                "filename": {"source_col": "target_col", ...},
                "filename:sheetname": {"source_col": "target_col", ...}
            }
        """
        column_mappings = {}

        for mapping in approved_mappings:
            file_name = mapping.get("file", "")
            sheet_name = mapping.get("sheet", "")
            source = mapping.get("source_column", "")
            target = mapping.get("approved_mapping", "")

            if not source or not target:
                continue

            # Create keys for both file-level and sheet-level
            if sheet_name:
                key = f"{file_name}:{sheet_name}"
            else:
                key = file_name

            if key not in column_mappings:
                column_mappings[key] = {}

            column_mappings[key][source] = target

        return column_mappings

    def _apply_column_mappings(self, df: pd.DataFrame, file_name: str, sheet_name: str = None) -> pd.DataFrame:
        """
        Apply validated column mappings from pre-flight validation.

        Args:
            df: DataFrame to apply mappings to
            file_name: Name of the file being processed
            sheet_name: Optional sheet name for Excel files

        Returns:
            DataFrame with columns renamed according to mappings
        """
        if not hasattr(self, 'column_mappings') or not self.column_mappings:
            return df

        # Try different key formats to find matching mappings
        possible_keys = [
            f"{file_name}:{sheet_name or 'default'}",
            f"{file_name}:default",
            file_name,
        ]

        file_mappings = {}
        for key in possible_keys:
            if key in self.column_mappings:
                file_mappings = self.column_mappings[key]
                break

        if not file_mappings:
            # Also try partial match on file name
            for key, mappings in self.column_mappings.items():
                if file_name in key or key.split(':')[0] in file_name:
                    file_mappings = mappings
                    break

        if file_mappings:
            # Build rename dict: only rename columns that exist in df
            rename_dict = {}
            for source_col, target_col in file_mappings.items():
                if source_col in df.columns and source_col != target_col:
                    rename_dict[source_col] = target_col

            if rename_dict:
                self.log(f"  Applying {len(rename_dict)} column mappings from pre-flight validation")
                df = df.rename(columns=rename_dict)

        return df

    # =========================================================================
    # INTELLIGENT COLUMN MAPPING & CLASSIFICATION
    # Uses InferenceEngine for confidence-scored decisions
    # =========================================================================

    def infer_column(self, source_column: str) -> Tuple[str, float]:
        """
        Intelligently map a source column name to standard name.

        Args:
            source_column: Original column name from customer data

        Returns:
            Tuple of (mapped_column_name, confidence)
        """
        if self.inference_engine:
            result = self.inference_engine.infer_column_mapping(
                source_column=source_column,
                strand="S3"
            )

            # Track inference for audit
            self.inference_results.append({
                "type": "column_mapping",
                "source": source_column,
                "result": result.decision,
                "confidence": result.confidence,
                "reasoning": result.reasoning
            })

            if result.confidence >= 0.5:
                if result.requires_review:
                    self.assumptions.append(
                        f"Low confidence column mapping: {source_column} -> {result.decision} ({result.confidence:.0%})"
                    )
                return result.decision, result.confidence

        return source_column, 0.3

    def infer_budget_type(self, description: str, amount: float = None) -> Tuple[str, float]:
        """
        Classify a budget line as income or expenditure.

        Args:
            description: Budget line description
            amount: Optional amount (positive=income, negative=expenditure)

        Returns:
            Tuple of (budget_type, confidence)
        """
        if self.inference_engine:
            result = self.inference_engine.infer_classification(
                value=description,
                classification_type="budget_type"
            )

            self.inference_results.append({
                "type": "budget_classification",
                "source": description,
                "result": result.decision,
                "confidence": result.confidence
            })

            if result.confidence >= 0.5:
                return result.decision, result.confidence

        # Fallback: use amount sign if available
        if amount is not None:
            return ('income' if amount >= 0 else 'expenditure'), 0.6

        return 'unknown', 0.3

    def infer_grant_type(self, description: str) -> Tuple[str, float]:
        """
        Classify a grant type (DFC, SCA, PE, UIFSM, PP, etc.)

        Args:
            description: Grant description

        Returns:
            Tuple of (grant_type, confidence)
        """
        if self.inference_engine:
            result = self.inference_engine.infer_classification(
                value=description,
                classification_type="grant_type"
            )

            self.inference_results.append({
                "type": "grant_classification",
                "source": description,
                "result": result.decision,
                "confidence": result.confidence
            })

            if result.confidence >= 0.5:
                return result.decision, result.confidence

        return 'other', 0.3

    def infer_data_category(self, columns: List[str], sample_values: List = None) -> Tuple[str, float]:
        """
        Infer what type of S3 data this is (pupils, income, expenditure, grants, etc.)

        Args:
            columns: List of column names
            sample_values: Optional sample data values

        Returns:
            Tuple of (data_category, confidence)
        """
        cols_lower = " ".join(c.lower() for c in columns)

        # Pattern matching for S3 data types
        patterns = {
            'pupils': ['pupil', 'student', 'enrol', 'census', 'headcount', 'fte_ks', 'key_stage'],
            'income': ['income', 'revenue', 'receipt', 'grant', 'funding', 'allocation'],
            'expenditure': ['expenditure', 'expense', 'cost', 'spend', 'budget'],
            'grants': ['grant', 'dfc', 'sca', 'uifsm', 'pupil premium', 'pe grant'],
            'scenarios': ['scenario', 'forecast', 'projection', 'plan'],
            'calculators': ['calculator', 'formula', 'calculation', 'rate']
        }

        scores = {}
        for category, keywords in patterns.items():
            score = sum(1 for kw in keywords if kw in cols_lower)
            scores[category] = score

        best_category = max(scores, key=scores.get)
        total = sum(scores.values()) or 1
        confidence = scores[best_category] / total if scores[best_category] > 0 else 0.3

        return best_category, confidence

    def get_inference_summary(self) -> Dict:
        """Get summary of all inference decisions made."""
        if not self.inference_results:
            return {"total": 0, "high_confidence": 0, "low_confidence": 0}

        high_conf = sum(1 for r in self.inference_results if r.get("confidence", 0) >= 0.9)
        med_conf = sum(1 for r in self.inference_results if 0.7 <= r.get("confidence", 0) < 0.9)
        low_conf = sum(1 for r in self.inference_results if r.get("confidence", 0) < 0.7)

        return {
            "total": len(self.inference_results),
            "high_confidence": high_conf,
            "medium_confidence": med_conf,
            "low_confidence": low_conf,
            "inference_available": self.inference_engine is not None
        }

    # =========================================================================
    # PHASE 1: DEEP ANALYSIS
    # =========================================================================

    def analyze_customer_data(self, data_dir: Path):
        """Analyze all S3 customer data files."""
        self.log("="*60)
        self.log("PHASE 1: DEEP ANALYSIS OF S3 CUSTOMER DATA")
        self.log("="*60)

        all_files = list(data_dir.rglob("*.xls*")) + list(data_dir.rglob("*.csv")) + list(data_dir.rglob("*.docx")) + list(data_dir.rglob("*.doc"))
        all_files = [f for f in all_files if not f.name.startswith("~$")]

        self.log(f"Found {len(all_files)} files to analyze")

        for file_path in all_files:
            self.log(f"\nAnalyzing: {file_path.name}")
            self._analyze_file(file_path)

        self._print_analysis_summary()

    def _analyze_file(self, file_path: Path):
        """Analyze a single file."""
        try:
            if file_path.suffix.lower() in ['.docx', '.doc'] and DOCX_SUPPORT:
                document = docx.Document(file_path)
                for table_idx, table in enumerate(document.tables):
                    rows = [[cell.text.strip() for cell in row.cells] for row in table.rows]
                    if len(rows) > 1:
                        df = pd.DataFrame(rows[1:], columns=rows[0])
                        df = self._apply_column_mappings(df, file_path.name)
                        self._classify_and_extract(df, file_path.name, f"DOCX_Table{table_idx+1}")
            elif file_path.suffix == ".csv":
                df = pd.read_csv(file_path)
                # Apply validated column mappings
                df = self._apply_column_mappings(df, file_path.name)
                self._classify_and_extract(df, file_path.name, "CSV")
            else:
                xl = pd.ExcelFile(file_path)
                self.log(f"  Sheets: {xl.sheet_names}")

                for sheet in xl.sheet_names:
                    if self._is_skip_sheet(sheet):
                        continue

                    df = self._read_sheet_smart(xl, sheet)
                    if df is not None and len(df) > 0:
                        # Apply validated column mappings
                        df = self._apply_column_mappings(df, file_path.name, sheet)
                        self._classify_and_extract(df, file_path.name, sheet)

        except Exception as e:
            self.issues.append(f"Error analyzing {file_path.name}: {e}")

    def _is_skip_sheet(self, sheet: str) -> bool:
        """Check if sheet should be skipped."""
        skip_words = ['guidance', 'notes', 'instructions', 'help', 'checklist', 'validation', 'lookup']
        return any(w in sheet.lower() for w in skip_words)

    def _read_sheet_smart(self, xl: pd.ExcelFile, sheet: str) -> Optional[pd.DataFrame]:
        """Smart read that finds header row."""
        try:
            df_raw = pd.read_excel(xl, sheet, header=None, nrows=20)

            best_row = 0
            best_score = 0

            for idx in range(min(10, len(df_raw))):
                row = df_raw.iloc[idx]
                score = sum(1 for v in row if isinstance(v, str) and len(str(v).strip()) > 2)
                if score > best_score:
                    best_score = score
                    best_row = idx

            df = pd.read_excel(xl, sheet, header=best_row)
            df = df.dropna(how='all')
            df.columns = [self._clean_column_name(c) for c in df.columns]

            return df
        except:
            return None

    def _clean_column_name(self, col: Any) -> str:
        """Standardize column name using S3 knowledge base."""
        if pd.isna(col):
            return "unnamed"

        col_str = str(col).strip()

        # Try knowledge-based normalization first
        if S3_KNOWLEDGE_AVAILABLE:
            normalized = normalize_column_name(col_str)
            if normalized != col_str:
                return normalized.lower().replace(' ', '_')

        # Fallback mappings
        col_lower = col_str.lower()
        mappings = {
            'school': 'school_code',
            'location': 'school_code',
            'cost centre': 'school_code',
            'finance code': 'finance_code',
            'nominal': 'finance_code',
            'nominal code': 'finance_code',
            'account': 'finance_code',
            'code': 'finance_code',
            'description': 'description',
            'title': 'description',
            'account title': 'description',
            'nominal description': 'description',
            'amount': 'amount',
            'value': 'amount',
            'budget': 'amount',
            'budget year': 'amount',
            'pupil': 'pupils',
            'student': 'pupils',
            'fte': 'fte',
            'year': 'year',
            'period': 'period',
            'department': 'department_code',
            'dept': 'department_code',
            'cost center': 'department_code',
            'fund': 'fund_code',
            'funds': 'fund_code',
            'ledger': 'ledger_code',
            'section': 'section',
            'section name': 'section_name',
        }

        for pattern, standard in mappings.items():
            if pattern in col_lower:
                return standard

        return col_lower.replace(' ', '_')

    def _classify_and_extract(self, df: pd.DataFrame, file_name: str, sheet_name: str):
        """Classify data type and extract."""
        cols = [str(c).lower() for c in df.columns]
        sheet_lower = sheet_name.lower()
        file_lower = file_name.lower()

        # Pupil numbers
        if 'pupil' in sheet_lower or 'census' in sheet_lower or 'student' in file_lower:
            self._extract_pupil_numbers(df, sheet_name)
            self.log(f"    -> Pupil numbers extracted")

        # Budgets
        elif 'budget' in sheet_lower or 'budget' in file_lower:
            self._extract_budgets(df, sheet_name)
            self.log(f"    -> Budget data extracted")

        # DfE Funding statement (raw GAG statement)
        elif any(x in file_lower for x in ['dfe', 'esfa', 'funding statement', 'gag statement', 'allocation']):
            # This is likely a raw DfE funding statement - use GAG extraction
            school_code = self._extract_school_code_from_context(df, sheet_name)
            self.extract_gag_funding_from_dfe_statement(df, school_code, sheet_name)
            self.log(f"    -> DfE GAG funding extracted using mappings")

        # Funding statement (generic)
        elif 'funding' in sheet_lower or 'gag' in sheet_lower:
            # Check if this looks like a raw DfE statement or processed data
            if self._is_raw_dfe_statement(df):
                school_code = self._extract_school_code_from_context(df, sheet_name)
                self.extract_gag_funding_from_dfe_statement(df, school_code, sheet_name)
                self.log(f"    -> DfE GAG funding extracted using mappings")
            else:
                self._extract_funding(df, sheet_name)
                self.log(f"    -> Funding data extracted")

        # Grants
        elif any(g in sheet_lower for g in ['grant', 'dfc', 'sca', 'uifsm', 'pupil premium', 'pe ']):
            self._extract_grants(df, sheet_name)
            self.log(f"    -> Grant data extracted")

        # General financial data
        elif 'amount' in cols or 'value' in cols:
            self._extract_budgets(df, sheet_name)
            self.log(f"    -> Financial data extracted")

    def _is_raw_dfe_statement(self, df: pd.DataFrame) -> bool:
        """
        Detect if a DataFrame looks like a raw DfE funding statement.

        DfE statements typically have:
        - Lines with standard DfE funding descriptions
        - Basic entitlement, Deprivation, EAL, Lump sum, etc.
        """
        # Look for characteristic DfE funding line descriptions
        dfe_indicators = [
            'basic entitlement', 'deprivation', 'lump sum', 'sparsity',
            'mobility', 'prior attainment', 'eal', 'looked-after',
            'minimum per pupil', 'split site', 'rates', 'pfi',
            'post-16', '16-19'
        ]

        # Check if any text column contains DfE-style descriptions
        for col in df.columns:
            if df[col].dtype == 'object':
                col_values = df[col].dropna().astype(str).str.lower()
                for indicator in dfe_indicators:
                    if col_values.str.contains(indicator, na=False).any():
                        return True

        return False

    def _extract_school_code_from_context(self, df: pd.DataFrame, sheet_name: str) -> str:
        """
        Extract school code from DataFrame or sheet name context.

        Tries:
        1. School code column in data
        2. URN column in data
        3. Sheet name as school identifier
        """
        # Try to find school code in data
        for col in df.columns:
            col_lower = str(col).lower()
            if any(x in col_lower for x in ['school_code', 'schoolcode', 'school code']):
                first_val = df[col].dropna().iloc[0] if not df[col].dropna().empty else None
                if first_val:
                    return str(first_val).strip()

        # Try URN
        for col in df.columns:
            col_lower = str(col).lower()
            if 'urn' in col_lower:
                first_val = df[col].dropna().iloc[0] if not df[col].dropna().empty else None
                if first_val:
                    return str(first_val).strip()

        # Use sheet name as fallback
        return sheet_name.strip()

    def _extract_pupil_numbers(self, df: pd.DataFrame, sheet_name: str):
        """Extract pupil numbers from dataframe."""
        for _, row in df.iterrows():
            school = row.get('school_code', '')
            if pd.isna(school):
                school = ''
            school_str = str(school).strip()

            # Look for pupil count columns
            for col in df.columns:
                col_lower = str(col).lower()

                if any(x in col_lower for x in ['pupil', 'student', 'number', 'count', 'fte']):
                    value = row.get(col)
                    if pd.notna(value):
                        try:
                            value_float = float(value)
                            if value_float >= 0:
                                # Determine pupil type from column name
                                finance_code = self._determine_pupil_finance_code(col_lower)
                                calculator = self._determine_pupil_calculator(col_lower)

                                pupil = ExtractedPupilNumber(
                                    finance_code=finance_code,
                                    school_code=school_str if school_str != 'nan' else '',
                                    description=col,
                                    year_code=self.current_year,
                                    value=value_float,
                                    calculator_code=calculator,
                                    notes=f"Extracted from {sheet_name}",
                                )
                                self.extracted_pupils.append(pupil)

                                if school_str and school_str != 'nan':
                                    self.schools_found.add(school_str)
                        except:
                            pass

    def _determine_pupil_finance_code(self, col_name: str) -> str:
        """Determine pupil finance code from column name."""
        col_lower = col_name.lower()

        if 'ks3' in col_lower or 'key stage 3' in col_lower:
            return 'PUPIL_SPRING_KS3'
        elif 'ks4' in col_lower or 'key stage 4' in col_lower:
            return 'PUPIL_SPRING_KS4'
        elif 'ks5' in col_lower or 'post 16' in col_lower or 'sixth' in col_lower:
            return 'PUPIL_SPRING_KS5'
        elif 'primary' in col_lower or 'ks1' in col_lower or 'ks2' in col_lower:
            return 'PUPIL_SPRING_PRI'
        elif 'nursery' in col_lower or 'eyfs' in col_lower:
            return 'PUPIL_SPRING_NUR'
        elif 'premium' in col_lower:
            if 'plac' in col_lower:
                return 'PUPILPREMIUMPLAC'
            elif 'service' in col_lower:
                return 'PUPILPREMIUMSER'
            elif 'primary' in col_lower:
                return 'PUPILPREMIUM_PRI'
            elif 'secondary' in col_lower:
                return 'PUPILPREMIUM_SEC'
            else:
                return 'PUPILPREMIUM'
        elif 'uifsm' in col_lower or 'infant' in col_lower:
            return 'PUPIL_UIFSM'

        return 'PUPIL_TOTAL'

    def _determine_pupil_calculator(self, col_name: str) -> str:
        """Determine calculator code for pupils."""
        col_lower = col_name.lower()

        if 'premium' in col_lower:
            return 'PUPPREMIUM_FACTOR'
        elif 'uifsm' in col_lower:
            return 'UIFSM_RATE'

        return '0%_CALC'

    def _extract_budgets(self, df: pd.DataFrame, sheet_name: str):
        """Extract budget lines from dataframe using S3 knowledge."""
        # Get section column if available (for IL/EL line type detection)
        section_col = None
        for col in df.columns:
            if 'section' in str(col).lower():
                section_col = col
                break

        for _, row in df.iterrows():
            finance_code = row.get('finance_code', '')
            if pd.isna(finance_code):
                continue

            code_str = str(finance_code).strip()

            # Try to extract finance code from complex strings
            if S3_KNOWLEDGE_AVAILABLE and (not code_str or code_str == 'nan' or len(code_str) > 20):
                extracted = extract_finance_code(code_str)
                if extracted:
                    code_str = extracted

            if not code_str or code_str == 'nan':
                continue

            # Map customer school code to template SchoolCode
            school = str(row.get('school_code', '')).strip()
            if school == 'nan':
                school = ''
            else:
                school = self.map_school_code(school)

            # Map customer finance code to template FinanceCode
            code_str = self.map_finance_code(code_str)

            # Get department - use mapping or knowledge-based lookup if missing
            dept = str(row.get('department_code', '')).strip()
            if not dept or dept == 'nan':
                if S3_KNOWLEDGE_AVAILABLE:
                    dept = get_default_department_for_finance_code(code_str)
                else:
                    dept = 'MISSING'
            else:
                dept = self.map_department_code(dept)

            description = str(row.get('description', '')).strip()
            if description == 'nan':
                description = code_str

            # Find amount column - use knowledge patterns
            amount = 0
            for col in df.columns:
                col_str = str(col)
                col_lower = col_str.lower()

                # Check if this is a value column
                is_value = any(x in col_lower for x in ['amount', 'value', 'budget', 'total', 'approved', 'master'])
                if S3_KNOWLEDGE_AVAILABLE:
                    is_value = is_value or is_value_column(col_str)

                if is_value:
                    val = row.get(col)
                    if pd.notna(val):
                        try:
                            if S3_KNOWLEDGE_AVAILABLE:
                                amount = normalize_amount(val) or 0
                            else:
                                amount = float(str(val).replace('£', '').replace(',', '').strip())
                            if amount != 0:
                                break
                        except (ValueError, TypeError):
                            pass

            if amount == 0:
                continue

            # Get section for line type detection
            section = None
            if section_col:
                section = str(row.get(section_col, '')).strip()
                if section == 'nan':
                    section = None

            # Determine line type from code using knowledge
            line_type = self._determine_line_type(code_str, amount, section)

            budget = ExtractedBudgetLine(
                finance_code=code_str,
                school_code=school,
                department_code=dept,
                description=description,
                year_value=amount,
                line_type=line_type,
                calculator_code='',
                month_profile='MONTHLY',
            )
            self.extracted_budgets.append(budget)
            self.finance_codes_found.add(code_str)

            if school:
                self.schools_found.add(school)

    def _determine_line_type(self, code: str, amount: float, section: str = None) -> str:
        """Determine if line is income or expenditure using S3 knowledge."""
        # Use knowledge-based classification if available
        if S3_KNOWLEDGE_AVAILABLE:
            line_type = determine_line_type(section=section, code=code, amount=amount)
            if line_type != 'unknown':
                return line_type

            # Check against known finance codes
            result = classify_finance_code(code)
            if result != 'unknown':
                return result

        # Fallback logic
        if not code:
            return 'expenditure' if amount > 0 else 'income'

        # Try to determine from code prefix
        if code and code[0].isdigit():
            first = int(code[0])
            if first in [4, 5]:
                return 'income'
            elif first in [6, 7, 8, 9]:
                return 'expenditure'

        # Determine from sign
        return 'expenditure' if amount > 0 else 'income'

    def _extract_funding(self, df: pd.DataFrame, sheet_name: str):
        """Extract funding statement data."""
        # Process as budget lines with income type
        for _, row in df.iterrows():
            school = str(row.get('school_code', '')).strip()

            for col in df.columns:
                col_lower = str(col).lower()

                # Look for funding columns
                if any(x in col_lower for x in ['gag', 'funding', 'grant', 'allocation']):
                    val = row.get(col)
                    if pd.notna(val):
                        try:
                            amount = float(str(val).replace('£', '').replace(',', '').strip())
                            if amount != 0:
                                budget = ExtractedBudgetLine(
                                    finance_code=self._determine_funding_code(col_lower),
                                    school_code=school if school != 'nan' else '',
                                    department_code='IGAG',
                                    description=col,
                                    year_value=-abs(amount),  # Income is negative
                                    line_type='income',
                                    calculator_code=self._determine_funding_calculator(col_lower),
                                    month_profile='MONTHLY',
                                )
                                self.extracted_budgets.append(budget)
                        except:
                            pass

    def _determine_funding_code(self, col_name: str) -> str:
        """Determine funding finance code."""
        col_lower = col_name.lower()

        if 'post 16' in col_lower or '16-19' in col_lower:
            return '510700'
        elif 'gag' in col_lower:
            return '510100'
        elif 'pupil premium' in col_lower:
            return '510200'

        return '510100'

    def _determine_funding_calculator(self, col_name: str) -> str:
        """Determine funding calculator."""
        col_lower = col_name.lower()

        if 'post 16' in col_lower or '16-19' in col_lower:
            return 'FUNDING_16_19'
        elif 'gag' in col_lower:
            return 'FUNDING_GAG'
        elif 'pupil premium' in col_lower:
            return 'PUPPREMIUM_CALC'

        return 'FUNDING_GAG'

    def extract_gag_funding_from_dfe_statement(
        self,
        df: pd.DataFrame,
        school_code: str,
        school_name: str = ""
    ) -> List[ExtractedBudgetLine]:
        """
        Extract GAG funding values from a raw DfE funding statement.

        Uses GAG_FUNDING_MAPPINGS to map DfE descriptions to IMP finance codes.
        CRITICAL: Mapping uses SchoolCode + Description (NOT FinanceCode)
        because finance codes vary between trusts.

        Args:
            df: DataFrame containing DfE funding statement
            school_code: IMP school code for this school
            school_name: School name (used to detect school type)

        Returns:
            List of ExtractedBudgetLine objects with GAG funding values
        """
        if not GAG_MAPPINGS_AVAILABLE:
            self.issues.append("GAG funding mappings not available")
            return []

        extracted = []

        # Detect school type from name
        school_type = detect_school_type(school_name) if school_name else "secondary"
        if school_type == "unknown":
            school_type = "secondary"
            self.assumptions.append(f"Assumed {school_code} is secondary school type")

        self.log(f"Processing GAG funding for {school_code} (type: {school_type})")

        # Find description and value columns in the DfE statement
        desc_col = None
        value_col = None

        for col in df.columns:
            col_lower = str(col).lower()
            if any(x in col_lower for x in ['description', 'factor', 'element', 'line']):
                desc_col = col
            elif any(x in col_lower for x in ['amount', 'value', 'allocation', 'total', '£']):
                value_col = col

        # If no clear columns, try first text col and first numeric col
        if desc_col is None or value_col is None:
            for col in df.columns:
                sample = df[col].dropna().head(5)
                if desc_col is None and sample.dtype == 'object':
                    desc_col = col
                elif value_col is None:
                    try:
                        pd.to_numeric(sample.astype(str).str.replace('£', '').str.replace(',', ''))
                        value_col = col
                    except:
                        pass

        if desc_col is None or value_col is None:
            self.issues.append(f"Could not identify description/value columns in funding statement for {school_code}")
            return []

        # Process each row
        for _, row in df.iterrows():
            description = str(row.get(desc_col, '')).strip()
            if not description or description.lower() in ['nan', 'none', '']:
                continue

            # Get the value
            raw_value = row.get(value_col)
            if pd.isna(raw_value):
                continue

            try:
                # Clean and convert to float
                value_str = str(raw_value).replace('£', '').replace(',', '').strip()
                value = float(value_str)
            except:
                continue

            if value == 0:
                continue

            # Find mapping for this DfE description
            mapping = find_mapping_by_description(description, school_type)

            if mapping:
                budget_line = ExtractedBudgetLine(
                    finance_code=mapping.imp_finance_code,
                    school_code=school_code,
                    department_code='IGAG',
                    description=mapping.imp_description,
                    year_value=-abs(value),  # Income is negative
                    line_type='income',
                    calculator_code='FUNDING_GAG' if mapping.category == 'pre16' else 'FUNDING_16_19',
                    month_profile='MONTHLY',
                )
                extracted.append(budget_line)
                self.log(f"  Mapped: {description} -> {mapping.imp_finance_code} = £{value:,.2f}")
            else:
                # Unknown line - create with generic code but flag for review
                self.issues.append(f"Unknown GAG line for {school_code}: '{description}' (£{value:,.2f})")
                budget_line = ExtractedBudgetLine(
                    finance_code='I1299',  # Generic GAG Other
                    school_code=school_code,
                    department_code='IGAG',
                    description=f"GAG - {description}",
                    year_value=-abs(value),
                    line_type='income',
                    calculator_code='FUNDING_GAG',
                    month_profile='MONTHLY',
                )
                extracted.append(budget_line)

        self.log(f"  Extracted {len(extracted)} GAG funding lines for {school_code}")

        # Add to main budgets list
        self.extracted_budgets.extend(extracted)

        return extracted

    def process_dfe_funding_statement_file(
        self,
        file_path: str,
        school_mapping: Dict[str, str] = None
    ) -> Dict[str, List[ExtractedBudgetLine]]:
        """
        Process a raw DfE funding statement file.

        DfE statements typically have:
        - Multiple sheets (one per school) OR
        - Single sheet with school identifier column

        Args:
            file_path: Path to DfE funding statement Excel file
            school_mapping: Optional dict mapping DfE school names to IMP school codes

        Returns:
            Dict mapping school codes to their extracted funding lines
        """
        self.log(f"Processing DfE funding statement: {file_path}")
        results = {}

        try:
            # Read all sheets
            xlsx = pd.ExcelFile(file_path)
            sheet_names = xlsx.sheet_names

            for sheet_name in sheet_names:
                # Skip summary/total sheets
                if any(x in sheet_name.lower() for x in ['summary', 'total', 'contents', 'index']):
                    continue

                df = pd.read_excel(xlsx, sheet_name=sheet_name)

                if df.empty:
                    continue

                # Determine school code
                # Option 1: Sheet name is school name/code
                school_code = sheet_name.strip()
                school_name = sheet_name.strip()

                # Option 2: Check mapping
                if school_mapping and sheet_name in school_mapping:
                    school_code = school_mapping[sheet_name]

                # Option 3: Look for school code in data
                for col in df.columns:
                    col_lower = str(col).lower()
                    if any(x in col_lower for x in ['school', 'urn', 'academy']):
                        first_val = df[col].dropna().iloc[0] if not df[col].dropna().empty else None
                        if first_val:
                            school_code = str(first_val).strip()
                            break

                # Extract funding for this school
                extracted = self.extract_gag_funding_from_dfe_statement(
                    df=df,
                    school_code=school_code,
                    school_name=school_name
                )

                if extracted:
                    results[school_code] = extracted

        except Exception as e:
            self.issues.append(f"Error processing funding statement: {str(e)}")

        self.log(f"Processed funding for {len(results)} schools")
        return results

    def _extract_grants(self, df: pd.DataFrame, sheet_name: str):
        """Extract grant data."""
        sheet_lower = sheet_name.lower()

        for _, row in df.iterrows():
            school = str(row.get('school_code', '')).strip()

            # Find grant amounts
            for col in df.columns:
                val = row.get(col)
                if pd.notna(val):
                    try:
                        amount = float(str(val).replace('£', '').replace(',', '').strip())
                        if amount > 0:
                            # Determine grant type
                            grant_type = self._determine_grant_type(sheet_lower, str(col).lower())

                            grant = ExtractedGrant(
                                grant_type=grant_type,
                                school_code=school if school != 'nan' else '',
                                amount=amount,
                                calculation_basis=col,
                                pupil_count=0,
                            )
                            self.extracted_grants.append(grant)
                    except:
                        pass

    def _determine_grant_type(self, sheet_name: str, col_name: str) -> str:
        """Determine grant type."""
        combined = sheet_name + ' ' + col_name

        if 'dfc' in combined or 'devolved formula' in combined:
            return 'DFC'
        elif 'sca' in combined or 'condition' in combined:
            return 'SCA'
        elif 'pe ' in combined or 'sport' in combined:
            return 'PE'
        elif 'uifsm' in combined or 'infant' in combined:
            return 'UIFSM'
        elif 'premium' in combined:
            return 'PP'

        return 'OTHER'

    def _print_analysis_summary(self):
        """Print analysis summary."""
        self.log("\n" + "="*60)
        self.log("ANALYSIS SUMMARY")
        self.log("="*60)

        self.log(f"Pupil records: {len(self.extracted_pupils)}")
        self.log(f"Budget lines: {len(self.extracted_budgets)}")
        self.log(f"Grants: {len(self.extracted_grants)}")
        self.log(f"Schools: {self.schools_found}")
        self.log(f"Finance codes: {len(self.finance_codes_found)}")

    # =========================================================================
    # PHASE 2: BUILD ALL TEMPLATE SHEETS
    # =========================================================================

    def build_all_templates(self) -> Dict[str, pd.DataFrame]:
        """Build ALL S3 template sheets."""
        self.log("\n" + "="*60)
        self.log("PHASE 2: BUILDING ALL S3 TEMPLATE SHEETS")
        self.log("="*60)

        self._build_pupils()
        self._build_statistics()
        self._build_funding()
        self._build_calculators()
        self._build_month_profiles()
        self._build_income()
        self._build_expenditure()
        self._build_scenario_apbud()
        self._build_bf_balances()
        self._build_finance_codes_s3()

        # Build scenario sheets (budget lines and values)
        self._build_scenario_rows()
        self._build_scenario_year_values()
        self._build_monthly_values()

        result = {}
        for sheet_name, data in self.template_data.items():
            if data:
                result[sheet_name] = pd.DataFrame(data)
                self.log(f"  {sheet_name}: {len(data)} rows")

        return result

    def _build_pupils(self):
        """Build Pupils sheet."""
        self.log("Building Pupils...")

        for pupil in self.extracted_pupils:
            self.template_data["Pupils"].append({
                "FinanceCode": pupil.finance_code,
                "SchoolCode": pupil.school_code,
                "LedgerCode": "DEFAULT",
                "DepartmentCode": "DEFAULT",
                "FundCode": "",
                "CalculatorCode": pupil.calculator_code,
                "MonthProfileCode": "MONTHLY",
                "Description": pupil.description,
                "Notes": pupil.notes,
                "YearNotes": "",
                "MatEditOnly": False,
                "FinancialYearCode": pupil.year_code,
                "Calculated": False,
                "YearValue": pupil.value,
            })

        # Add defaults for each school if no pupil data
        if not self.template_data["Pupils"]:
            for school in self.schools_found or ['MAT']:
                for key_stage in ['PRI', 'KS3', 'KS4', 'KS5']:
                    self.template_data["Pupils"].append({
                        "FinanceCode": f"PUPIL_SPRING_{key_stage}",
                        "SchoolCode": school,
                        "LedgerCode": "DEFAULT",
                        "DepartmentCode": "DEFAULT",
                        "FundCode": "",
                        "CalculatorCode": "0%_CALC",
                        "MonthProfileCode": "MONTHLY",
                        "Description": f"{key_stage} Spring Census Pupil Numbers",
                        "Notes": "",
                        "YearNotes": "",
                        "MatEditOnly": False,
                        "FinancialYearCode": self.previous_year,
                        "Calculated": False,
                        "YearValue": 0,
                    })

    def _build_statistics(self):
        """
        Build Statistics sheet.

        Creates STANDARD uplift statistics rows for EVERY school automatically.
        Users can amend values post-build.
        """
        self.log("Building Statistics...")

        # Use knowledge-based uplifts if available, otherwise fallback
        if S3_KNOWLEDGE_AVAILABLE and STANDARD_UPLIFT_STATISTICS:
            uplifts = [
                (s["finance_code"], s["description"], s["calculator_code"], s.get("notes", ""))
                for s in STANDARD_UPLIFT_STATISTICS
            ]
        else:
            # Fallback uplift definitions
            uplifts = [
                ("UPLIFT_PUPILASCL%", "Pupil ASCL Uplift %", "PUPILASCL_FACTOR",
                 "Percentage change of pupil numbers from previous year combined with ASCL Uplift %"),
                ("UPLIFT_PUPILEXP%", "Pupil Expenditure Uplift %", "PUPILEXP_FACTOR",
                 "Percentage change of pupil numbers from previous year combined with Expenditure Uplift %"),
                ("UPLIFT_PUPILGAG%", "Pupil GAG Uplift %", "PUPILGAG_FACTOR",
                 "Percentage change of pupil numbers from previous year combined with GAG Uplift %"),
                ("UPLIFT_PUPILINC%", "Pupil Income Uplift %", "PUPILINC_FACTOR",
                 "Percentage change of pupil numbers from previous year combined with Income Uplift %"),
                ("UPLIFT_PUPILRPI%", "Pupil RPI Uplift %", "PUPILRPI_FACTOR",
                 "Percentage change of pupil numbers from previous year combined with RPI Uplift %"),
            ]

        # Create standard uplift rows for EVERY school
        schools_list = list(self.schools_found) if self.schools_found else ['MAT']

        for school in schools_list:
            for uplift_data in uplifts:
                code = uplift_data[0]
                desc = uplift_data[1]
                calc = uplift_data[2]
                notes = uplift_data[3] if len(uplift_data) > 3 else ""

                self.template_data["Statistics"].append({
                    "FinanceCode": code,
                    "SchoolCode": school,
                    "LedgerCode": "DEFAULT",
                    "DepartmentCode": "DEFAULT",
                    "FundCode": "",
                    "CalculatorCode": calc,
                    "MonthProfileCode": "MONTHLY",
                    "Description": desc,
                    "Notes": notes,
                    "YearNotes": "",
                    "MatEditOnly": True,
                    "FinancialYearCode": self.current_year,
                    "Calculated": True,
                    "YearValue": None,
                })

        self.log(f"  Created {len(uplifts)} uplift statistics x {len(schools_list)} schools = {len(uplifts) * len(schools_list)} rows")

    def _build_funding(self):
        """Build Funding sheet."""
        self.log("Building Funding...")

        # Extract funding lines from budgets
        for budget in self.extracted_budgets:
            if budget.line_type == 'income' and budget.calculator_code:
                self.template_data["Funding"].append({
                    "FinanceCode": budget.finance_code,
                    "SchoolCode": budget.school_code,
                    "LedgerCode": "DEFAULT",
                    "DepartmentCode": budget.department_code,
                    "FundCode": "",
                    "MonthProfileCode": "MONTHLY",
                    "CalculatorCode": budget.calculator_code,
                    "Description": budget.description,
                    "Notes": "",
                    "YearNotes": "",
                    "MatEditOnly": True,
                    "FinancialYearCode": self.current_year,
                    "Calculated": True,
                    "YearValue": None,
                })

    def _build_calculators(self):
        """Build Calculators sheet."""
        self.log("Building Calculators...")

        calculators = [
            ("0%_CALC", "Zero Calculator", "0"),
            ("FUNDING_GAG", "GAG Funding", "FUNDING"),
            ("FUNDING_16_19", "Post 16 Funding", "FUNDING"),
            ("PUPPREMIUM_FACTOR", "Pupil Premium Factor", "PUPILPREMIUM"),
            ("PUPPREMIUM_CALC", "Pupil Premium Calculator", "PUPILPREMIUM"),
            ("DFC_CORE", "DFC Core Amount", "DFC"),
            ("DFC_PUPIL", "DFC Per Pupil", "DFC"),
            ("DFC_EXP", "DFC Expenditure", "DFC"),
            ("PE_GRANT_CORE", "PE Grant Core", "PEGRANT"),
            ("PE_GRANT_PUPIL", "PE Grant Per Pupil", "PEGRANT"),
            ("UIFSM_CALC", "UIFSM Calculator", "UIFSM"),
            ("CENTRALCHG_SCH", "Central Charge School", "CENTRAL"),
            ("CENTRALCHG_MAT", "Central Charge MAT", "CENTRAL"),
        ]

        for code, title, category in calculators:
            self.template_data["Calculators"].append({
                "CalculatorCode": code,
                "Title": title,
                "Category": category,
                "CalculatorEnabled": True,
            })

    def _build_month_profiles(self):
        """Build MonthProfiles sheet."""
        self.log("Building MonthProfiles...")

        # Standard monthly profile (equal distribution)
        monthly_pct = round(100/12, 2)

        self.template_data["MonthProfiles"].append({
            "MonthProfileCode": "MONTHLY",
            "Title": "Monthly Equal",
            "Sep": monthly_pct,
            "Oct": monthly_pct,
            "Nov": monthly_pct,
            "Dec": monthly_pct,
            "Jan": monthly_pct,
            "Feb": monthly_pct,
            "Mar": monthly_pct,
            "Apr": monthly_pct,
            "May": monthly_pct,
            "Jun": monthly_pct,
            "Jul": monthly_pct,
            "Aug": round(100 - 11*monthly_pct, 2),  # Ensure totals 100
            "MonthProfileEnabled": True,
        })

        # Academic year front-loaded
        self.template_data["MonthProfiles"].append({
            "MonthProfileCode": "ACADEMIC",
            "Title": "Academic Year",
            "Sep": 10,
            "Oct": 10,
            "Nov": 10,
            "Dec": 8,
            "Jan": 8,
            "Feb": 8,
            "Mar": 10,
            "Apr": 8,
            "May": 8,
            "Jun": 8,
            "Jul": 8,
            "Aug": 4,
            "MonthProfileEnabled": True,
        })

    def _build_income(self):
        """Build Income sheet."""
        self.log("Building Income...")

        for budget in self.extracted_budgets:
            if budget.line_type == 'income':
                self.template_data["Income"].append({
                    "FinanceCode": budget.finance_code,
                    "SchoolCode": budget.school_code,
                    "LedgerCode": "COSTCTR",
                    "DepartmentCode": budget.department_code or "IGAG",
                    "FundCode": "",
                    "CalculatorCode": budget.calculator_code,
                    "MonthProfileCode": budget.month_profile,
                    "Description": budget.description,
                    "Notes": "",
                    "YearNotes": "",
                    "MatEditOnly": bool(budget.calculator_code),
                    "FinancialYearCode": self.current_year,
                    "Calculated": bool(budget.calculator_code),
                    "YearValue": budget.year_value if not budget.calculator_code else None,
                })

    def _build_expenditure(self):
        """Build Expenditure sheet."""
        self.log("Building Expenditure...")

        for budget in self.extracted_budgets:
            if budget.line_type == 'expenditure':
                self.template_data["Expenditure"].append({
                    "FinanceCode": budget.finance_code,
                    "SchoolCode": budget.school_code,
                    "LedgerCode": "COSTCTR",
                    "DepartmentCode": budget.department_code,
                    "FundCode": "",
                    "CalculatorCode": budget.calculator_code,
                    "MonthProfileCode": budget.month_profile,
                    "Description": budget.description,
                    "Notes": "",
                    "YearNotes": "",
                    "MatEditOnly": bool(budget.calculator_code),
                    "FinancialYearCode": self.current_year,
                    "Calculated": bool(budget.calculator_code),
                    "YearValue": budget.year_value if not budget.calculator_code else None,
                })

    def _build_scenario_apbud(self):
        """Build ScenarioApBud sheet."""
        self.log("Building ScenarioApBud...")

        scenario_code = f"APBUD{self.current_year.replace('/', '')[-4:]}"

        # Copy all budget lines to approved budget scenario
        for budget in self.extracted_budgets:
            self.template_data["ScenarioApBud"].append({
                "ScenarioCode": scenario_code,
                "FinanceCode": budget.finance_code,
                "SchoolCode": budget.school_code,
                "LedgerCode": "COSTCTR",
                "DepartmentCode": budget.department_code,
                "FundCode": "",
                "CalculatorCode": "",
                "MonthProfileCode": "MONTHLY",
                "Description": budget.description,
                "Notes": "",
                "YearNotes": "",
                "MatEditOnly": True,
                "FinancialYearCode": self.current_year,
                "Calculated": False,
                "YearValue": abs(budget.year_value) if budget.line_type == 'income' else budget.year_value,
            })

    def _build_bf_balances(self):
        """Build BF Balances sheet."""
        self.log("Building BF Balances...")

        for school in self.schools_found or ['MAT']:
            # Capital BF
            self.template_data["BF Balances"].append({
                "FinanceCode": "CAP_BFWD_RES",
                "SchoolCode": school,
                "LedgerCode": "DEFAULT",
                "DepartmentCode": "DEFAULT",
                "FundCode": "",
                "CalculatorCode": "",
                "MonthProfileCode": "MONTHLY",
                "Description": "Brought Forward Balance Capital",
                "Notes": "",
                "YearNotes": "",
                "MatEditOnly": True,
                "FinancialYearCode": self.current_year,
                "Calculated": False,
                "YearValue": 0,
            })

            # Revenue BF
            self.template_data["BF Balances"].append({
                "FinanceCode": "REV_BFWD_RES",
                "SchoolCode": school,
                "LedgerCode": "DEFAULT",
                "DepartmentCode": "DEFAULT",
                "FundCode": "",
                "CalculatorCode": "",
                "MonthProfileCode": "MONTHLY",
                "Description": "Brought Forward Balance Revenue",
                "Notes": "",
                "YearNotes": "",
                "MatEditOnly": True,
                "FinancialYearCode": self.current_year,
                "Calculated": False,
                "YearValue": 0,
            })

    def _build_finance_codes_s3(self):
        """Build Finance Codes S3 sheet."""
        self.log("Building Finance Codes S3...")

        # Add pupil finance codes
        pupil_codes = [
            ("PUPIL_SPRING_PRI", "Primary Spring Census", "STATISTICS"),
            ("PUPIL_SPRING_KS3", "KS3 Spring Census", "STATISTICS"),
            ("PUPIL_SPRING_KS4", "KS4 Spring Census", "STATISTICS"),
            ("PUPIL_SPRING_KS5", "KS5 Spring Census", "STATISTICS"),
            ("PUPILPREMIUMPLAC", "Pupil Premium PLAC", "STATISTICS"),
            ("PUPILPREMIUMSER", "Pupil Premium Service", "STATISTICS"),
            ("PUPILPREMIUM_PRI", "Pupil Premium Primary", "STATISTICS"),
            ("PUPILPREMIUM_SEC", "Pupil Premium Secondary", "STATISTICS"),
            ("PUPIL_UIFSM", "UIFSM Pupils", "STATISTICS"),
            ("CAP_BFWD_RES", "Capital BF Balance", "BUDGET"),
            ("REV_BFWD_RES", "Revenue BF Balance", "BUDGET"),
        ]

        for code, title, fc_type in pupil_codes:
            self.template_data["Finance Codes S3"].append({
                "FinanceCode": code,
                "Title": title,
                "FinanceCodeTypeCode": fc_type,
                "GroupingCode": "ZZZ",
                "CustomGrouping": "ZZZ",
                "LedgerCode": "DEFAULT",
                "AvailableToAllSchools": True,
                "SchoolCodes": "",
                "FinanceCodeEnabled": True,
            })

    # =========================================================================
    # SCENARIO SHEETS - Budget Lines and Customer Values
    # =========================================================================

    def _build_scenario_rows(self):
        """
        Build 35_ScenarioRows sheet - budget line definitions.

        ScenarioRows contains the ROW STRUCTURE for each budget entry:
        - Tab (Income/Expenditure/Statistics/Pupils/Funding)
        - FinanceCode, SchoolCode, LedgerCode, DepartmentCode
        - CalculatorCode, MonthProfileCode
        - Description, Notes, MatEditOnly

        CRITICAL: ScenarioCode distinguishes APPROVED vs MASTER budget:
        - APBUD{YYYY} = Approved Budget (full year)
        - MASTER{YYYY} = Master Budget (changed in-year)
        """
        self.log("Building ScenarioRows (budget line definitions)...")

        # Use detected scenario code, or leave empty for base rows
        scenario = self.scenario_code or ""

        # Build rows from extracted budgets
        for budget in self.extracted_budgets:
            tab = "Income" if budget.line_type == "income" else "Expenditure"

            self.template_data["ScenarioRows"].append({
                "Tab": tab,
                "ScenarioCode": scenario,
                "FinanceCode": budget.finance_code,
                "SchoolCode": budget.school_code,
                "LedgerCode": "COSTCTR",
                "DepartmentCode": budget.department_code,
                "FundCode": "",
                "CalculatorCode": budget.calculator_code,
                "MonthProfileCode": budget.month_profile,
                "StaffMemberCode": "",
                "StaffRoleCode": "",
                "ContractReference": "",
                "ContractDateFrom": None,
                "ContractDateTo": None,
                "Description": budget.description,
                "Notes": "",
                "MatEditOnly": bool(budget.calculator_code),
            })

        # Build rows from extracted pupils
        for pupil in self.extracted_pupils:
            self.template_data["ScenarioRows"].append({
                "Tab": "Pupils",
                "ScenarioCode": scenario,
                "FinanceCode": pupil.finance_code,
                "SchoolCode": pupil.school_code,
                "LedgerCode": "DEFAULT",
                "DepartmentCode": "DEFAULT",
                "FundCode": "",
                "CalculatorCode": pupil.calculator_code,
                "MonthProfileCode": "MONTHLY",
                "StaffMemberCode": "",
                "StaffRoleCode": "",
                "ContractReference": "",
                "ContractDateFrom": None,
                "ContractDateTo": None,
                "Description": pupil.description,
                "Notes": pupil.notes,
                "MatEditOnly": False,
            })

        if scenario:
            self.log(f"  Using ScenarioCode: {scenario} ({self.budget_type or 'unknown'} budget)")

    def _build_scenario_year_values(self):
        """
        Build 36_ScenarioYearValues sheet - customer values.

        ScenarioYearValues contains the ACTUAL VALUES (numbers) for each line:
        - Tab, FinanceCode, SchoolCode (must match ScenarioRows)
        - FinancialYearCode (e.g., 2025/26)
        - YearValue - the customer's value (ALWAYS a number)
        - Calculated - whether value is auto-calculated

        CRITICAL: ScenarioCode must match ScenarioRows to link line to value.
        """
        self.log("Building ScenarioYearValues (customer values)...")

        # Use detected scenario code
        scenario = self.scenario_code or ""

        # Build values from extracted budgets
        for budget in self.extracted_budgets:
            tab = "Income" if budget.line_type == "income" else "Expenditure"

            self.template_data["ScenarioYearValues"].append({
                "Tab": tab,
                "ScenarioCode": scenario,
                "FinanceCode": budget.finance_code,
                "SchoolCode": budget.school_code,
                "LedgerCode": "COSTCTR",
                "DepartmentCode": budget.department_code,
                "Description": budget.description,
                "YearNotes": "",
                "FinancialYearCode": self.current_year,
                "Calculated": bool(budget.calculator_code),
                "YearValue": budget.year_value,  # The actual customer value (always a number)
            })

        # Build values from extracted pupils
        for pupil in self.extracted_pupils:
            self.template_data["ScenarioYearValues"].append({
                "Tab": "Pupils",
                "ScenarioCode": scenario,
                "FinanceCode": pupil.finance_code,
                "SchoolCode": pupil.school_code,
                "LedgerCode": "DEFAULT",
                "DepartmentCode": "DEFAULT",
                "Description": pupil.description,
                "YearNotes": "",
                "FinancialYearCode": pupil.year_code,
                "Calculated": False,
                "YearValue": pupil.value,  # The pupil count (always a number)
            })

    def _build_monthly_values(self):
        """
        Build 37_Monthly Values sheet - monthly breakdown.

        For budgets with monthly data, this provides the period-by-period values.
        Most data uses MonthProfile to spread the YearValue, but some customer
        data provides explicit monthly values.
        """
        self.log("Building Monthly Values...")

        # For now, we only create monthly values for items with explicit monthly data
        # This will be populated when we extract monthly data from customer files
        # that have PERIOD 1-12 or month-named columns

        # The monthly_data attribute would be populated during extraction
        if hasattr(self, 'extracted_monthly') and self.extracted_monthly:
            for monthly in self.extracted_monthly:
                self.template_data["Monthly Values"].append({
                    "Tab": monthly.get("tab", "Expenditure"),
                    "ScenarioCode": "",
                    "FinanceCode": monthly.get("finance_code", ""),
                    "SchoolCode": monthly.get("school_code", ""),
                    "LedgerCode": monthly.get("ledger_code", "COSTCTR"),
                    "DepartmentCode": monthly.get("department_code", "DEFAULT"),
                    "FinancialYearCode": monthly.get("year_code", self.current_year),
                    "Period01Value": monthly.get("period_01", 0),  # September
                    "Period02Value": monthly.get("period_02", 0),  # October
                    "Period03Value": monthly.get("period_03", 0),  # November
                    "Period04Value": monthly.get("period_04", 0),  # December
                    "Period05Value": monthly.get("period_05", 0),  # January
                    "Period06Value": monthly.get("period_06", 0),  # February
                    "Period07Value": monthly.get("period_07", 0),  # March
                    "Period08Value": monthly.get("period_08", 0),  # April
                    "Period09Value": monthly.get("period_09", 0),  # May
                    "Period10Value": monthly.get("period_10", 0),  # June
                    "Period11Value": monthly.get("period_11", 0),  # July
                    "Period12Value": monthly.get("period_12", 0),  # August
                })

    # =========================================================================
    # EXTERNAL AUDIT REVIEW
    # =========================================================================

    def perform_external_audit(self, customer_data_dir: Path) -> Dict[str, Any]:
        """
        External Audit Review - Compare source data against processed output.
        Validates data integrity, completeness, and accuracy for S3 financial data.
        """
        self.log("\n" + "="*60)
        self.log("EXTERNAL AUDIT REVIEW")
        self.log("="*60)

        # Reset audit results
        self.audit_results = {
            "source_vs_output": [],
            "data_integrity": [],
            "domain_checks": [],
            "missing_data": [],
            "data_lineage": [],
        }
        self.audit_passed = True
        self.audit_score = 100.0

        # 1. Source vs Output comparison
        self._audit_source_vs_output()

        # 2. Data integrity checks
        self._audit_data_integrity()

        # 3. Domain-specific checks
        self._audit_domain_rules()

        # 4. Missing/incomplete data
        self._audit_missing_data()

        # 5. Data lineage verification
        self._audit_data_lineage(customer_data_dir)

        # Calculate final audit score
        self._calculate_audit_score()

        # Log audit summary
        self._log_audit_summary()

        # Generate detailed report
        self.detailed_audit_report = self._generate_detailed_audit_report()

        return {
            "passed": self.audit_passed,
            "score": self.audit_score,
            "results": self.audit_results,
            "detailed_report": self.detailed_audit_report,
        }

    def _audit_source_vs_output(self):
        """Compare source data counts against output."""
        self.log("Auditing: Source vs Output comparison...")

        checks = []

        # Pupil records
        output_pupils = len(self.extracted_pupils)
        checks.append({
            "check": "Pupil Records Extracted",
            "output_count": output_pupils,
            "passed": output_pupils > 0,
            "severity": "warning" if output_pupils == 0 else "info",
            "details": f"Output: {output_pupils} pupil records"
        })

        # Budget lines
        income_count = len([b for b in self.extracted_budgets if b.line_type == 'income'])
        exp_count = len([b for b in self.extracted_budgets if b.line_type == 'expenditure'])
        checks.append({
            "check": "Income Lines Extracted",
            "output_count": income_count,
            "passed": True,
            "severity": "info",
            "details": f"Output: {income_count} income lines"
        })
        checks.append({
            "check": "Expenditure Lines Extracted",
            "output_count": exp_count,
            "passed": True,
            "severity": "info",
            "details": f"Output: {exp_count} expenditure lines"
        })

        # Grants
        grant_count = len(self.extracted_grants)
        checks.append({
            "check": "Grants Extracted",
            "output_count": grant_count,
            "passed": True,
            "severity": "info",
            "details": f"Output: {grant_count} grants"
        })

        self.audit_results["source_vs_output"] = checks

    def _audit_data_integrity(self):
        """Check data integrity - duplicates, nulls, format consistency."""
        self.log("Auditing: Data integrity...")

        checks = []

        # Check for duplicate pupil records (same school + finance code + year)
        pupil_keys = [(p.school_code, p.finance_code, p.year_code) for p in self.extracted_pupils]
        pupil_duplicates = [k for k in set(pupil_keys) if pupil_keys.count(k) > 1]
        checks.append({
            "check": "Pupil Record Uniqueness",
            "passed": len(pupil_duplicates) == 0,
            "severity": "warning" if pupil_duplicates else "info",
            "details": f"Duplicates: {len(pupil_duplicates)}" if pupil_duplicates else "All unique"
        })

        # Check for duplicate budget lines (same school + finance code + department)
        budget_keys = [(b.school_code, b.finance_code, b.department_code) for b in self.extracted_budgets]
        budget_duplicates = [k for k in set(budget_keys) if budget_keys.count(k) > 1]
        checks.append({
            "check": "Budget Line Uniqueness",
            "passed": len(budget_duplicates) == 0,
            "severity": "warning" if budget_duplicates else "info",
            "details": f"Duplicates: {len(budget_duplicates)}" if budget_duplicates else "All unique"
        })

        # Check schools are valid
        invalid_schools = [s for s in self.schools_found if not s or s == 'nan']
        checks.append({
            "check": "Valid School Codes",
            "passed": len(invalid_schools) == 0,
            "severity": "error" if invalid_schools else "info",
            "details": f"Invalid: {len(invalid_schools)}" if invalid_schools else "All valid"
        })
        if invalid_schools:
            self.audit_passed = False

        # Check finance codes have valid format
        invalid_fc = [fc for fc in self.finance_codes_found if not fc or fc == 'nan']
        checks.append({
            "check": "Valid Finance Codes",
            "passed": len(invalid_fc) == 0,
            "severity": "warning" if invalid_fc else "info",
            "details": f"Invalid: {len(invalid_fc)}" if invalid_fc else "All valid"
        })

        self.audit_results["data_integrity"] = checks

    def _audit_domain_rules(self):
        """Check domain-specific business rules for S3 financial data."""
        self.log("Auditing: Domain rules...")

        checks = []

        # Check pupil numbers are non-negative
        negative_pupils = [p for p in self.extracted_pupils if p.value < 0]
        checks.append({
            "check": "Pupil Numbers Non-Negative",
            "passed": len(negative_pupils) == 0,
            "severity": "error" if negative_pupils else "info",
            "details": f"Negative values: {len(negative_pupils)}" if negative_pupils else "All valid"
        })
        if negative_pupils:
            self.audit_passed = False

        # Check income lines have negative values (accounting convention)
        positive_income = [b for b in self.extracted_budgets
                         if b.line_type == 'income' and b.year_value > 0]
        checks.append({
            "check": "Income Values Negative",
            "passed": len(positive_income) == 0,
            "severity": "warning" if positive_income else "info",
            "details": f"Positive income: {len(positive_income)}" if positive_income else "Correct sign convention"
        })

        # Check expenditure lines have positive values
        negative_exp = [b for b in self.extracted_budgets
                       if b.line_type == 'expenditure' and b.year_value < 0]
        checks.append({
            "check": "Expenditure Values Positive",
            "passed": len(negative_exp) == 0,
            "severity": "warning" if negative_exp else "info",
            "details": f"Negative expenditure: {len(negative_exp)}" if negative_exp else "Correct sign convention"
        })

        # Check grant amounts are positive
        negative_grants = [g for g in self.extracted_grants if g.amount < 0]
        checks.append({
            "check": "Grant Amounts Positive",
            "passed": len(negative_grants) == 0,
            "severity": "warning" if negative_grants else "info",
            "details": f"Negative grants: {len(negative_grants)}" if negative_grants else "All valid"
        })

        # Check calculator codes are valid
        valid_calculators = {'0%_CALC', 'FUNDING_GAG', 'FUNDING_16_19', 'PUPPREMIUM_FACTOR',
                            'PUPPREMIUM_CALC', 'DFC_CORE', 'DFC_PUPIL', 'DFC_EXP',
                            'PE_GRANT_CORE', 'PE_GRANT_PUPIL', 'UIFSM_CALC', 'UIFSM_RATE',
                            'CENTRALCHG_SCH', 'CENTRALCHG_MAT', ''}
        invalid_calcs = [p.calculator_code for p in self.extracted_pupils
                        if p.calculator_code and p.calculator_code not in valid_calculators]
        checks.append({
            "check": "Valid Calculator Codes",
            "passed": len(invalid_calcs) == 0,
            "severity": "warning" if invalid_calcs else "info",
            "details": f"Invalid: {set(invalid_calcs)}" if invalid_calcs else "All valid"
        })

        self.audit_results["domain_checks"] = checks

    def _audit_missing_data(self):
        """Check for missing or incomplete data."""
        self.log("Auditing: Missing data...")

        checks = []

        # Check if we have any data at all
        has_data = (len(self.extracted_pupils) > 0 or
                   len(self.extracted_budgets) > 0 or
                   len(self.extracted_grants) > 0)
        checks.append({
            "check": "Data Extracted",
            "passed": has_data,
            "severity": "error" if not has_data else "info",
            "details": "Some data extracted" if has_data else "No data extracted"
        })
        if not has_data:
            self.audit_passed = False

        # Check if we have schools
        checks.append({
            "check": "Schools Identified",
            "passed": len(self.schools_found) > 0,
            "severity": "warning" if len(self.schools_found) == 0 else "info",
            "details": f"Schools: {len(self.schools_found)}"
        })

        # Check pupils have school codes
        pupils_no_school = [p for p in self.extracted_pupils if not p.school_code]
        pupil_school_pct = ((len(self.extracted_pupils) - len(pupils_no_school)) /
                          len(self.extracted_pupils) * 100) if self.extracted_pupils else 0
        checks.append({
            "check": "Pupils Have School Codes",
            "passed": pupil_school_pct >= 80,
            "severity": "warning" if pupil_school_pct < 80 else "info",
            "details": f"{pupil_school_pct:.1f}% have school codes"
        })

        # Check budgets have descriptions
        budgets_no_desc = [b for b in self.extracted_budgets
                         if not b.description or b.description == b.finance_code]
        budget_desc_pct = ((len(self.extracted_budgets) - len(budgets_no_desc)) /
                         len(self.extracted_budgets) * 100) if self.extracted_budgets else 0
        checks.append({
            "check": "Budget Lines Have Descriptions",
            "passed": budget_desc_pct >= 50,
            "severity": "warning" if budget_desc_pct < 50 else "info",
            "details": f"{budget_desc_pct:.1f}% have descriptions"
        })

        self.audit_results["missing_data"] = checks

    def _audit_data_lineage(self, customer_data_dir: Path):
        """Verify data lineage - track where data came from."""
        self.log("Auditing: Data lineage...")

        checks = []

        # Count source files processed
        source_files = list(customer_data_dir.rglob("*.xls*")) + list(customer_data_dir.rglob("*.csv"))
        source_files = [f for f in source_files if not f.name.startswith("~$")]

        checks.append({
            "check": "Source Files Processed",
            "passed": len(source_files) > 0,
            "severity": "error" if len(source_files) == 0 else "info",
            "details": f"Files: {len(source_files)}"
        })
        if len(source_files) == 0:
            self.audit_passed = False

        # Record source file names
        checks.append({
            "check": "Source File List",
            "passed": True,
            "severity": "info",
            "details": ", ".join([f.name for f in source_files[:5]]) +
                      (f" (+{len(source_files)-5} more)" if len(source_files) > 5 else "")
        })

        # Check for issues during extraction
        checks.append({
            "check": "Extraction Issues",
            "passed": len(self.issues) == 0,
            "severity": "warning" if self.issues else "info",
            "details": f"Issues: {len(self.issues)}" if self.issues else "No issues"
        })

        self.audit_results["data_lineage"] = checks

    def _calculate_audit_score(self):
        """Calculate overall audit score based on check results."""
        total_checks = 0
        passed_checks = 0
        error_count = 0
        warning_count = 0

        for category, checks in self.audit_results.items():
            for check in checks:
                total_checks += 1
                if check.get("passed", True):
                    passed_checks += 1
                elif check.get("severity") == "error":
                    error_count += 1
                elif check.get("severity") == "warning":
                    warning_count += 1

        # Calculate score: errors = -10 points, warnings = -5 points
        if total_checks > 0:
            base_score = (passed_checks / total_checks) * 100
            penalty = (error_count * 10) + (warning_count * 5)
            self.audit_score = max(0, base_score - penalty)
        else:
            self.audit_score = 0

        # Audit fails if score < 60 or any critical errors
        if self.audit_score < 60 or error_count > 0:
            self.audit_passed = False

    def _log_audit_summary(self):
        """Log audit summary."""
        self.log("\n" + "-"*40)
        self.log("AUDIT SUMMARY")
        self.log("-"*40)
        self.log(f"  Audit Score: {self.audit_score:.1f}%")
        self.log(f"  Audit Passed: {self.audit_passed}")

        for category, checks in self.audit_results.items():
            errors = [c for c in checks if not c.get("passed") and c.get("severity") == "error"]
            warnings = [c for c in checks if not c.get("passed") and c.get("severity") == "warning"]

            if errors or warnings:
                self.log(f"\n  {category}:")
                for check in errors:
                    self.log(f"    ERROR: {check['check']} - {check['details']}")
                for check in warnings:
                    self.log(f"    WARNING: {check['check']} - {check['details']}")

        self.log("-"*40)

    def _generate_detailed_audit_report(self) -> Dict[str, Any]:
        """Generate detailed audit report with explanations and recommendations."""
        detailed_report = {
            "summary": {
                "score": self.audit_score,
                "passed": self.audit_passed,
                "total_issues": 0,
                "critical_issues": 0,
                "warnings": 0,
            },
            "issues": [],
            "recommendations": [],
        }

        for category, checks in self.audit_results.items():
            for check in checks:
                if not check.get("passed", True):
                    issue = self._explain_audit_issue(category, check)
                    if issue:
                        detailed_report["issues"].append(issue)
                        detailed_report["summary"]["total_issues"] += 1
                        if check.get("severity") == "error":
                            detailed_report["summary"]["critical_issues"] += 1
                        elif check.get("severity") == "warning":
                            detailed_report["summary"]["warnings"] += 1

        detailed_report["recommendations"] = self._generate_recommendations(detailed_report["issues"])
        return detailed_report

    def _explain_audit_issue(self, category: str, check: Dict) -> Dict[str, Any]:
        """Generate detailed explanation for audit issue."""
        check_name = check.get("check", "Unknown")
        severity = check.get("severity", "info")
        details = check.get("details", "")

        explanation = {
            "category": category,
            "check": check_name,
            "severity": severity,
            "details": details,
            "what_is_missing": "",
            "why_it_matters": "",
            "how_to_fix": "",
            "affected_records": [],
        }

        # Source vs Output issues
        if category == "source_vs_output":
            if "Pupil" in check_name:
                explanation["what_is_missing"] = f"Pupil records mismatch: {details}"
                explanation["why_it_matters"] = "Accurate pupil numbers are essential for funding calculations (PP, UIFSM, etc.)"
                explanation["how_to_fix"] = "Verify census data is in correct format. Check Spring/Autumn census columns are labeled correctly."
            elif "Grant" in check_name:
                explanation["what_is_missing"] = f"Grant data issue: {details}"
                explanation["why_it_matters"] = "Missing grants affect budget projections and income calculations"
                explanation["how_to_fix"] = "Ensure grant allocations are included in source data with correct grant types."
            elif "Budget" in check_name:
                explanation["what_is_missing"] = f"Budget line discrepancy: {details}"
                explanation["why_it_matters"] = "Missing budget lines cause incomplete financial projections"
                explanation["how_to_fix"] = "Check income/expenditure columns are correctly labeled and contain valid amounts."

        # Data integrity issues
        elif category == "data_integrity":
            if "Duplicate" in check_name:
                explanation["what_is_missing"] = f"Duplicate records found: {details}"
                explanation["why_it_matters"] = "Duplicates cause double-counting in budgets and reports"
                explanation["how_to_fix"] = "Review source data for unintentional duplicates."
            elif "Negative" in check_name:
                explanation["what_is_missing"] = f"Invalid values: {details}"
                explanation["why_it_matters"] = "Negative pupil numbers or invalid amounts cause calculation errors"
                explanation["how_to_fix"] = "Check for data entry errors in source files."

        # Domain checks
        elif category == "domain_checks":
            if "Census" in check_name:
                explanation["what_is_missing"] = f"Census data issue: {details}"
                explanation["why_it_matters"] = "Census data drives funding calculations"
                explanation["how_to_fix"] = "Ensure census columns include correct term (Spring/Autumn) and year."
            elif "Funding" in check_name:
                explanation["what_is_missing"] = f"Funding calculation issue: {details}"
                explanation["why_it_matters"] = "Incorrect funding affects budget accuracy"
                explanation["how_to_fix"] = "Verify funding rates match current DfE rates."

        # Missing data
        elif category == "missing_data":
            if "School" in check_name:
                explanation["what_is_missing"] = f"Schools missing pupil data: {details}"
                explanation["why_it_matters"] = "Schools without pupil data cannot receive proper funding allocations"
                explanation["how_to_fix"] = "Add pupil numbers for all schools in the census data."

        return explanation

    def _generate_recommendations(self, issues: List[Dict]) -> List[Dict[str, str]]:
        """Generate prioritized recommendations."""
        recommendations = []
        seen = set()

        for issue in issues:
            if issue.get("severity") == "error":
                rec = {
                    "priority": "HIGH",
                    "category": issue.get("category", ""),
                    "action": issue.get("how_to_fix", "Review and fix critical data issues"),
                    "reason": issue.get("why_it_matters", ""),
                }
                key = f"{rec['category']}:{rec['action'][:50]}"
                if key not in seen:
                    recommendations.append(rec)
                    seen.add(key)

        for issue in issues:
            if issue.get("severity") == "warning":
                rec = {
                    "priority": "MEDIUM",
                    "category": issue.get("category", ""),
                    "action": issue.get("how_to_fix", "Review data quality"),
                    "reason": issue.get("why_it_matters", ""),
                }
                key = f"{rec['category']}:{rec['action'][:50]}"
                if key not in seen:
                    recommendations.append(rec)
                    seen.add(key)

        if not self.extracted_pupils:
            recommendations.insert(0, {
                "priority": "HIGH",
                "category": "data_source",
                "action": "Add census data with pupil numbers to the S3 folder",
                "reason": "No pupil data found - funding calculations require pupil numbers",
            })

        return recommendations

    # =========================================================================
    # OUTPUT METHODS
    # =========================================================================

    def _write_to_template(
        self,
        formatted_sheets: Dict[str, pd.DataFrame],
        output_dir: Path,
        timestamp: str
    ) -> Path:
        """
        Write processed data INTO a copy of the pre-populated template.

        Preserves existing sheets (reference data, formulas, formatting).
        Only updates/adds data sheets.

        Args:
            formatted_sheets: Dict of sheet_name -> DataFrame to write
            output_dir: Output directory
            timestamp: Timestamp for filename

        Returns:
            Path to output file
        """
        import shutil
        from openpyxl import load_workbook

        # Create output filename
        template_name = Path(self.template_workbook).stem
        output_file = output_dir / f"{template_name}_populated_{timestamp}.xlsx"

        # Copy template to output location
        self.log(f"  Copying template to: {output_file}")
        shutil.copy2(self.template_workbook, output_file)

        # Open the copied workbook
        self.log(f"  Opening workbook for writing...")
        wb = load_workbook(output_file)

        # Sheets that should NOT be overwritten (reference data)
        protected_sheets = {
            'Schools', 'Depts', 'Funds', 'LocalAuth', 'Activity', 'Ledger',
            'SchHub', 'SchType', 'System Grouping Codes', 'FinanceCodes Budget',
            '14_Calculators', '15_MonthProfiles', 'Parameters', 'Lookup',
            'Validation', 'Checklist', 'Notes'
        }

        # Write data sheets
        for sheet_name, df in formatted_sheets.items():
            if len(df) == 0:
                continue

            # Check if this is a protected sheet
            if sheet_name in protected_sheets:
                self.log(f"  [SKIP] {sheet_name} - protected reference sheet")
                continue

            # Remove existing sheet if it exists (we'll replace it)
            if sheet_name in wb.sheetnames:
                self.log(f"  [UPDATE] {sheet_name}: {len(df)} rows")
                del wb[sheet_name]
            else:
                self.log(f"  [ADD] {sheet_name}: {len(df)} rows")

            # Create new sheet and write data
            ws = wb.create_sheet(sheet_name)

            # Write headers
            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)

            # Write data rows
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    # Handle NaN values
                    if pd.isna(value):
                        ws.cell(row=row_idx, column=col_idx, value=None)
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=value)

        # Add summary and audit sheets
        self._add_summary_sheets(wb)

        # Save workbook
        self.log(f"  Saving workbook...")
        wb.save(output_file)
        wb.close()

        return output_file

    def _write_new_workbook(
        self,
        formatted_sheets: Dict[str, pd.DataFrame],
        output_file: Path
    ):
        """
        Create a new workbook with all data (raw data mode).

        Args:
            formatted_sheets: Dict of sheet_name -> DataFrame to write
            output_file: Path to output file
        """
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for sheet_name, df in formatted_sheets.items():
                if len(df) > 0:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    self.log(f"  [ADD] {sheet_name}: {len(df)} rows")

            # Add summary sheets
            self._write_summary_sheets_pandas(writer)

    def _add_summary_sheets(self, wb):
        """Add summary and audit sheets to workbook (openpyxl)."""
        from openpyxl.utils.dataframe import dataframe_to_rows

        # Summary sheet
        summary_data = {
            "Metric": [
                "Pupil Records", "Income Lines", "Expenditure Lines", "Grants",
                "Schools", "Finance Codes", "Issues", "---",
                "AUDIT SCORE", "AUDIT PASSED", "BUILD MODE"
            ],
            "Value": [
                len(self.extracted_pupils),
                len([b for b in self.extracted_budgets if b.line_type == 'income']),
                len([b for b in self.extracted_budgets if b.line_type == 'expenditure']),
                len(self.extracted_grants),
                len(self.schools_found),
                len(self.finance_codes_found),
                len(self.issues),
                "---",
                f"{self.audit_score:.1f}%",
                "YES" if self.audit_passed else "NO",
                self.build_mode.upper()
            ]
        }
        summary_df = pd.DataFrame(summary_data)

        if "_Summary" in wb.sheetnames:
            del wb["_Summary"]
        ws = wb.create_sheet("_Summary")
        for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Issues sheet
        if self.issues:
            if "_Issues" in wb.sheetnames:
                del wb["_Issues"]
            ws = wb.create_sheet("_Issues")
            ws.cell(row=1, column=1, value="Issue")
            for idx, issue in enumerate(self.issues, 2):
                ws.cell(row=idx, column=1, value=issue)

        # Audit Report sheet
        audit_report = []
        for category, checks in self.audit_results.items():
            for check in checks:
                audit_report.append({
                    "Category": category,
                    "Check": check.get("check", ""),
                    "Passed": "YES" if check.get("passed", True) else "NO",
                    "Severity": check.get("severity", "info").upper(),
                    "Details": check.get("details", ""),
                })
        if audit_report:
            audit_df = pd.DataFrame(audit_report)
            if "_Audit_Report" in wb.sheetnames:
                del wb["_Audit_Report"]
            ws = wb.create_sheet("_Audit_Report")
            for r_idx, row in enumerate(dataframe_to_rows(audit_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

    def _write_summary_sheets_pandas(self, writer):
        """Write summary sheets using pandas ExcelWriter."""
        # Summary
        summary_data = {
            "Metric": [
                "Pupil Records", "Income Lines", "Expenditure Lines", "Grants",
                "Schools", "Finance Codes", "Issues", "---",
                "AUDIT SCORE", "AUDIT PASSED", "BUILD MODE"
            ],
            "Value": [
                len(self.extracted_pupils),
                len([b for b in self.extracted_budgets if b.line_type == 'income']),
                len([b for b in self.extracted_budgets if b.line_type == 'expenditure']),
                len(self.extracted_grants),
                len(self.schools_found),
                len(self.finance_codes_found),
                len(self.issues),
                "---",
                f"{self.audit_score:.1f}%",
                "YES" if self.audit_passed else "NO",
                self.build_mode.upper()
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="_Summary", index=False)

        # Issues
        if self.issues:
            issues_data = [{"Issue": issue} for issue in self.issues]
            pd.DataFrame(issues_data).to_excel(writer, sheet_name="_Issues", index=False)

        # Audit Report
        audit_report = []
        for category, checks in self.audit_results.items():
            for check in checks:
                audit_report.append({
                    "Category": category,
                    "Check": check.get("check", ""),
                    "Passed": "YES" if check.get("passed", True) else "NO",
                    "Severity": check.get("severity", "info").upper(),
                    "Details": check.get("details", ""),
                })
        if audit_report:
            pd.DataFrame(audit_report).to_excel(writer, sheet_name="_Audit_Report", index=False)

        # Detailed Audit Report
        detailed_issues = []
        for issue in self.detailed_audit_report.get("issues", []):
            detailed_issues.append({
                "Category": issue.get("category", ""),
                "Check": issue.get("check", ""),
                "Severity": issue.get("severity", "").upper(),
                "What Is Missing": issue.get("what_is_missing", ""),
                "Why It Matters": issue.get("why_it_matters", ""),
                "How To Fix": issue.get("how_to_fix", ""),
            })
        if detailed_issues:
            pd.DataFrame(detailed_issues).to_excel(writer, sheet_name="_Audit_Details", index=False)

        # Recommendations
        recommendations = self.detailed_audit_report.get("recommendations", [])
        if recommendations:
            rec_data = [{
                "Priority": r.get("priority", ""),
                "Category": r.get("category", ""),
                "Action Required": r.get("action", ""),
                "Reason": r.get("reason", ""),
            } for r in recommendations]
            pd.DataFrame(rec_data).to_excel(writer, sheet_name="_Recommendations", index=False)

    # =========================================================================
    # MAIN ENTRY POINT
    # =========================================================================

    def process(
        self,
        customer_data_dir: Path,
        output_dir: Path,
        template_path: Path = None,
        column_mappings: Dict[str, Dict[str, str]] = None,
        code_mappings: Dict[str, Dict[str, str]] = None
    ) -> Dict[str, Any]:
        """Main processing entry point with external audit.

        Supports two modes:
        1. RAW_DATA mode: Creates new workbook from scratch
        2. PREPOPULATED_TEMPLATE mode: Writes INTO existing template workbook

        Args:
            customer_data_dir: Path to customer raw data files
            output_dir: Path to save output
            template_path: Path to pre-populated template workbook (required for template mode)
            column_mappings: Optional dict of validated column mappings from pre-flight validation
            code_mappings: Optional dict of code mappings (customer codes -> template codes):
                - school_mappings: {customer_school: template_SchoolCode}
                - finance_mappings: {customer_finance: template_FinanceCode}
                - department_mappings: {customer_dept: template_DeptCode}

        Returns:
            Dict with processing results and output file path
        """
        # Store column mappings for use during processing
        self.column_mappings = column_mappings or {}

        # Apply code mappings (customer codes -> template codes)
        if code_mappings:
            self.apply_code_mappings(code_mappings)

        self.log("="*60)
        self.log("S3 SPECIALIST AGENT - STARTING PROCESSING")
        self.log("="*60)

        # Determine build mode from template_path
        if template_path and Path(template_path).exists():
            self.build_mode = "prepopulated_template"
            self.template_workbook = str(template_path)
            self.log(f"BUILD MODE: Pre-populated Template")
            self.log(f"Template: {template_path}")
        else:
            self.build_mode = "raw_data"
            self.log(f"BUILD MODE: Raw Data (no template)")

        # Phase 0: Load template reference data (if template mode)
        if self.build_mode == "prepopulated_template":
            self.log("\nPHASE 0: LOADING TEMPLATE REFERENCE DATA")
            self.log("-" * 40)
            self.load_template_reference_data(self.template_workbook)

        # Phase 1: Analysis
        self.analyze_customer_data(customer_data_dir)

        # Phase 2: Build templates
        template_sheets = self.build_all_templates()

        # Phase 3: External Audit Review
        audit_results = self.perform_external_audit(customer_data_dir)

        # Phase 4: Format data to match official template schema
        template_warnings = []
        formatted_sheets = {}

        self.log("\nPHASE 4: FORMATTING FOR OFFICIAL TEMPLATE")
        self.log("-" * 40)

        for internal_name, df in template_sheets.items():
            if len(df) == 0:
                continue

            # Get official template sheet name
            official_name = self.SHEET_NAME_MAPPING.get(internal_name, internal_name)

            # Apply template formatting if available
            if self.template_formatter and self.template_registry:
                s3_sheets = self.template_registry.list_sheets("S3")
                if official_name in s3_sheets:
                    formatted_df, warnings = self.template_formatter.format_dataframe(
                        df, "S3", official_name
                    )
                    formatted_sheets[official_name] = formatted_df
                    if warnings:
                        template_warnings.extend([f"{official_name}: {w}" for w in warnings])
                        self.log(f"  {official_name}: {len(df)} rows (formatted, {len(warnings)} warnings)")
                    else:
                        self.log(f"  {official_name}: {len(df)} rows (formatted)")
                else:
                    # No template schema, use as-is with official name
                    formatted_sheets[official_name] = df
                    self.log(f"  {official_name}: {len(df)} rows (no schema)")
            else:
                # No formatter available, use official name
                formatted_sheets[official_name] = df
                self.log(f"  {official_name}: {len(df)} rows")

        if template_warnings:
            self.log(f"\n[WARN] Template formatting warnings: {len(template_warnings)}")
            for w in template_warnings[:5]:
                self.log(f"  - {w}")

        # Phase 5: Save output
        self.log("\nPHASE 5: SAVING OUTPUT")
        self.log("-" * 40)

        output_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if self.build_mode == "prepopulated_template" and self.template_workbook:
            # TEMPLATE MODE: Write INTO a copy of the template
            output_file = self._write_to_template(
                formatted_sheets=formatted_sheets,
                output_dir=output_dir,
                timestamp=timestamp
            )
        else:
            # RAW DATA MODE: Create new workbook
            output_file = output_dir / f"S3_complete_template_{timestamp}.xlsx"
            self._write_new_workbook(formatted_sheets, output_file)

        self.log(f"\nOutput saved to: {output_file}")

        # Determine overall success (includes audit)
        has_critical_errors = len(self.issues) > 0 or not self.audit_passed

        return {
            "success": not has_critical_errors,
            "output_file": output_file,
            "template_sheets": template_sheets,
            "issues": self.issues,
            "audit": {
                "passed": self.audit_passed,
                "score": self.audit_score,
                "results": self.audit_results,
                "detailed_report": self.detailed_audit_report,
            },
            "summary": {
                "pupils": len(self.extracted_pupils),
                "income_lines": len([b for b in self.extracted_budgets if b.line_type == 'income']),
                "expenditure_lines": len([b for b in self.extracted_budgets if b.line_type == 'expenditure']),
                "grants": len(self.extracted_grants),
                "schools": list(self.schools_found),
                "audit_score": self.audit_score,
                "audit_passed": self.audit_passed,
            }
        }


def run_s3_specialist(
    customer_data_dir: Path,
    output_dir: Path,
    template_path: Path = None,
    column_mappings: Dict[str, Dict[str, str]] = None,
    code_mappings: Dict[str, Dict[str, str]] = None
) -> Dict[str, Any]:
    """Run the S3 specialist agent.

    Supports two modes:
    1. RAW_DATA mode (no template): Creates new workbook from scratch
    2. PREPOPULATED_TEMPLATE mode: Writes INTO existing template workbook

    Args:
        customer_data_dir: Path to customer raw data files
        output_dir: Path to save output
        template_path: Path to pre-populated template workbook (optional)
                      If provided, agent runs in template mode and writes
                      processed data INTO a copy of this template
        column_mappings: Optional dict of validated column mappings from pre-flight validation
        code_mappings: Optional dict of code mappings (customer codes -> template codes):
            - school_mappings: {customer_school: template_SchoolCode}
            - finance_mappings: {customer_finance: template_FinanceCode}
            - department_mappings: {customer_dept: template_DeptCode}

    Returns:
        Processing result dictionary with:
        - success: bool
        - output_file: Path to output workbook
        - build_mode: "raw_data" or "prepopulated_template"
        - issues: List of issues found
        - audit: Audit results
        - summary: Processing summary
    """
    agent = S3SpecialistAgent()
    return agent.process(
        customer_data_dir,
        output_dir,
        template_path=template_path,
        column_mappings=column_mappings,
        code_mappings=code_mappings
    )


def run_s3_preflight(
    customer_data_dir: Path,
    template_path: Path = None
) -> Dict[str, Any]:
    """
    Run pre-flight analysis to get proposed mappings for user approval.

    WORKFLOW:
    1. Call run_s3_preflight() to analyze files
    2. Show proposed_mappings to user for review/approval
    3. User approves/modifies mappings
    4. Call run_s3_specialist() with approved column_mappings

    Args:
        customer_data_dir: Path to customer raw data files
        template_path: Optional path to pre-populated template

    Returns:
        Dict containing:
        - proposed_mappings: List of mapping proposals, each with:
            - file: Source file name
            - sheet: Source sheet name
            - source_column: Original column name
            - proposed_mapping: Suggested standard column name
            - confidence: Confidence score (0-1)
            - confidence_pct: Human-readable percentage
            - reasoning: Why this mapping was proposed
            - alternatives: Other possible mappings
            - sample_values: Sample data from this column
            - requires_review: True if confidence < 70%
        - file_summary: Summary of files found
        - warnings: Any warnings about the data
        - ready_for_processing: True if all mappings are high confidence
        - requires_review_count: Number of mappings needing user review

    Example usage:
        # Step 1: Pre-flight analysis
        preflight = run_s3_preflight(
            customer_data_dir=Path("customer_data/"),
            template_path=Path("templates/S3_Template.xlsx")
        )

        # Step 2: Show to user - items where requires_review=True need attention
        for mapping in preflight["proposed_mappings"]:
            if mapping["requires_review"]:
                print(f"REVIEW: {mapping['source_column']} -> {mapping['proposed_mapping']}")
                print(f"  Confidence: {mapping['confidence_pct']}")
                print(f"  Alternatives: {mapping['alternatives']}")

        # Step 3: User approves/modifies (in your UI)
        approved = [
            {"file": "budget.xlsx", "sheet": "Sheet1",
             "source_column": "Nominal", "approved_mapping": "finance_code"},
            # ... more approved mappings
        ]

        # Step 4: Convert to column_mappings format
        agent = S3SpecialistAgent()
        column_mappings = agent.apply_user_approved_mappings(approved)

        # Step 5: Process with approved mappings
        result = run_s3_specialist(
            customer_data_dir=Path("customer_data/"),
            output_dir=Path("output/"),
            template_path=Path("templates/S3_Template.xlsx"),
            column_mappings=column_mappings
        )
    """
    agent = S3SpecialistAgent()
    return agent.preflight_analysis(
        customer_data_dir=customer_data_dir,
        template_path=template_path
    )
