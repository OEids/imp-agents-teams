"""
Process S2 Customer Data into Template
Reads customer data and writes to S2 template tabs
"""
import pandas as pd
from pathlib import Path
from datetime import datetime
import shutil
import warnings
warnings.filterwarnings('ignore')

# Paths
CUSTOMER_DATA = Path(r'C:\claude\customer data\S2\Planner - Staffing.xlsm')
TEMPLATE_PATH = Path(r'C:\claude\agent-teams\knowledge\Templates\Strand 2\AA_New - Strand 2 Standard Workbook API 1.101.xlsx')
OUTPUT_PATH = Path(r'C:\claude\agent-teams\reports\S2_Customer_Output.xlsx')

def clean_date(val):
    """Convert date to string format."""
    if pd.isna(val):
        return ''
    if isinstance(val, (datetime, pd.Timestamp)):
        return val.strftime('%Y-%m-%d')
    return str(val)

def clean_val(val):
    """Clean value for output."""
    if pd.isna(val):
        return ''
    if isinstance(val, (datetime, pd.Timestamp)):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, float) and val != val:  # NaN
        return ''
    return val

def read_customer_data():
    """Read and parse customer staffing data."""
    print(f"Reading customer data from: {CUSTOMER_DATA}")
    df = pd.read_excel(CUSTOMER_DATA, sheet_name='Staff Contract Information', header=4)

    # Column mapping
    col_map = {
        'Unique payroll number / person identifier': 'payroll_number',
        'Last Name': 'last_name',
        'First Name': 'first_name',
        'Continuous Service start (dd/mm/yyyy)': 'service_start',
        'Work location (school name or code)': 'school',
        'Contract reference (unique to individual) (if applicable)': 'contract_ref',
        'Staff Role or Job Title name': 'job_title',
        'Gross Salary Finance Code / Nominal': 'finance_code',
        'Department Code / cost centre': 'department',
        'Fund Code (if applicable)': 'fund_code',
        'Full time hours for role': 'ft_hours',
        'Weekly Hours': 'weekly_hours',
        'Weekly FTE': 'fte',
        'EQWP/ TTO Weeks worked per year': 'weeks_worked',
        'EQWP/ TTO Weeks paid per year': 'weeks_paid',
        'Weeks Paid Rate Code': 'weeks_paid_code',
        'Pay Scale name': 'pay_scale',
        'Current Scale Point': 'scale_point',
        'Grade title / max scale point': 'grade',
        'FTE Annual salary rate': 'fte_salary',
        'Actual Annual salary': 'actual_salary',
        'Scheme TPS/ LGPS/ Other': 'pension_scheme',
        'Pension Opt Out': 'pension_opt_out',
        'Gender': 'gender',
        'Contract type (perm, fixed term etc)': 'contract_type',
        'Contract Start (if after budget period start date)': 'contract_start',
        'Contract End (if in budget period)': 'contract_end',
    }

    df = df.rename(columns=col_map)

    # Remove invalid rows
    df = df[df['payroll_number'].notna()]
    df = df[~df['payroll_number'].astype(str).str.contains('Example|nan', case=False, na=False)]

    print(f"  Found {len(df)} contract records")
    return df

def extract_staff_members(df):
    """Extract unique staff members."""
    staff = []
    seen = set()

    for _, row in df.iterrows():
        code = str(row['payroll_number']).strip()
        if code and code not in seen and code != 'nan':
            seen.add(code)

            # Get school code from location
            school = str(row.get('school', '')).strip()
            school_code = school[:3].upper() if school else 'MAT'

            staff.append({
                'StaffMemberCode': code,
                'FirstName': clean_val(row.get('first_name')),
                'LastName': clean_val(row.get('last_name')),
                'Title': f"{clean_val(row.get('first_name'))} {clean_val(row.get('last_name'))}".strip(),
                'ServiceStartDate': clean_date(row.get('service_start')),
                'ServiceEndDate': '',
                'DateOfBirth': '',
                'Apprenticeship': False,
                'PensionOptOut': bool(row.get('pension_opt_out')) if pd.notna(row.get('pension_opt_out')) else False,
                'AvailableToAllSchools': False,
                'SchoolCodes': school_code,
                'StaffMemberEnabled': True,
                'PayrollLocation': '',
                'GenderCode': clean_val(row.get('gender')) or '',
                'Casual': False
            })

    print(f"  Extracted {len(staff)} unique staff members")
    return staff

def extract_contracts(df):
    """Extract teaching and support contracts."""
    teaching = []
    support = []

    for _, row in df.iterrows():
        code = str(row['payroll_number']).strip()
        if not code or code == 'nan':
            continue

        # Determine school code
        school = str(row.get('school', '')).strip()
        school_code = school[:3].upper() if school else 'MAT'

        # Determine if teaching or support
        pay_scale = str(row.get('pay_scale', '')).upper()
        job_title = str(row.get('job_title', '')).lower()
        pension = str(row.get('pension_scheme', '')).upper()

        is_teaching = (
            'TEACH' in pay_scale or
            'teacher' in job_title or
            'head' in job_title or
            pension == 'TPS'
        )

        # Map pension scheme
        pension_code = ''
        if 'TPS' in pension:
            pension_code = 'TPS'
        elif 'LGPS' in pension:
            pension_code = 'LGPS'

        # Map pay scale
        ps = clean_val(row.get('pay_scale'))
        if ps:
            ps = ps.upper().replace(' ', '_')

        contract = {
            'SchoolCode': school_code,
            'StaffMemberCode': code,
            'Reference': clean_val(row.get('contract_ref')) or f"{code}A",
            'Title': f"{clean_val(row.get('first_name'))} {clean_val(row.get('last_name'))}".strip(),
            'StaffRoleCode': clean_val(row.get('job_title', '')).upper().replace(' ', '_')[:20] if row.get('job_title') else '',
            'PayScaleCode': ps,
            'PayScaleGradeCode': clean_val(row.get('grade')),
            'PayScalePointCode': str(int(row.get('scale_point'))) if pd.notna(row.get('scale_point')) else '',
            'DepartmentCode': clean_val(row.get('department')),
            'FundCode': clean_val(row.get('fund_code')) or 'GAG',
            'PensionCode': pension_code,
            'EquatedWeekPatternCode': clean_val(row.get('weeks_paid_code')) or 'AYR',
            'DateFrom': clean_date(row.get('contract_start')) or clean_date(row.get('service_start')),
            'DateTo': clean_date(row.get('contract_end')),
            'WeeklyFteOrHpw': clean_val(row.get('fte')) if is_teaching else clean_val(row.get('weekly_hours')),
            'NoIncrement': False,
            'ContractTypeCode': clean_val(row.get('contract_type', '')).upper()[:4] if row.get('contract_type') else 'PERM',
        }

        if is_teaching:
            teaching.append(contract)
        else:
            support.append(contract)

    print(f"  Extracted {len(teaching)} teaching contracts")
    print(f"  Extracted {len(support)} support contracts")
    return teaching, support

def write_to_template(staff_members, teaching_contracts, support_contracts):
    """Write data to template."""
    print(f"\nCopying template to: {OUTPUT_PATH}")
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(TEMPLATE_PATH, OUTPUT_PATH)

    # Use openpyxl to write to specific cells
    from openpyxl import load_workbook
    wb = load_workbook(OUTPUT_PATH)

    # Write StaffMembers (starting at row 2, column O which is 15)
    print("Writing StaffMembers...")
    ws = wb['25_StaffMembers']
    for i, staff in enumerate(staff_members, start=2):
        ws.cell(row=i, column=15, value=staff['StaffMemberCode'])
        ws.cell(row=i, column=16, value=staff['FirstName'])
        ws.cell(row=i, column=17, value=staff['LastName'])
        ws.cell(row=i, column=18, value=staff['Title'])
        ws.cell(row=i, column=19, value=staff['ServiceStartDate'])
        ws.cell(row=i, column=20, value=staff['ServiceEndDate'])
        ws.cell(row=i, column=21, value=staff['DateOfBirth'])
        ws.cell(row=i, column=22, value=staff['Apprenticeship'])
        ws.cell(row=i, column=23, value=staff['PensionOptOut'])
        ws.cell(row=i, column=24, value=staff['AvailableToAllSchools'])
        ws.cell(row=i, column=25, value=staff['SchoolCodes'])
        ws.cell(row=i, column=26, value=staff['StaffMemberEnabled'])
        ws.cell(row=i, column=27, value=staff['PayrollLocation'])
        ws.cell(row=i, column=28, value=staff['GenderCode'])
        ws.cell(row=i, column=29, value=staff['Casual'])

    # Write ContractsTeachFTE (starting at row 2, column Z which is 26)
    print("Writing ContractsTeachFTE...")
    ws = wb['28_ContractsTeachFTE']
    for i, contract in enumerate(teaching_contracts, start=2):
        ws.cell(row=i, column=27, value=contract['SchoolCode'])
        ws.cell(row=i, column=28, value=contract['StaffMemberCode'])
        ws.cell(row=i, column=29, value=contract['Reference'])
        ws.cell(row=i, column=30, value=contract['Title'])
        ws.cell(row=i, column=31, value=contract['StaffRoleCode'])
        ws.cell(row=i, column=34, value=contract['PayScaleCode'])
        ws.cell(row=i, column=36, value=contract['PayScaleGradeCode'])
        ws.cell(row=i, column=37, value=contract['PayScalePointCode'])
        ws.cell(row=i, column=38, value=contract['DepartmentCode'])
        ws.cell(row=i, column=39, value=contract['FundCode'])
        ws.cell(row=i, column=40, value=contract['PensionCode'])
        ws.cell(row=i, column=41, value=contract['EquatedWeekPatternCode'])
        ws.cell(row=i, column=42, value=contract['DateFrom'])
        ws.cell(row=i, column=43, value=contract['DateTo'])
        ws.cell(row=i, column=44, value=contract['WeeklyFteOrHpw'])
        ws.cell(row=i, column=46, value=contract['NoIncrement'])
        ws.cell(row=i, column=47, value=contract['ContractTypeCode'])

    # Write ContractsSupportHours (starting at row 2, column Z which is 26)
    print("Writing ContractsSupportHours...")
    ws = wb['29_ContractsSupportHours']
    for i, contract in enumerate(support_contracts, start=2):
        ws.cell(row=i, column=27, value=contract['SchoolCode'])
        ws.cell(row=i, column=28, value=contract['StaffMemberCode'])
        ws.cell(row=i, column=29, value=contract['Reference'])
        ws.cell(row=i, column=30, value=contract['Title'])
        ws.cell(row=i, column=31, value=contract['StaffRoleCode'])
        ws.cell(row=i, column=34, value=contract['PayScaleCode'])
        ws.cell(row=i, column=36, value=contract['PayScaleGradeCode'])
        ws.cell(row=i, column=37, value=contract['PayScalePointCode'])
        ws.cell(row=i, column=38, value=contract['DepartmentCode'])
        ws.cell(row=i, column=39, value=contract['FundCode'])
        ws.cell(row=i, column=40, value=contract['PensionCode'])
        ws.cell(row=i, column=41, value=contract['EquatedWeekPatternCode'])
        ws.cell(row=i, column=42, value=contract['DateFrom'])
        ws.cell(row=i, column=43, value=contract['DateTo'])
        ws.cell(row=i, column=44, value=contract['WeeklyFteOrHpw'])
        ws.cell(row=i, column=46, value=contract['NoIncrement'])
        ws.cell(row=i, column=47, value=contract['ContractTypeCode'])

    wb.save(OUTPUT_PATH)
    print(f"\nOutput saved to: {OUTPUT_PATH}")

def main():
    print("="*60)
    print("S2 CUSTOMER DATA PROCESSING")
    print("="*60)

    # Read customer data
    df = read_customer_data()

    # Extract records
    staff_members = extract_staff_members(df)
    teaching_contracts, support_contracts = extract_contracts(df)

    # Write to template
    write_to_template(staff_members, teaching_contracts, support_contracts)

    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"Staff Members: {len(staff_members)}")
    print(f"Teaching Contracts: {len(teaching_contracts)}")
    print(f"Support Contracts: {len(support_contracts)}")
    print(f"Total Contracts: {len(teaching_contracts) + len(support_contracts)}")

if __name__ == '__main__':
    main()
